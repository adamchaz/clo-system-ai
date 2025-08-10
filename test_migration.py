#!/usr/bin/env python3
"""
Simple Asset Migration Test Script
Test basic Asset model and database connectivity before full migration
"""

import sys
import json
import logging
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, List, Any, Optional
import pandas as pd
import numpy as np

# Excel processing
import openpyxl

# Database and ORM
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.exc import SQLAlchemyError

# Add backend to path
sys.path.append('backend')

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def test_database_connection():
    """Test database connection"""
    try:
        # Test with a simple SQLite database first
        database_url = "sqlite:///test_clo.db"
        engine = create_engine(database_url)
        
        with engine.connect() as conn:
            result = conn.execute(text("SELECT 1"))
            logger.info("Database connection successful")
            return True, engine
    except Exception as e:
        logger.error(f"Database connection failed: {e}")
        return False, None

def extract_sample_assets(excel_path: str, sample_size: int = 10) -> List[Dict]:
    """Extract sample assets from Excel for testing"""
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Find All Assets sheet or similar
        asset_sheet = None
        for sheet_name in workbook.sheetnames:
            if 'asset' in sheet_name.lower():
                asset_sheet = workbook[sheet_name]
                break
        
        if not asset_sheet:
            asset_sheet = workbook.worksheets[0]  # Use first sheet
        
        logger.info(f"Using sheet: {asset_sheet.title}")
        
        # Extract headers
        headers = []
        for col in range(1, asset_sheet.max_column + 1):
            header_value = asset_sheet.cell(row=1, column=col).value
            headers.append(str(header_value) if header_value else f"Column_{col}")
        
        # Extract sample data
        sample_data = []
        for row in range(2, min(sample_size + 2, asset_sheet.max_row + 1)):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = asset_sheet.cell(row=row, column=col_idx).value
                row_data[header] = cell_value
            
            # Skip completely empty rows
            if any(value not in [None, '', ' '] for value in row_data.values()):
                sample_data.append(row_data)
        
        logger.info(f"Extracted {len(sample_data)} sample assets")
        return sample_data
        
    except Exception as e:
        logger.error(f"Error extracting sample data: {e}")
        return []

def create_simple_assets_table(engine):
    """Create a simple assets table for testing"""
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS simple_assets (
                    id INTEGER PRIMARY KEY,
                    blkrock_id TEXT,
                    issue_name TEXT,
                    issuer_name TEXT,
                    par_amount REAL,
                    market_value REAL,
                    coupon REAL,
                    maturity DATE,
                    currency TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))
            conn.commit()
            logger.info("Simple assets table created")
            return True
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        return False

def insert_sample_assets(engine, sample_data: List[Dict]):
    """Insert sample assets into simple table"""
    try:
        with engine.connect() as conn:
            inserted_count = 0
            
            for asset in sample_data:
                # Basic data cleaning and insertion
                blkrock_id = asset.get('BLKRockID') or asset.get('BlkRockID') or asset.get('ID')
                issue_name = asset.get('Issue Name') or asset.get('Security Name')
                issuer_name = asset.get('Issuer Name') or asset.get('Issuer')
                
                # Skip if no ID
                if not blkrock_id:
                    continue
                
                # Clean numeric values
                par_amount = None
                if asset.get('Par Amount'):
                    try:
                        par_str = str(asset['Par Amount']).replace('$', '').replace(',', '').strip()
                        par_amount = float(par_str) if par_str else None
                    except:
                        pass
                
                market_value = None
                if asset.get('Market Value'):
                    try:
                        mv_str = str(asset['Market Value']).replace('%', '').strip()
                        market_value = float(mv_str) if mv_str else None
                    except:
                        pass
                
                coupon = None
                if asset.get('Coupon'):
                    try:
                        coupon_str = str(asset['Coupon']).replace('%', '').strip()
                        coupon = float(coupon_str) if coupon_str else None
                    except:
                        pass
                
                # Insert record
                conn.execute(text("""
                    INSERT INTO simple_assets 
                    (blkrock_id, issue_name, issuer_name, par_amount, market_value, coupon, currency)
                    VALUES (:blkrock_id, :issue_name, :issuer_name, :par_amount, :market_value, :coupon, :currency)
                """), {
                    'blkrock_id': str(blkrock_id) if blkrock_id else None,
                    'issue_name': str(issue_name) if issue_name else None,
                    'issuer_name': str(issuer_name) if issuer_name else None,
                    'par_amount': par_amount,
                    'market_value': market_value,
                    'coupon': coupon,
                    'currency': asset.get('Currency', 'USD')
                })
                inserted_count += 1
            
            conn.commit()
            logger.info(f"Inserted {inserted_count} sample assets")
            return inserted_count
            
    except Exception as e:
        logger.error(f"Error inserting sample assets: {e}")
        return 0

def validate_insertion(engine):
    """Validate the inserted data"""
    try:
        with engine.connect() as conn:
            result = conn.execute(text("SELECT COUNT(*) as count FROM simple_assets"))
            count = result.fetchone()[0]
            logger.info(f"Total assets in database: {count}")
            
            # Show sample records
            result = conn.execute(text("SELECT * FROM simple_assets LIMIT 5"))
            records = result.fetchall()
            
            logger.info("Sample records:")
            for record in records:
                logger.info(f"  ID: {record[1]}, Name: {record[2]}, Par: {record[4]}")
            
            return count > 0
            
    except Exception as e:
        logger.error(f"Error validating data: {e}")
        return False

def main():
    """Main test execution"""
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        logger.error(f"Excel file not found: {excel_file}")
        return
    
    logger.info("Starting All Assets Migration Test")
    logger.info("=" * 50)
    
    # Test database connection
    logger.info("Testing database connection...")
    db_success, engine = test_database_connection()
    if not db_success:
        return
    
    # Extract sample assets
    logger.info("Extracting sample assets from Excel...")
    sample_assets = extract_sample_assets(excel_file, 20)
    if not sample_assets:
        logger.error("No sample data extracted")
        return
    
    # Create simple table
    logger.info("Creating simple assets table...")
    if not create_simple_assets_table(engine):
        return
    
    # Insert sample assets
    logger.info("Inserting sample assets...")
    inserted_count = insert_sample_assets(engine, sample_assets)
    
    # Validate
    logger.info("Validating migration...")
    validation_success = validate_insertion(engine)
    
    # Report results
    logger.info("\n" + "=" * 50)
    logger.info("MIGRATION TEST RESULTS")
    logger.info("=" * 50)
    logger.info(f"Database Connection: {'✅ SUCCESS' if db_success else '❌ FAILED'}")
    logger.info(f"Sample Data Extraction: {'✅ SUCCESS' if sample_assets else '❌ FAILED'}")
    logger.info(f"Assets Inserted: {inserted_count}")
    logger.info(f"Validation: {'✅ SUCCESS' if validation_success else '❌ FAILED'}")
    
    if validation_success:
        logger.info("\n🎉 Migration test completed successfully!")
        logger.info("The system is ready for full All Assets migration.")
    else:
        logger.error("\n❌ Migration test failed. Please check the logs for details.")

if __name__ == "__main__":
    main()