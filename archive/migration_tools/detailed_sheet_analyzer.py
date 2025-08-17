#!/usr/bin/env python3
"""
Detailed Sheet-by-Sheet Analyzer for CLO Excel System
Analyzes each worksheet in detail to understand the business logic and data structure
"""

import openpyxl
from openpyxl import load_workbook
import pandas as pd
from pathlib import Path
import json
from typing import Dict, List, Any
import re

class DetailedSheetAnalyzer:
    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.workbook = None
        self.sheet_details = {}
        
    def analyze_all_sheets(self):
        """Analyze all sheets in detail"""
        print("=" * 80)
        print("DETAILED SHEET-BY-SHEET ANALYSIS")
        print("=" * 80)
        
        try:
            # Load with data_only=False to get formulas
            self.workbook = load_workbook(self.filepath, data_only=False, keep_vba=True)
            
            for sheet_name in self.workbook.sheetnames:
                print(f"\n{'='*60}")
                print(f"ANALYZING SHEET: {sheet_name}")
                print(f"{'='*60}")
                
                sheet = self.workbook[sheet_name]
                sheet_analysis = self._analyze_sheet_comprehensive(sheet, sheet_name)
                self.sheet_details[sheet_name] = sheet_analysis
                
                self._print_sheet_summary(sheet_analysis)
                
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
            
        return True
    
    def _analyze_sheet_comprehensive(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Comprehensive analysis of a single sheet"""
        analysis = {
            'name': sheet_name,
            'dimensions': {
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'used_range': f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
            },
            'protection': {
                'sheet_protected': sheet.protection.sheet if hasattr(sheet.protection, 'sheet') else False,
                'cells_locked': 0,
                'cells_unlocked': 0
            },
            'data_structure': {
                'headers_detected': [],
                'data_types_by_column': {},
                'has_merged_cells': len(sheet.merged_cells.ranges) > 0,
                'merged_cell_count': len(sheet.merged_cells.ranges)
            },
            'formulas': {
                'total_count': 0,
                'unique_formulas': [],
                'complex_formulas': [],
                'cell_references': [],
                'function_usage': {}
            },
            'data_validation': {
                'validation_rules': [],
                'dropdown_lists': [],
                'numeric_constraints': []
            },
            'business_indicators': {
                'clo_keywords_found': [],
                'calculation_patterns': [],
                'input_controls': [],
                'output_summaries': []
            },
            'sample_data': {
                'first_10_rows': [],
                'key_value_pairs': []
            }
        }
        
        # Analyze cell content and formulas
        self._analyze_cell_contents(sheet, analysis)
        
        # Detect data structure
        self._detect_data_structure(sheet, analysis)
        
        # Analyze formulas in detail
        self._analyze_formulas_detailed(sheet, analysis)
        
        # Look for CLO business logic indicators
        self._identify_clo_business_logic(sheet, analysis)
        
        # Extract sample data
        self._extract_sample_data(sheet, analysis)
        
        return analysis
    
    def _analyze_cell_contents(self, sheet, analysis):
        """Analyze all cell contents"""
        formula_count = 0
        locked_count = 0
        unlocked_count = 0
        
        # Limit analysis to reasonable range to avoid memory issues
        max_row = min(sheet.max_row, 500)
        max_col = min(sheet.max_column, 50)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                try:
                    cell = sheet.cell(row=row, column=col)
                    
                    # Check protection status
                    if hasattr(cell, 'protection') and cell.protection.locked:
                        locked_count += 1
                    else:
                        unlocked_count += 1
                    
                    # Analyze formulas
                    if cell.data_type == 'f' and cell.value:
                        formula_count += 1
                        
                        # Store unique formulas (limit to 50)
                        if len(analysis['formulas']['unique_formulas']) < 50:
                            formula_info = {
                                'cell': cell.coordinate,
                                'formula': cell.value,
                                'row': row,
                                'column': col
                            }
                            analysis['formulas']['unique_formulas'].append(formula_info)
                        
                        # Count function usage
                        if cell.value:
                            functions = re.findall(r'([A-Z]{2,})\s*\(', str(cell.value))
                            for func in functions:
                                analysis['formulas']['function_usage'][func] = \
                                    analysis['formulas']['function_usage'].get(func, 0) + 1
                
                except Exception as e:
                    # Skip problematic cells
                    continue
        
        analysis['formulas']['total_count'] = formula_count
        analysis['protection']['cells_locked'] = locked_count
        analysis['protection']['cells_unlocked'] = unlocked_count
    
    def _detect_data_structure(self, sheet, analysis):
        """Detect the data structure and headers"""
        headers = []
        
        # Check first few rows for headers
        for row_num in range(1, min(6, sheet.max_row + 1)):
            row_values = []
            for col in range(1, min(sheet.max_column + 1, 20)):
                try:
                    cell_value = sheet.cell(row=row_num, column=col).value
                    if cell_value is not None:
                        row_values.append(str(cell_value).strip())
                    else:
                        row_values.append("")
                except:
                    row_values.append("")
            
            # If this looks like a header row (text values, no numbers)
            if row_values and any(val and not str(val).replace('.','').replace('-','').isdigit() 
                                 for val in row_values if val):
                headers = row_values
                analysis['data_structure']['headers_detected'] = headers
                break
        
        # Analyze data types by column
        if headers:
            for col_idx, header in enumerate(headers[:10]):  # Limit to first 10 columns
                if header:
                    col_data = []
                    for row in range(2, min(sheet.max_row + 1, 20)):  # Sample first 20 rows
                        try:
                            cell_value = sheet.cell(row=row, column=col_idx + 1).value
                            if cell_value is not None:
                                col_data.append(cell_value)
                        except:
                            continue
                    
                    if col_data:
                        # Determine predominant data type
                        types = [type(val).__name__ for val in col_data]
                        most_common_type = max(set(types), key=types.count) if types else 'unknown'
                        analysis['data_structure']['data_types_by_column'][header] = {
                            'predominant_type': most_common_type,
                            'sample_values': col_data[:3]
                        }
    
    def _analyze_formulas_detailed(self, sheet, analysis):
        """Detailed formula analysis"""
        complex_formulas = []
        
        for formula_info in analysis['formulas']['unique_formulas']:
            formula = formula_info['formula']
            
            if formula and len(formula) > 50:  # Consider formulas >50 chars as complex
                complex_info = {
                    'cell': formula_info['cell'],
                    'length': len(formula),
                    'nesting_level': formula.count('('),
                    'sheet_references': len(re.findall(r"'[^']+'\!", formula)),
                    'preview': formula[:100] + "..." if len(formula) > 100 else formula
                }
                complex_formulas.append(complex_info)
        
        analysis['formulas']['complex_formulas'] = complex_formulas
        
        # Extract cell references
        all_formulas_text = " ".join([f['formula'] for f in analysis['formulas']['unique_formulas'] 
                                     if f['formula']])
        
        # Find sheet references
        sheet_refs = re.findall(r"'([^']+)'\!", all_formulas_text)
        analysis['formulas']['cell_references'] = list(set(sheet_refs))
    
    def _identify_clo_business_logic(self, sheet, analysis):
        """Identify CLO-specific business logic"""
        clo_keywords = {
            'portfolio_management': ['portfolio', 'asset', 'loan', 'security', 'position'],
            'risk_metrics': ['default', 'recovery', 'loss', 'rating', 'spread', 'correlation'],
            'compliance': ['coverage', 'ratio', 'test', 'compliance', 'overcollateralization', 'oc'],
            'cash_flow': ['cashflow', 'payment', 'interest', 'principal', 'coupon', 'yield'],
            'modeling': ['simulation', 'monte carlo', 'scenario', 'stress', 'hypothesis'],
            'valuation': ['npv', 'pv', 'fair value', 'mark', 'price', 'valuation']
        }
        
        # Check all cell values for CLO keywords
        keyword_matches = {category: [] for category in clo_keywords}
        
        # Sample cells for keyword analysis
        for row in range(1, min(sheet.max_row + 1, 100)):
            for col in range(1, min(sheet.max_column + 1, 20)):
                try:
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower()
                        
                        for category, keywords in clo_keywords.items():
                            for keyword in keywords:
                                if keyword in cell_lower:
                                    keyword_matches[category].append({
                                        'cell': sheet.cell(row=row, column=col).coordinate,
                                        'keyword': keyword,
                                        'context': cell_value[:50]
                                    })
                except:
                    continue
        
        # Store significant keyword matches
        significant_matches = []
        for category, matches in keyword_matches.items():
            if matches:
                significant_matches.append({
                    'category': category,
                    'count': len(matches),
                    'examples': matches[:3]  # First 3 examples
                })
        
        analysis['business_indicators']['clo_keywords_found'] = significant_matches
        
        # Identify calculation patterns
        calculation_patterns = []
        
        for formula_info in analysis['formulas']['unique_formulas'][:20]:  # Check first 20 formulas
            formula = formula_info['formula']
            if formula:
                formula_upper = formula.upper()
                
                if any(func in formula_upper for func in ['SUMPRODUCT', 'SUMIFS', 'SUMIF']):
                    calculation_patterns.append('Weighted calculations/aggregations')
                
                if any(func in formula_upper for func in ['VLOOKUP', 'INDEX', 'MATCH']):
                    calculation_patterns.append('Data lookup operations')
                
                if any(func in formula_upper for func in ['IF', 'IFS', 'CHOOSE']):
                    calculation_patterns.append('Conditional logic')
                
                if any(func in formula_upper for func in ['MAX', 'MIN', 'AVERAGE']):
                    calculation_patterns.append('Statistical calculations')
        
        analysis['business_indicators']['calculation_patterns'] = list(set(calculation_patterns))
    
    def _extract_sample_data(self, sheet, analysis):
        """Extract sample data for understanding"""
        sample_rows = []
        
        # Extract first 10 rows
        for row in range(1, min(sheet.max_row + 1, 11)):
            row_data = []
            for col in range(1, min(sheet.max_column + 1, 10)):
                try:
                    cell_value = sheet.cell(row=row, column=col).value
                    # Convert to string for JSON serialization
                    if cell_value is not None:
                        row_data.append(str(cell_value)[:100])  # Limit string length
                    else:
                        row_data.append(None)
                except:
                    row_data.append(None)
            sample_rows.append(row_data)
        
        analysis['sample_data']['first_10_rows'] = sample_rows
        
        # Extract key-value pairs (for configuration sheets)
        key_value_pairs = []
        for row in range(1, min(sheet.max_row + 1, 50)):
            try:
                key_cell = sheet.cell(row=row, column=1).value
                value_cell = sheet.cell(row=row, column=2).value
                
                if key_cell and isinstance(key_cell, str) and len(key_cell.strip()) > 0:
                    key_value_pairs.append({
                        'key': str(key_cell).strip(),
                        'value': str(value_cell) if value_cell is not None else None,
                        'row': row
                    })
            except:
                continue
        
        analysis['sample_data']['key_value_pairs'] = key_value_pairs[:20]  # Limit to 20
    
    def _print_sheet_summary(self, analysis):
        """Print summary of sheet analysis"""
        print(f"Sheet: {analysis['name']}")
        print(f"Dimensions: {analysis['dimensions']['max_row']} rows × {analysis['dimensions']['max_column']} columns")
        
        # Data structure
        if analysis['data_structure']['headers_detected']:
            print(f"Headers detected: {len(analysis['data_structure']['headers_detected'])}")
            print(f"Sample headers: {', '.join(analysis['data_structure']['headers_detected'][:5])}")
        
        # Formulas
        print(f"Formulas: {analysis['formulas']['total_count']} total")
        if analysis['formulas']['function_usage']:
            top_functions = sorted(analysis['formulas']['function_usage'].items(), 
                                 key=lambda x: x[1], reverse=True)[:5]
            print(f"Top functions: {', '.join([f'{func}({count})' for func, count in top_functions])}")
        
        # Business logic indicators
        if analysis['business_indicators']['clo_keywords_found']:
            print("CLO Business Logic Found:")
            for indicator in analysis['business_indicators']['clo_keywords_found']:
                print(f"  - {indicator['category']}: {indicator['count']} instances")
        
        # Calculation patterns
        if analysis['business_indicators']['calculation_patterns']:
            print(f"Calculation patterns: {', '.join(analysis['business_indicators']['calculation_patterns'])}")
        
        # Sample data insight
        if analysis['sample_data']['key_value_pairs']:
            print(f"Configuration pairs found: {len(analysis['sample_data']['key_value_pairs'])}")
            # Show a few examples
            for pair in analysis['sample_data']['key_value_pairs'][:3]:
                print(f"  {pair['key']}: {pair['value']}")
    
    def save_detailed_analysis(self, output_path: str = None):
        """Save detailed analysis to JSON"""
        if not output_path:
            output_path = self.filepath.parent / f"{self.filepath.stem}_detailed_analysis.json"
        
        # Convert to JSON-serializable format
        serializable_data = {}
        for sheet_name, sheet_data in self.sheet_details.items():
            serializable_data[sheet_name] = sheet_data
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(serializable_data, f, indent=2, default=str)
        
        print(f"\nDetailed analysis saved to: {output_path}")
        return output_path

def main():
    filepath = "TradeHypoPrelimv32.xlsm"
    
    if not Path(filepath).exists():
        print(f"Error: File '{filepath}' not found")
        return
    
    analyzer = DetailedSheetAnalyzer(filepath)
    
    if analyzer.analyze_all_sheets():
        # Save results
        output_file = analyzer.save_detailed_analysis()
        
        print("\n" + "=" * 80)
        print("SUMMARY OF CLO SYSTEM ANALYSIS")
        print("=" * 80)
        
        total_sheets = len(analyzer.sheet_details)
        total_formulas = sum(sheet['formulas']['total_count'] for sheet in analyzer.sheet_details.values())
        
        print(f"Total Sheets Analyzed: {total_sheets}")
        print(f"Total Formulas Found: {total_formulas}")
        
        # Categorize sheets by business function
        input_sheets = []
        calculation_sheets = []
        output_sheets = []
        
        for sheet_name, sheet_data in analyzer.sheet_details.items():
            name_lower = sheet_name.lower()
            if 'input' in name_lower:
                input_sheets.append(sheet_name)
            elif 'output' in name_lower:
                output_sheets.append(sheet_name)
            else:
                calculation_sheets.append(sheet_name)
        
        print(f"\nSheet Categories:")
        print(f"  Input Sheets: {len(input_sheets)} - {', '.join(input_sheets)}")
        print(f"  Calculation Sheets: {len(calculation_sheets)} - {', '.join(calculation_sheets[:5])}...")
        print(f"  Output Sheets: {len(output_sheets)} - {', '.join(output_sheets)}")
        
        print(f"\nResults saved to: {output_file}")

if __name__ == "__main__":
    main()