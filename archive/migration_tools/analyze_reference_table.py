#!/usr/bin/env python3
"""
Analyze Reference Table tab for migration planning
"""

import openpyxl
from pathlib import Path
import json
from datetime import datetime

def analyze_reference_table():
    """Analyze the Reference Table worksheet structure"""
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        print(f"Excel file not found: {excel_file}")
        return
    
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    
    # Check if Reference Table exists
    if 'Reference Table' not in workbook.sheetnames:
        print("Available worksheets:", workbook.sheetnames)
        return
    
    sheet = workbook['Reference Table']
    
    print("REFERENCE TABLE ANALYSIS")
    print("=" * 60)
    print(f"Worksheet: Reference Table")
    print(f"Dimensions: {sheet.max_column} columns × {sheet.max_row} rows")
    
    # Analyze the structure - reference tables often have multiple sections
    print(f"\nSTRUCTURE ANALYSIS:")
    
    # Look for section headers and data patterns
    sections = []
    current_section = None
    
    for row_num in range(1, min(100, sheet.max_row + 1)):  # Check first 100 rows
        # Check if this row contains section headers
        row_values = []
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row_num, column=col).value
            row_values.append(value)
        
        # Look for patterns that indicate section breaks
        non_empty_values = [v for v in row_values if v is not None and str(v).strip()]
        
        if len(non_empty_values) == 1 and str(non_empty_values[0]).strip():
            # Potential section header
            section_name = str(non_empty_values[0]).strip()
            if current_section:
                sections.append(current_section)
            current_section = {
                'name': section_name,
                'start_row': row_num,
                'headers': [],
                'data_rows': [],
                'end_row': None
            }
        elif current_section and len(non_empty_values) > 1:
            # Potential data row
            if not current_section['headers']:
                # This might be headers
                current_section['headers'] = [str(v) if v else f"Column_{i+1}" for i, v in enumerate(row_values[:10])]
                current_section['header_row'] = row_num
            else:
                # This is data
                current_section['data_rows'].append({
                    'row_num': row_num,
                    'values': row_values[:len(current_section['headers'])]
                })
    
    # Add the last section
    if current_section:
        sections.append(current_section)
    
    # Display sections found
    print(f"\nSECTIONS IDENTIFIED: {len(sections)}")
    
    for i, section in enumerate(sections, 1):
        print(f"\n{i}. {section['name']}")
        print(f"   Start Row: {section['start_row']}")
        print(f"   Headers: {len(section['headers'])} columns")
        print(f"   Data Rows: {len(section['data_rows'])}")
        
        if section['headers']:
            print(f"   Columns: {', '.join(section['headers'][:5])}{'...' if len(section['headers']) > 5 else ''}")
        
        if section['data_rows']:
            print(f"   Sample Data:")
            for data_row in section['data_rows'][:3]:
                sample_values = [str(v) if v else '' for v in data_row['values'][:3]]
                print(f"     Row {data_row['row_num']}: {', '.join(sample_values)}")
    
    # If no clear sections, analyze as single table
    if not sections:
        print("\nNo clear sections found. Analyzing as single reference table...")
        
        # Extract headers from row 1
        headers = []
        for col in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=1, column=col).value
            headers.append(str(header_value) if header_value else f"Column_{col}")
        
        # Count data rows
        data_row_count = 0
        sample_data = []
        
        for row_num in range(2, sheet.max_row + 1):
            row_values = []
            has_data = False
            
            for col in range(1, len(headers) + 1):
                value = sheet.cell(row=row_num, column=col).value
                row_values.append(value)
                if value is not None and str(value).strip():
                    has_data = True
            
            if has_data:
                data_row_count += 1
                if len(sample_data) < 5:
                    sample_data.append({
                        'row': row_num,
                        'values': row_values
                    })
        
        sections = [{
            'name': 'Reference Table',
            'start_row': 1,
            'headers': headers,
            'data_rows': data_row_count,
            'sample_data': sample_data
        }]
        
        print(f"Single table with {len(headers)} columns and {data_row_count} data rows")
    
    # Analyze data types in each section
    print(f"\nDATA TYPE ANALYSIS:")
    
    for section in sections:
        if not section.get('data_rows'):
            continue
            
        print(f"\n{section['name']}:")
        
        if isinstance(section['data_rows'], list) and section['data_rows']:
            # Analyze first few data rows
            for col_idx, header in enumerate(section['headers'][:10]):
                values = []
                for data_row in section['data_rows'][:10]:
                    if col_idx < len(data_row['values']):
                        value = data_row['values'][col_idx]
                        if value is not None:
                            values.append(value)
                
                if values:
                    # Determine data type
                    numeric_count = sum(1 for v in values if isinstance(v, (int, float)))
                    text_count = sum(1 for v in values if isinstance(v, str))
                    date_count = sum(1 for v in values if hasattr(v, 'date'))
                    
                    if numeric_count > len(values) * 0.7:
                        data_type = "Numeric"
                    elif date_count > len(values) * 0.7:
                        data_type = "Date"
                    else:
                        data_type = "Text"
                    
                    sample_vals = [str(v)[:20] for v in values[:3]]
                    print(f"   {header}: {data_type} | Sample: {', '.join(sample_vals)}")
    
    # Migration recommendations
    print(f"\nMIGRATION RECOMMENDATIONS:")
    print("-" * 40)
    
    for section in sections:
        section_name = section['name']
        
        print(f"\n{section_name}:")
        
        if 'rating' in section_name.lower() or 'ratings' in section_name.lower():
            print("  - Suggested Table: rating_mappings")
            print("  - Purpose: Credit rating standardization and conversion")
            print("  - Schema: rating_agency, rating_symbol, numeric_value, description")
        
        elif 'industry' in section_name.lower() or 'sector' in section_name.lower():
            print("  - Suggested Table: industry_classifications") 
            print("  - Purpose: Industry and sector lookup codes")
            print("  - Schema: classification_type, code, description, parent_code")
        
        elif 'country' in section_name.lower() or 'geography' in section_name.lower():
            print("  - Suggested Table: geographic_regions")
            print("  - Purpose: Country and geographic region mappings")
            print("  - Schema: country_code, country_name, region, sub_region")
        
        elif 'currency' in section_name.lower():
            print("  - Suggested Table: currencies")
            print("  - Purpose: Currency codes and conversion factors")
            print("  - Schema: currency_code, currency_name, symbol, decimal_places")
        
        elif 'index' in section_name.lower() or 'benchmark' in section_name.lower():
            print("  - Suggested Table: interest_rate_indices")
            print("  - Purpose: Interest rate benchmark definitions")
            print("  - Schema: index_code, index_name, description, base_currency")
        
        else:
            print(f"  - Suggested Table: {section_name.lower().replace(' ', '_')}")
            print("  - Purpose: General reference lookup")
            print("  - Schema: Based on column analysis")
    
    print(f"\nNEXT STEPS:")
    print("1. Create database tables for each reference section")
    print("2. Design appropriate primary keys and indexes")
    print("3. Implement data extraction and transformation")
    print("4. Create validation rules for reference data integrity")
    print("5. Establish foreign key relationships with main asset table")

if __name__ == "__main__":
    analyze_reference_table()