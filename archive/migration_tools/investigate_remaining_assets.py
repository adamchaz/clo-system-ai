#!/usr/bin/env python3
"""
Investigate why only 384 assets were extracted instead of ~1,003
"""

import openpyxl
from pathlib import Path
import pandas as pd

def investigate_excel_data():
    """Investigate the Excel file to understand the asset count"""
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        print(f"Excel file not found: {excel_file}")
        return
    
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook['All Assets']
    
    print(f"Excel Investigation Report")
    print("=" * 50)
    print(f"Sheet dimensions: {sheet.max_column} columns × {sheet.max_row} rows")
    
    # Extract headers
    headers = []
    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=col).value
        headers.append(str(header_value) if header_value else f"Column_{col}")
    
    # Analyze row by row
    total_rows = 0
    rows_with_blkrock_id = 0
    rows_with_some_data = 0
    empty_rows = 0
    
    blkrock_id_col = 1  # BLKRockID is first column
    
    print(f"\nRow Analysis:")
    print(f"Row 1: Headers")
    
    sample_empty_rows = []
    sample_populated_rows = []
    
    for row_num in range(2, sheet.max_row + 1):
        total_rows += 1
        
        # Check BLKRockID
        blkrock_id = sheet.cell(row=row_num, column=blkrock_id_col).value
        
        # Check all values in row
        row_values = []
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row_num, column=col).value
            row_values.append(value)
        
        # Categorize row
        has_blkrock_id = blkrock_id is not None and str(blkrock_id).strip()
        has_any_data = any(value not in [None, '', ' '] for value in row_values)
        
        if has_blkrock_id:
            rows_with_blkrock_id += 1
            if len(sample_populated_rows) < 5:
                sample_populated_rows.append({
                    'row': row_num,
                    'blkrock_id': blkrock_id,
                    'issue_name': sheet.cell(row=row_num, column=2).value,
                    'par_amount': sheet.cell(row=row_num, column=7).value
                })
        
        if has_any_data:
            rows_with_some_data += 1
        else:
            empty_rows += 1
            if len(sample_empty_rows) < 5:
                sample_empty_rows.append(row_num)
    
    print(f"Total data rows (excluding header): {total_rows}")
    print(f"Rows with BLKRockID: {rows_with_blkrock_id}")
    print(f"Rows with some data: {rows_with_some_data}")
    print(f"Empty rows: {empty_rows}")
    
    print(f"\nSample populated rows:")
    for row_info in sample_populated_rows:
        print(f"  Row {row_info['row']}: {row_info['blkrock_id']} | {row_info['issue_name']} | Par: {row_info['par_amount']}")
    
    print(f"\nSample empty rows: {sample_empty_rows}")
    
    # Check specific row ranges
    print(f"\nChecking specific row ranges:")
    ranges_to_check = [(2, 50), (100, 150), (200, 250), (400, 450), (500, 550), (800, 850), (950, 1004)]
    
    for start_row, end_row in ranges_to_check:
        assets_in_range = 0
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            blkrock_id = sheet.cell(row=row, column=1).value
            if blkrock_id and str(blkrock_id).strip():
                assets_in_range += 1
        print(f"  Rows {start_row}-{end_row}: {assets_in_range} assets")
    
    # Check where assets stop
    print(f"\nFinding where populated data ends...")
    last_asset_row = None
    for row_num in range(2, sheet.max_row + 1):
        blkrock_id = sheet.cell(row=row_num, column=1).value
        if blkrock_id and str(blkrock_id).strip():
            last_asset_row = row_num
    
    if last_asset_row:
        print(f"Last row with BLKRockID: {last_asset_row}")
        print(f"Asset at last row: {sheet.cell(row=last_asset_row, column=1).value} | {sheet.cell(row=last_asset_row, column=2).value}")
        
        # Check a few rows after
        print(f"Checking rows after last asset:")
        for row in range(last_asset_row + 1, min(last_asset_row + 11, sheet.max_row + 1)):
            blkrock_id = sheet.cell(row=row, column=1).value
            issue_name = sheet.cell(row=row, column=2).value
            if blkrock_id or issue_name:
                print(f"  Row {row}: {blkrock_id} | {issue_name}")

if __name__ == "__main__":
    investigate_excel_data()