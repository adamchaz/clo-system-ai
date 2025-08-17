#!/usr/bin/env python3
"""
Comprehensive Excel/VBA Analyzer for CLO System
Analyzes TradeHypoPrelimv32.xlsm for structure, VBA code, formulas, and business logic
"""

import zipfile
import xml.etree.ElementTree as ET
import os
import re
import json
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl

class CLOExcelAnalyzer:
    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.analysis_results = {
            'file_structure': {},
            'worksheets': {},
            'vba_analysis': {},
            'business_logic': {},
            'data_flows': {},
            'formulas': {},
            'external_dependencies': {}
        }
    
    def analyze_comprehensive(self) -> Dict[str, Any]:
        """Run comprehensive analysis of the Excel file"""
        print("=" * 70)
        print("COMPREHENSIVE CLO EXCEL SYSTEM ANALYSIS")
        print("=" * 70)
        
        # Basic file info
        self._analyze_file_structure()
        
        # Worksheet analysis with openpyxl
        self._analyze_worksheets_detailed()
        
        # VBA and macro analysis from ZIP structure
        self._analyze_vba_from_zip()
        
        # Business logic identification
        self._identify_business_logic()
        
        # Data flow analysis
        self._analyze_data_flows()
        
        # Formula analysis
        self._analyze_formulas()
        
        # External dependencies
        self._analyze_external_dependencies()
        
        return self.analysis_results
    
    def _analyze_file_structure(self):
        """Analyze basic file structure"""
        print("\n1. FILE STRUCTURE ANALYSIS")
        print("-" * 40)
        
        file_size_mb = self.filepath.stat().st_size / (1024 * 1024)
        
        self.analysis_results['file_structure'] = {
            'filename': self.filepath.name,
            'size_mb': round(file_size_mb, 2),
            'type': 'Excel Macro-Enabled Workbook (.xlsm)',
            'contains_macros': True
        }
        
        print(f"File: {self.filepath.name}")
        print(f"Size: {file_size_mb:.2f} MB")
        print(f"Type: Excel Macro-Enabled Workbook")
        
        # Analyze ZIP structure
        try:
            with zipfile.ZipFile(self.filepath, 'r') as zip_ref:
                file_list = zip_ref.namelist()
                
                structure_info = {
                    'total_files': len(file_list),
                    'worksheet_files': len([f for f in file_list if f.startswith('xl/worksheets/')]),
                    'vba_files': len([f for f in file_list if 'vbaProject' in f]),
                    'external_connection_files': len([f for f in file_list if 'connections' in f.lower()]),
                    'has_custom_ui': any('customUI' in f for f in file_list),
                    'has_external_links': any('externalLinks' in f for f in file_list)
                }
                
                self.analysis_results['file_structure'].update(structure_info)
                
                print(f"Total internal files: {structure_info['total_files']}")
                print(f"Worksheet files: {structure_info['worksheet_files']}")
                print(f"VBA project files: {structure_info['vba_files']}")
                print(f"External connections: {structure_info['external_connection_files']}")
                print(f"Custom UI elements: {structure_info['has_custom_ui']}")
                print(f"External links: {structure_info['has_external_links']}")
                
        except Exception as e:
            print(f"Error analyzing ZIP structure: {e}")
    
    def _analyze_worksheets_detailed(self):
        """Detailed worksheet analysis using openpyxl"""
        print("\n2. DETAILED WORKSHEET ANALYSIS")
        print("-" * 40)
        
        try:
            workbook = openpyxl.load_workbook(self.filepath, data_only=False, keep_vba=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_analysis = self._analyze_single_worksheet(sheet, sheet_name)
                self.analysis_results['worksheets'][sheet_name] = sheet_analysis
                
                print(f"\nSheet: {sheet_name}")
                print(f"  Purpose: {sheet_analysis['inferred_purpose']}")
                print(f"  Data Range: {sheet_analysis['dimensions']}")
                print(f"  Formulas: {len(sheet_analysis['formulas'])}")
                print(f"  Data Validation Rules: {len(sheet_analysis['data_validations'])}")
                print(f"  Named Ranges: {len(sheet_analysis['named_ranges'])}")
                
        except Exception as e:
            print(f"Error analyzing worksheets: {e}")
    
    def _analyze_single_worksheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Analyze a single worksheet in detail"""
        analysis = {
            'name': sheet_name,
            'dimensions': f"{sheet.max_column} cols x {sheet.max_row} rows",
            'inferred_purpose': self._infer_sheet_purpose(sheet, sheet_name),
            'formulas': [],
            'data_validations': [],
            'named_ranges': [],
            'key_cells': [],
            'data_types': {},
            'conditional_formatting': [],
            'protection_status': {
                'sheet_protected': sheet.protection.sheet,
                'password_protected': sheet.protection.password is not None
            }
        }
        
        # Extract formulas (sample first 20)
        formula_count = 0
        for row in sheet.iter_rows(max_row=min(200, sheet.max_row)):
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    analysis['formulas'].append({
                        'cell': cell.coordinate,
                        'formula': cell.value,
                        'result': cell.internal_value if hasattr(cell, 'internal_value') else None
                    })
                    formula_count += 1
                    if formula_count >= 20:  # Limit output
                        break
            if formula_count >= 20:
                break
        
        # Check for data validations
        for validation in sheet.data_validations.dataValidation:
            analysis['data_validations'].append({
                'range': str(validation.sqref),
                'type': validation.type,
                'formula1': validation.formula1,
                'formula2': validation.formula2
            })
        
        # Identify key cells with important data
        analysis['key_cells'] = self._identify_key_cells(sheet)
        
        return analysis
    
    def _infer_sheet_purpose(self, sheet, sheet_name: str) -> str:
        """Infer the purpose of a worksheet based on its name and content"""
        name_lower = sheet_name.lower()
        
        # CLO-specific sheet purposes
        if 'input' in name_lower:
            return 'Data Input Sheet - Contains input parameters for CLO modeling'
        elif 'output' in name_lower:
            return 'Results Output Sheet - Contains calculated results and reports'
        elif 'model' in name_lower or 'run' in name_lower:
            return 'Model Control Sheet - Controls model execution and parameters'
        elif 'asset' in name_lower:
            return 'Asset Data Sheet - Contains loan/asset information'
        elif 'correlation' in name_lower:
            return 'Asset Correlation Matrix - Contains correlation data between assets'
        elif 'reference' in name_lower:
            return 'Reference Data Sheet - Contains lookup tables and reference data'
        elif 'filter' in name_lower:
            return 'Data Filter Sheet - Contains filtering criteria and rules'
        elif 'ranking' in name_lower:
            return 'Asset Ranking Sheet - Contains asset ranking and selection logic'
        elif 'rebalance' in name_lower:
            return 'Portfolio Rebalancing Sheet - Contains rebalancing calculations'
        elif 'hypo' in name_lower:
            return 'Hypothesis Testing Sheet - Contains scenario testing and analysis'
        elif 'mag' in name_lower:
            return 'Magnitude Analysis Sheet - Contains sensitivity/magnitude analysis'
        else:
            return 'General Worksheet - Purpose needs further analysis'
    
    def _identify_key_cells(self, sheet) -> List[Dict[str, Any]]:
        """Identify key cells that contain important calculations or controls"""
        key_cells = []
        
        # Look for cells with specific patterns
        for row in sheet.iter_rows(max_row=min(50, sheet.max_row), max_col=min(20, sheet.max_column)):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value_lower = cell.value.lower()
                    
                    # Identify control cells
                    if any(keyword in value_lower for keyword in [
                        'run model', 'execute', 'calculate', 'portfolio value',
                        'compliance', 'overcollateralization', 'coverage ratio',
                        'default rate', 'recovery rate', 'interest rate'
                    ]):
                        key_cells.append({
                            'cell': cell.coordinate,
                            'value': cell.value,
                            'type': 'control_parameter'
                        })
                
                # Look for cells with complex formulas
                elif cell.data_type == 'f' and cell.value:
                    if len(cell.value) > 50:  # Complex formula
                        key_cells.append({
                            'cell': cell.coordinate,
                            'formula': cell.value[:100] + "..." if len(cell.value) > 100 else cell.value,
                            'type': 'complex_calculation'
                        })
        
        return key_cells[:10]  # Limit to top 10
    
    def _analyze_vba_from_zip(self):
        """Analyze VBA code by examining the ZIP structure"""
        print("\n3. VBA PROJECT ANALYSIS")
        print("-" * 40)
        
        try:
            with zipfile.ZipFile(self.filepath, 'r') as zip_ref:
                vba_files = [f for f in zip_ref.namelist() if 'vbaProject' in f]
                
                vba_analysis = {
                    'has_vba': len(vba_files) > 0,
                    'vba_files': vba_files,
                    'estimated_modules': 0,
                    'potential_functionality': []
                }
                
                for vba_file in vba_files:
                    print(f"VBA File: {vba_file}")
                    
                    try:
                        vba_content = zip_ref.read(vba_file)
                        
                        # Basic VBA analysis
                        if b'Microsoft Visual Basic for Applications' in vba_content:
                            print("  - Contains VBA code modules")
                            vba_analysis['estimated_modules'] += 1
                        
                        # Look for common VBA patterns (simplified)
                        if b'Sub ' in vba_content or b'Function ' in vba_content:
                            vba_analysis['potential_functionality'].append('Subroutines and Functions')
                        
                        if b'Solver' in vba_content:
                            vba_analysis['potential_functionality'].append('Excel Solver Integration')
                        
                        if b'SQL' in vba_content or b'ADODB' in vba_content:
                            vba_analysis['potential_functionality'].append('Database Connectivity')
                        
                        if b'FileSystem' in vba_content or b'Dir(' in vba_content:
                            vba_analysis['potential_functionality'].append('File System Operations')
                            
                        # Look for CLO-specific functionality
                        clo_keywords = [b'portfolio', b'asset', b'rating', b'default', b'recovery',
                                      b'correlation', b'simulation', b'monte', b'carlo', b'compliance']
                        
                        for keyword in clo_keywords:
                            if keyword in vba_content.lower():
                                vba_analysis['potential_functionality'].append(f'CLO Logic: {keyword.decode()}')
                        
                    except Exception as e:
                        print(f"  - Error analyzing {vba_file}: {e}")
                
                self.analysis_results['vba_analysis'] = vba_analysis
                
                print(f"VBA Files Found: {len(vba_files)}")
                print(f"Estimated Modules: {vba_analysis['estimated_modules']}")
                if vba_analysis['potential_functionality']:
                    print("Potential VBA Functionality:")
                    for func in set(vba_analysis['potential_functionality']):
                        print(f"  - {func}")
                
        except Exception as e:
            print(f"Error analyzing VBA: {e}")
    
    def _identify_business_logic(self):
        """Identify CLO business logic patterns"""
        print("\n4. BUSINESS LOGIC IDENTIFICATION")
        print("-" * 40)
        
        business_logic = {
            'clo_components': [],
            'calculation_types': [],
            'workflow_stages': [],
            'compliance_tests': [],
            'risk_management': []
        }
        
        # Analyze worksheet names and purposes for business logic
        for sheet_name, sheet_info in self.analysis_results['worksheets'].items():
            purpose = sheet_info['inferred_purpose']
            
            # Identify CLO components
            if 'asset' in sheet_name.lower():
                business_logic['clo_components'].append('Asset Portfolio Management')
            if 'correlation' in sheet_name.lower():
                business_logic['clo_components'].append('Asset Correlation Modeling')
            if 'rating' in sheet_name.lower():
                business_logic['clo_components'].append('Credit Rating Analysis')
            
            # Identify calculation types
            if len(sheet_info['formulas']) > 10:
                if 'mag' in sheet_name.lower():
                    business_logic['calculation_types'].append('Magnitude/Sensitivity Analysis')
                elif 'hypo' in sheet_name.lower():
                    business_logic['calculation_types'].append('Hypothesis Testing')
                elif 'ranking' in sheet_name.lower():
                    business_logic['calculation_types'].append('Asset Ranking/Selection')
                elif 'rebalance' in sheet_name.lower():
                    business_logic['calculation_types'].append('Portfolio Rebalancing')
            
            # Look for compliance tests in formulas
            for formula_info in sheet_info['formulas'][:10]:  # Check first 10 formulas
                formula = formula_info['formula'].upper() if formula_info['formula'] else ""
                
                if any(keyword in formula for keyword in ['COVERAGE', 'RATIO', 'TEST', 'COMPLIANCE']):
                    business_logic['compliance_tests'].append(f"Formula in {sheet_name}: {formula[:50]}...")
                
                if any(keyword in formula for keyword in ['DEFAULT', 'LOSS', 'RECOVERY', 'RISK']):
                    business_logic['risk_management'].append(f"Risk calculation in {sheet_name}")
        
        # Identify workflow stages based on sheet analysis
        input_sheets = [name for name, info in self.analysis_results['worksheets'].items() 
                       if 'input' in info['inferred_purpose'].lower()]
        output_sheets = [name for name, info in self.analysis_results['worksheets'].items() 
                        if 'output' in info['inferred_purpose'].lower()]
        control_sheets = [name for name, info in self.analysis_results['worksheets'].items() 
                         if 'control' in info['inferred_purpose'].lower()]
        
        business_logic['workflow_stages'] = [
            f"1. Data Input Stage: {len(input_sheets)} sheets",
            f"2. Calculation/Processing Stage: {len(self.analysis_results['worksheets']) - len(input_sheets) - len(output_sheets)} sheets",
            f"3. Output/Reporting Stage: {len(output_sheets)} sheets"
        ]
        
        if control_sheets:
            business_logic['workflow_stages'].insert(0, f"0. Model Control Stage: {len(control_sheets)} sheets")
        
        self.analysis_results['business_logic'] = business_logic
        
        print("CLO System Components:")
        for component in set(business_logic['clo_components']):
            print(f"  - {component}")
        
        print("\nCalculation Types Identified:")
        for calc_type in set(business_logic['calculation_types']):
            print(f"  - {calc_type}")
        
        print("\nWorkflow Stages:")
        for stage in business_logic['workflow_stages']:
            print(f"  - {stage}")
    
    def _analyze_data_flows(self):
        """Analyze data flow between worksheets"""
        print("\n5. DATA FLOW ANALYSIS")
        print("-" * 40)
        
        data_flows = {
            'inter_sheet_references': {},
            'data_sources': [],
            'data_sinks': [],
            'calculation_chain': []
        }
        
        # Analyze formulas for inter-sheet references
        for sheet_name, sheet_info in self.analysis_results['worksheets'].items():
            references = []
            
            for formula_info in sheet_info['formulas']:
                formula = formula_info['formula'] if formula_info['formula'] else ""
                
                # Look for sheet references (e.g., 'SheetName'!A1)
                sheet_refs = re.findall(r"'([^']+)'!", formula)
                sheet_refs.extend(re.findall(r"([A-Za-z\s]+[A-Za-z])!", formula))
                
                for ref_sheet in sheet_refs:
                    if ref_sheet != sheet_name and ref_sheet in self.analysis_results['worksheets']:
                        references.append(ref_sheet)
            
            if references:
                data_flows['inter_sheet_references'][sheet_name] = list(set(references))
        
        # Identify data sources and sinks
        input_sheets = [name for name, info in self.analysis_results['worksheets'].items() 
                       if 'input' in info['inferred_purpose'].lower()]
        output_sheets = [name for name, info in self.analysis_results['worksheets'].items() 
                        if 'output' in info['inferred_purpose'].lower()]
        
        data_flows['data_sources'] = input_sheets
        data_flows['data_sinks'] = output_sheets
        
        self.analysis_results['data_flows'] = data_flows
        
        print("Data Flow Summary:")
        print(f"  Input Sheets (Data Sources): {len(data_flows['data_sources'])}")
        for source in data_flows['data_sources']:
            print(f"    - {source}")
        
        print(f"\n  Output Sheets (Data Sinks): {len(data_flows['data_sinks'])}")
        for sink in data_flows['data_sinks']:
            print(f"    - {sink}")
        
        print(f"\n  Inter-sheet Dependencies:")
        for sheet, refs in data_flows['inter_sheet_references'].items():
            if refs:
                print(f"    {sheet} -> {', '.join(refs)}")
    
    def _analyze_formulas(self):
        """Analyze complex formulas for business logic"""
        print("\n6. FORMULA COMPLEXITY ANALYSIS")
        print("-" * 40)
        
        formula_analysis = {
            'total_formulas': 0,
            'complex_formulas': [],
            'function_usage': {},
            'array_formulas': [],
            'lookup_formulas': []
        }
        
        excel_functions = {}
        
        for sheet_name, sheet_info in self.analysis_results['worksheets'].items():
            formula_analysis['total_formulas'] += len(sheet_info['formulas'])
            
            for formula_info in sheet_info['formulas']:
                formula = formula_info['formula'] if formula_info['formula'] else ""
                
                # Count function usage
                functions = re.findall(r'([A-Z]+)\s*\(', formula)
                for func in functions:
                    excel_functions[func] = excel_functions.get(func, 0) + 1
                
                # Identify complex formulas (>100 characters or nested functions)
                if len(formula) > 100 or formula.count('(') > 3:
                    formula_analysis['complex_formulas'].append({
                        'sheet': sheet_name,
                        'cell': formula_info['cell'],
                        'formula_length': len(formula),
                        'nesting_level': formula.count('('),
                        'preview': formula[:80] + "..." if len(formula) > 80 else formula
                    })
                
                # Identify array formulas
                if formula.startswith('{') and formula.endswith('}'):
                    formula_analysis['array_formulas'].append({
                        'sheet': sheet_name,
                        'cell': formula_info['cell'],
                        'formula': formula[:80] + "..." if len(formula) > 80 else formula
                    })
                
                # Identify lookup formulas
                if any(func in formula.upper() for func in ['VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH']):
                    formula_analysis['lookup_formulas'].append({
                        'sheet': sheet_name,
                        'cell': formula_info['cell'],
                        'type': 'lookup'
                    })
        
        formula_analysis['function_usage'] = dict(sorted(excel_functions.items(), 
                                                        key=lambda x: x[1], reverse=True))
        
        self.analysis_results['formulas'] = formula_analysis
        
        print(f"Total Formulas: {formula_analysis['total_formulas']}")
        print(f"Complex Formulas: {len(formula_analysis['complex_formulas'])}")
        print(f"Array Formulas: {len(formula_analysis['array_formulas'])}")
        print(f"Lookup Formulas: {len(formula_analysis['lookup_formulas'])}")
        
        print("\nTop Excel Functions Used:")
        for func, count in list(formula_analysis['function_usage'].items())[:10]:
            print(f"  {func}: {count} times")
    
    def _analyze_external_dependencies(self):
        """Analyze external dependencies and connections"""
        print("\n7. EXTERNAL DEPENDENCIES ANALYSIS")
        print("-" * 40)
        
        dependencies = {
            'external_files': [],
            'add_ins': [],
            'database_connections': [],
            'web_queries': [],
            'special_features': []
        }
        
        try:
            with zipfile.ZipFile(self.filepath, 'r') as zip_ref:
                file_list = zip_ref.namelist()
                
                # Check for external connections
                connection_files = [f for f in file_list if 'connections' in f.lower()]
                if connection_files:
                    dependencies['database_connections'] = connection_files
                
                # Check for external links
                if any('externalLinks' in f for f in file_list):
                    dependencies['external_files'].append('Excel external links detected')
                
                # Check for web queries
                if any('webQuery' in f for f in file_list):
                    dependencies['web_queries'].append('Web query connections detected')
                
                # Check for special Excel features
                if any('pivotCache' in f for f in file_list):
                    dependencies['special_features'].append('Pivot Tables')
                
                if any('chart' in f for f in file_list):
                    dependencies['special_features'].append('Charts/Graphs')
                
                if any('drawing' in f for f in file_list):
                    dependencies['special_features'].append('Drawings/Shapes')
        
        except Exception as e:
            print(f"Error analyzing external dependencies: {e}")
        
        # Check VBA for external dependencies
        vba_analysis = self.analysis_results.get('vba_analysis', {})
        if 'Excel Solver Integration' in vba_analysis.get('potential_functionality', []):
            dependencies['add_ins'].append('Excel Solver Add-in')
        
        if 'Database Connectivity' in vba_analysis.get('potential_functionality', []):
            dependencies['add_ins'].append('Database connectivity (ADODB/SQL)')
        
        self.analysis_results['external_dependencies'] = dependencies
        
        print("External Dependencies Found:")
        for dep_type, items in dependencies.items():
            if items:
                print(f"  {dep_type.replace('_', ' ').title()}:")
                for item in items:
                    print(f"    - {item}")
    
    def save_comprehensive_analysis(self, output_path: Optional[str] = None):
        """Save comprehensive analysis to JSON file"""
        if not output_path:
            output_path = self.filepath.parent / f"{self.filepath.stem}_comprehensive_analysis.json"
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.analysis_results, f, indent=2, default=str)
        
        print(f"\nComprehensive analysis saved to: {output_path}")
        return output_path
    
    def generate_conversion_strategy(self):
        """Generate Python conversion strategy based on analysis"""
        print("\n" + "=" * 70)
        print("PYTHON CONVERSION STRATEGY RECOMMENDATIONS")
        print("=" * 70)
        
        strategy = {
            'architecture_recommendations': [],
            'technology_stack': [],
            'conversion_priorities': [],
            'potential_challenges': [],
            'implementation_phases': []
        }
        
        # Architecture recommendations
        strategy['architecture_recommendations'] = [
            "Multi-layered architecture: Data Layer, Business Logic Layer, API Layer, UI Layer",
            "Microservices approach for different CLO functions (portfolio management, risk calculation, reporting)",
            "Event-driven architecture for real-time portfolio updates",
            "Database-first design replacing Excel's file-based storage"
        ]
        
        # Technology stack
        strategy['technology_stack'] = [
            "Backend: Python with FastAPI or Flask",
            "Data Processing: Pandas, NumPy for calculations",
            "Database: PostgreSQL or SQL Server for structured data",
            "Caching: Redis for performance optimization",
            "Frontend: React.js or Vue.js for modern UI",
            "API: RESTful APIs with OpenAPI documentation",
            "Testing: pytest for comprehensive testing",
            "Deployment: Docker containers with orchestration"
        ]
        
        # Conversion priorities based on analysis
        total_formulas = self.analysis_results.get('formulas', {}).get('total_formulas', 0)
        complex_formulas = len(self.analysis_results.get('formulas', {}).get('complex_formulas', []))
        
        strategy['conversion_priorities'] = [
            f"1. Data Model Design - Convert {len(self.analysis_results['worksheets'])} sheets to database schema",
            f"2. Business Logic Migration - Convert {total_formulas} formulas to Python functions",
            f"3. Complex Calculations - Handle {complex_formulas} complex formulas requiring special attention",
            "4. VBA Macro Conversion - Convert VBA business logic to Python",
            "5. User Interface Recreation - Build modern web interface",
            "6. Integration & Testing - Ensure calculation accuracy"
        ]
        
        # Potential challenges
        strategy['potential_challenges'] = [
            "Excel-specific functions need Python equivalents",
            "VBA macros require complete reengineering",
            "Complex formula dependencies must be preserved",
            "User workflow adaptation to web-based system",
            "Performance optimization for large datasets",
            "Data validation and business rule enforcement"
        ]
        
        # Add specific challenges based on analysis
        if self.analysis_results.get('external_dependencies', {}).get('add_ins'):
            strategy['potential_challenges'].append("Excel add-in dependencies need alternative solutions")
        
        if complex_formulas > 50:
            strategy['potential_challenges'].append(f"High number of complex formulas ({complex_formulas}) requires careful testing")
        
        # Implementation phases
        strategy['implementation_phases'] = [
            "Phase 1: Data Architecture & Core Models (4-6 weeks)",
            "Phase 2: Business Logic & Calculation Engine (8-10 weeks)", 
            "Phase 3: API Development & Integration (4-6 weeks)",
            "Phase 4: User Interface & Experience (6-8 weeks)",
            "Phase 5: Testing, Validation & Deployment (4-6 weeks)",
            "Phase 6: Training & Migration (2-4 weeks)"
        ]
        
        # Print strategy
        for section, items in strategy.items():
            print(f"\n{section.replace('_', ' ').upper()}:")
            print("-" * 50)
            for item in items:
                print(f"• {item}")
        
        return strategy

def main():
    """Main execution function"""
    filepath = "TradeHypoPrelimv32.xlsm"
    
    if not os.path.exists(filepath):
        print(f"Error: File '{filepath}' not found in current directory")
        print("Current directory contents:")
        for item in os.listdir('.'):
            print(f"  {item}")
        return
    
    # Run comprehensive analysis
    analyzer = CLOExcelAnalyzer(filepath)
    results = analyzer.analyze_comprehensive()
    
    # Save results
    output_file = analyzer.save_comprehensive_analysis()
    
    # Generate conversion strategy
    strategy = analyzer.generate_conversion_strategy()
    
    print(f"\n{'='*70}")
    print("ANALYSIS COMPLETE")
    print(f"{'='*70}")
    print(f"Detailed results saved to: {output_file}")
    print("\nThis analysis provides the foundation for converting the CLO Excel system to Python.")

if __name__ == "__main__":
    main()