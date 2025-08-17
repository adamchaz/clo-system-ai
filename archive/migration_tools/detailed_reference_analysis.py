#!/usr/bin/env python3
"""
Detailed analysis of Reference Table structure and content
"""

import openpyxl
from pathlib import Path
import pandas as pd
from datetime import datetime

def detailed_reference_analysis():
    """Detailed analysis of Reference Table"""
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        print(f"Excel file not found: {excel_file}")
        return
    
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook['Reference Table']
    
    print("DETAILED REFERENCE TABLE ANALYSIS")
    print("=" * 60)
    print(f"Dimensions: {sheet.max_column} columns × {sheet.max_row} rows")
    
    # Extract all headers
    headers = []
    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=col).value
        headers.append(str(header_value) if header_value else f"Column_{col}")
    
    print(f"\nCOLUMN HEADERS ({len(headers)} total):")
    for i, header in enumerate(headers, 1):
        print(f"  {i:2d}. {header}")
    
    # Sample the first 20 data rows to understand content
    print(f"\nSAMPLE DATA (first 10 rows):")
    
    sample_data = []
    for row_num in range(2, min(12, sheet.max_row + 1)):
        row_data = {}
        has_data = False
        
        for col_idx, header in enumerate(headers, 1):
            value = sheet.cell(row=row_num, column=col_idx).value
            row_data[header] = value
            if value is not None and str(value).strip():
                has_data = True
        
        if has_data:
            sample_data.append(row_data)
    
    # Display sample data focusing on key columns
    key_columns = ['Column_1', 'Column_2', 'Column_3']  # Will update based on actual headers
    if len(headers) >= 3:
        key_columns = headers[:3]
    
    print(f"\nKey columns for first 10 rows:")
    for i, row_data in enumerate(sample_data, 2):
        print(f"Row {i}:")
        for col in key_columns:
            value = row_data.get(col, 'None')
            print(f"  {col}: {value}")
        print()
        if i > 6:  # Limit output
            break
    
    # Analyze data distribution across all rows
    print(f"DATA DISTRIBUTION ANALYSIS:")
    
    # Count non-empty values per column
    non_empty_counts = {}
    data_types = {}
    sample_values = {}
    
    for col_idx, header in enumerate(headers[:20], 1):  # Analyze first 20 columns
        non_empty_count = 0
        numeric_count = 0
        text_count = 0
        date_count = 0
        sample_vals = []
        
        # Sample every 100th row for performance
        for row_num in range(2, sheet.max_row + 1, 100):
            value = sheet.cell(row=row_num, column=col_idx).value
            
            if value is not None and str(value).strip():
                non_empty_count += 1
                
                if len(sample_vals) < 5:
                    sample_vals.append(str(value))
                
                if isinstance(value, (int, float)):
                    numeric_count += 1
                elif isinstance(value, str):
                    text_count += 1
                elif hasattr(value, 'date'):
                    date_count += 1
        
        non_empty_counts[header] = non_empty_count
        sample_values[header] = sample_vals
        
        # Determine primary data type
        total_sampled = non_empty_count
        if total_sampled > 0:
            if numeric_count > total_sampled * 0.7:
                data_types[header] = "Numeric"
            elif date_count > total_sampled * 0.3:
                data_types[header] = "Date"
            else:
                data_types[header] = "Text"
        else:
            data_types[header] = "Empty"
    
    print(f"\nCOLUMN ANALYSIS (first 20 columns):")
    for header in headers[:20]:
        count = non_empty_counts.get(header, 0)
        dtype = data_types.get(header, "Unknown")
        samples = sample_values.get(header, [])
        
        print(f"{header[:30]:30} | {dtype:8} | {count:4d} values | {', '.join(samples[:3])}")
    
    # Look for patterns that suggest this is reference/lookup data
    print(f"\nREFERENCE DATA PATTERN ANALYSIS:")
    
    # Check if first column looks like keys/codes
    first_col_values = set()
    for row_num in range(2, min(102, sheet.max_row + 1)):  # Check first 100 rows
        value = sheet.cell(row=row_num, column=1).value
        if value:
            first_col_values.add(str(value))
    
    print(f"First column unique values (sample): {len(first_col_values)}")
    if first_col_values:
        sample_first_col = list(first_col_values)[:10]
        print(f"Sample values: {', '.join(sample_first_col)}")
    
    # Check for common reference table patterns
    header_text = ' '.join(headers).lower()
    
    patterns = {
        'ratings': ['rating', 'grade', 'score'],
        'industry': ['industry', 'sector', 'naics', 'sic'],
        'geography': ['country', 'region', 'state', 'geography'],
        'currency': ['currency', 'ccy', 'fx'],
        'indices': ['index', 'benchmark', 'rate'],
        'mappings': ['code', 'id', 'mapping', 'lookup']
    }
    
    detected_patterns = []
    for pattern_name, keywords in patterns.items():
        if any(keyword in header_text for keyword in keywords):
            detected_patterns.append(pattern_name)
    
    print(f"\nDETECTED PATTERNS: {', '.join(detected_patterns) if detected_patterns else 'General reference data'}")
    
    # Migration strategy recommendations
    print(f"\nMIGRATION STRATEGY RECOMMENDATIONS:")
    print("-" * 50)
    
    if detected_patterns:
        for pattern in detected_patterns:
            if pattern == 'ratings':
                print(f"✓ RATINGS DATA DETECTED")
                print(f"  - Create 'credit_ratings' table")
                print(f"  - Schema: rating_agency, rating_symbol, numeric_score, description")
                print(f"  - Use for asset credit rating lookups")
            
            elif pattern == 'industry':
                print(f"✓ INDUSTRY DATA DETECTED")
                print(f"  - Create 'industry_classifications' table")
                print(f"  - Schema: classification_system, industry_code, industry_name, parent_code")
                print(f"  - Use for asset industry categorization")
            
            elif pattern == 'geography':
                print(f"✓ GEOGRAPHIC DATA DETECTED")
                print(f"  - Create 'geographic_regions' table")
                print(f"  - Schema: country_code, country_name, region, sub_region")
                print(f"  - Use for asset geographic analysis")
            
            elif pattern == 'currency':
                print(f"✓ CURRENCY DATA DETECTED") 
                print(f"  - Create 'currencies' table")
                print(f"  - Schema: currency_code, currency_name, symbol, decimal_places")
                print(f"  - Use for multi-currency calculations")
            
            elif pattern == 'indices':
                print(f"✓ INDEX DATA DETECTED")
                print(f"  - Create 'interest_rate_indices' table") 
                print(f"  - Schema: index_code, index_name, description, base_currency")
                print(f"  - Use for floating rate calculations")
    
    else:
        print(f"✓ GENERAL REFERENCE DATA")
        print(f"  - Create comprehensive reference tables based on column groupings")
        print(f"  - Consider splitting large table into logical sections")
        print(f"  - Implement proper indexing for lookup performance")
    
    print(f"\nNEXT STEPS FOR MIGRATION:")
    print(f"1. Design normalized database schema based on detected patterns")
    print(f"2. Create staging tables for data transformation")
    print(f"3. Implement data quality validation rules") 
    print(f"4. Test foreign key relationships with asset table")
    print(f"5. Create indexes for optimal lookup performance")
    
    return {
        'total_columns': len(headers),
        'total_rows': sheet.max_row,
        'headers': headers,
        'detected_patterns': detected_patterns,
        'sample_data': sample_data[:5]
    }

if __name__ == "__main__":
    result = detailed_reference_analysis()