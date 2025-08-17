#!/usr/bin/env python3
"""
Analyze all remaining worksheets in the Excel file to identify additional data for migration
"""

import openpyxl
from pathlib import Path
from datetime import datetime, date
import json
from typing import Dict, List, Any

class WorksheetAnalyzer:
    """Analyze all worksheets to identify migration candidates"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.analysis_results = {}
        
        # Already migrated sheets
        self.migrated_sheets = {'All Assets', 'Reference Table'}
        
        # CLO data type patterns
        self.data_patterns = {
            'model_config': ['run', 'model', 'config', 'parameter', 'setting'],
            'scenario_data': ['mag', 'scenario', 'hypo', 'assumption'],
            'correlation_data': ['correlation', 'correl', 'cov', 'variance'],
            'portfolio_data': ['deal', 'portfolio', 'allocation', 'weight'],
            'output_data': ['output', 'result', 'ranking', 'rebalance'],
            'filter_criteria': ['filter', 'criteria', 'constraint', 'rule'],
            'cashflow_data': ['cashflow', 'payment', 'amortization', 'schedule'],
            'rating_data': ['rating', 'credit', 'migration', 'transition'],
            'pricing_data': ['price', 'yield', 'spread', 'curve', 'libor']
        }
    
    def load_workbook(self):
        """Load Excel workbook"""
        try:
            print(f"Loading Excel workbook: {self.excel_path}")
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            print(f"Successfully loaded workbook with {len(self.workbook.sheetnames)} worksheets")
            return True
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def analyze_worksheet(self, sheet_name: str) -> Dict[str, Any]:
        """Analyze individual worksheet structure and content"""
        
        if sheet_name in self.migrated_sheets:
            return {
                'status': 'already_migrated',
                'description': 'This sheet has already been migrated'
            }
        
        try:
            sheet = self.workbook[sheet_name]
            
            # Basic dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Sample data from key areas
            sample_data = self._extract_sample_data(sheet)
            
            # Determine data type and priority
            data_type = self._classify_data_type(sheet_name, sample_data)
            priority = self._assess_migration_priority(data_type, sample_data)
            
            # Analyze data density
            data_density = self._calculate_data_density(sheet, max_row, max_col)
            
            # Check for structured data
            structure_info = self._analyze_data_structure(sheet, sample_data)
            
            analysis = {
                'sheet_name': sheet_name,
                'dimensions': f"{max_col} cols × {max_row} rows",
                'total_cells': max_row * max_col,
                'data_density': data_density,
                'data_type': data_type,
                'migration_priority': priority,
                'structure_info': structure_info,
                'sample_data': sample_data,
                'recommended_approach': self._recommend_migration_approach(data_type, structure_info)
            }
            
            return analysis
            
        except Exception as e:
            return {
                'sheet_name': sheet_name,
                'error': f"Analysis failed: {e}",
                'migration_priority': 'unknown'
            }
    
    def _extract_sample_data(self, sheet, sample_rows=10, sample_cols=10):
        """Extract sample data from worksheet"""
        sample_data = {
            'headers': [],
            'first_rows': [],
            'data_types': {},
            'non_empty_cells': 0
        }
        
        try:
            # Extract potential headers (first few rows)
            for row in range(1, min(4, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(sample_cols + 1, sheet.max_column + 1)):
                    value = sheet.cell(row=row, column=col).value
                    row_data.append(value)
                sample_data['first_rows'].append(row_data)
            
            # Extract headers from row 1
            for col in range(1, min(sample_cols + 1, sheet.max_column + 1)):
                header_value = sheet.cell(row=1, column=col).value
                sample_data['headers'].append(str(header_value) if header_value else f"Col_{col}")
            
            # Sample data types
            for row in range(1, min(sample_rows + 1, sheet.max_row + 1)):
                for col in range(1, min(sample_cols + 1, sheet.max_column + 1)):
                    value = sheet.cell(row=row, column=col).value
                    if value is not None:
                        sample_data['non_empty_cells'] += 1
                        value_type = type(value).__name__
                        if value_type not in sample_data['data_types']:
                            sample_data['data_types'][value_type] = 0
                        sample_data['data_types'][value_type] += 1
            
        except Exception as e:
            sample_data['extraction_error'] = str(e)
        
        return sample_data
    
    def _classify_data_type(self, sheet_name: str, sample_data: Dict) -> str:
        """Classify the type of data in the worksheet"""
        sheet_name_lower = sheet_name.lower()
        
        # Check sheet name against patterns
        for data_type, keywords in self.data_patterns.items():
            if any(keyword in sheet_name_lower for keyword in keywords):
                return data_type
        
        # Check sample data content
        headers_text = ' '.join(sample_data.get('headers', [])).lower()
        
        for data_type, keywords in self.data_patterns.items():
            if any(keyword in headers_text for keyword in keywords):
                return data_type
        
        # Default classification
        if 'input' in sheet_name_lower:
            return 'scenario_data'
        elif 'output' in sheet_name_lower:
            return 'output_data'
        else:
            return 'unknown'
    
    def _assess_migration_priority(self, data_type: str, sample_data: Dict) -> str:
        """Assess migration priority based on data type and content"""
        
        priority_map = {
            'model_config': 'HIGH',      # Essential for model execution
            'scenario_data': 'HIGH',     # Core modeling scenarios  
            'correlation_data': 'HIGH',  # Risk management critical
            'portfolio_data': 'MEDIUM', # Important for deal management
            'pricing_data': 'HIGH',     # Valuation critical
            'rating_data': 'MEDIUM',    # Already have some rating data
            'cashflow_data': 'MEDIUM',  # Important but may be derived
            'output_data': 'LOW',       # Can be regenerated
            'filter_criteria': 'MEDIUM', # Important for compliance
            'unknown': 'LOW'
        }
        
        base_priority = priority_map.get(data_type, 'LOW')
        
        # Adjust based on data density
        data_density = sample_data.get('non_empty_cells', 0)
        
        if data_density > 50:  # Substantial data
            if base_priority == 'LOW':
                return 'MEDIUM'
        elif data_density < 10:  # Very sparse data
            if base_priority == 'HIGH':
                return 'MEDIUM'
            elif base_priority == 'MEDIUM':
                return 'LOW'
        
        return base_priority
    
    def _calculate_data_density(self, sheet, max_row: int, max_col: int) -> Dict[str, Any]:
        """Calculate data density in the worksheet"""
        
        # Sample approach for large sheets
        sample_rows = min(100, max_row)
        sample_cols = min(20, max_col)
        
        total_sampled = sample_rows * sample_cols
        non_empty_count = 0
        
        for row in range(1, sample_rows + 1):
            for col in range(1, sample_cols + 1):
                value = sheet.cell(row=row, column=col).value
                if value is not None and str(value).strip():
                    non_empty_count += 1
        
        density_percentage = (non_empty_count / total_sampled * 100) if total_sampled > 0 else 0
        
        return {
            'sampled_cells': total_sampled,
            'non_empty_cells': non_empty_count,
            'density_percentage': round(density_percentage, 2),
            'estimated_total_data_cells': int((non_empty_count / total_sampled) * max_row * max_col) if total_sampled > 0 else 0
        }
    
    def _analyze_data_structure(self, sheet, sample_data: Dict) -> Dict[str, Any]:
        """Analyze the structure of data in the worksheet"""
        
        structure = {
            'has_headers': False,
            'table_like': False,
            'multi_section': False,
            'data_starts_row': 1,
            'estimated_records': 0
        }
        
        # Check if first row looks like headers
        first_row = sample_data.get('first_rows', [[]])[0] if sample_data.get('first_rows') else []
        if first_row:
            text_count = sum(1 for val in first_row if isinstance(val, str) and val and val.strip())
            if text_count >= len(first_row) * 0.5:  # 50% or more text values
                structure['has_headers'] = True
                structure['data_starts_row'] = 2
        
        # Check if it's table-like (consistent column structure)
        first_rows = sample_data.get('first_rows', [])
        if len(first_rows) >= 3:
            consistent_structure = True
            first_row_len = len([v for v in first_rows[0] if v is not None])
            
            for row in first_rows[1:]:
                row_len = len([v for v in row if v is not None])
                if abs(row_len - first_row_len) > 2:  # Allow some variation
                    consistent_structure = False
                    break
            
            structure['table_like'] = consistent_structure
        
        # Estimate number of records based on density
        density_info = self._calculate_data_density(sheet, sheet.max_row, sheet.max_column)
        if density_info['density_percentage'] > 10:  # Significant data
            structure['estimated_records'] = max(1, int(sheet.max_row * density_info['density_percentage'] / 100))
        
        return structure
    
    def _recommend_migration_approach(self, data_type: str, structure_info: Dict) -> Dict[str, str]:
        """Recommend migration approach based on analysis"""
        
        if data_type == 'model_config':
            return {
                'approach': 'key_value_pairs',
                'table_design': 'model_parameters (parameter_name, parameter_value, section, description)',
                'special_handling': 'Parse configuration sections and nested parameters'
            }
        
        elif data_type == 'scenario_data':
            return {
                'approach': 'scenario_tables',
                'table_design': 'scenarios (scenario_name, parameter_name, parameter_value, data_type)',
                'special_handling': 'Group by MAG scenario and normalize parameters'
            }
        
        elif data_type == 'correlation_data':
            return {
                'approach': 'matrix_storage',
                'table_design': 'asset_correlations (asset1_id, asset2_id, correlation_value, correlation_type)',
                'special_handling': 'Parse correlation matrices and store as pairs'
            }
        
        elif data_type == 'portfolio_data':
            return {
                'approach': 'portfolio_holdings',
                'table_design': 'deal_assets (deal_name, asset_id, allocation_percent, weight)',
                'special_handling': 'Link to existing asset table via foreign keys'
            }
        
        elif data_type == 'output_data':
            return {
                'approach': 'results_storage',
                'table_design': 'model_outputs (run_date, asset_id, metric_name, metric_value)',
                'special_handling': 'Store as time-series data for trend analysis'
            }
        
        elif structure_info.get('table_like', False):
            return {
                'approach': 'table_migration',
                'table_design': 'Based on column analysis',
                'special_handling': 'Standard table migration with data type inference'
            }
        
        else:
            return {
                'approach': 'json_storage',
                'table_design': 'flexible_data (sheet_name, section, data_json)',
                'special_handling': 'Store as JSON for flexible querying'
            }
    
    def generate_comprehensive_analysis(self):
        """Generate analysis for all worksheets"""
        
        if not self.load_workbook():
            return None
        
        print(f"\nCOMPREHENSIVE WORKSHEET ANALYSIS")
        print("=" * 70)
        
        all_sheets = self.workbook.sheetnames
        print(f"Total worksheets: {len(all_sheets)}")
        print(f"Already migrated: {len(self.migrated_sheets)}")
        print(f"Remaining for analysis: {len(all_sheets) - len(self.migrated_sheets)}")
        
        # Analyze each worksheet
        for sheet_name in all_sheets:
            print(f"\nAnalyzing: {sheet_name}")
            analysis = self.analyze_worksheet(sheet_name)
            self.analysis_results[sheet_name] = analysis
            
            if analysis.get('status') == 'already_migrated':
                print(f"  Status: ✅ Already migrated")
            else:
                print(f"  Dimensions: {analysis.get('dimensions', 'unknown')}")
                print(f"  Data Type: {analysis.get('data_type', 'unknown')}")
                print(f"  Priority: {analysis.get('migration_priority', 'unknown')}")
                
                density = analysis.get('data_density', {})
                if density:
                    print(f"  Data Density: {density.get('density_percentage', 0):.1f}% ({density.get('non_empty_cells', 0)} cells)")
        
        return self.analysis_results
    
    def create_migration_roadmap(self):
        """Create prioritized migration roadmap"""
        
        if not self.analysis_results:
            print("No analysis results available. Run generate_comprehensive_analysis first.")
            return
        
        print(f"\n\nMIGRATION ROADMAP")
        print("=" * 70)
        
        # Group by priority
        priority_groups = {'HIGH': [], 'MEDIUM': [], 'LOW': []}
        
        for sheet_name, analysis in self.analysis_results.items():
            if analysis.get('status') == 'already_migrated':
                continue
            
            priority = analysis.get('migration_priority', 'LOW')
            if priority in priority_groups:
                priority_groups[priority].append({
                    'sheet': sheet_name,
                    'data_type': analysis.get('data_type', 'unknown'),
                    'estimated_records': analysis.get('structure_info', {}).get('estimated_records', 0),
                    'approach': analysis.get('recommended_approach', {})
                })
        
        # Display roadmap by priority
        phase = 1
        for priority in ['HIGH', 'MEDIUM', 'LOW']:
            sheets = priority_groups[priority]
            if sheets:
                print(f"\nPHASE {phase}: {priority} PRIORITY ({len(sheets)} sheets)")
                print("-" * 50)
                
                for sheet_info in sheets:
                    print(f"📋 {sheet_info['sheet']}")
                    print(f"   Type: {sheet_info['data_type']}")
                    print(f"   Est. Records: {sheet_info['estimated_records']}")
                    print(f"   Approach: {sheet_info['approach'].get('approach', 'TBD')}")
                    print(f"   Table: {sheet_info['approach'].get('table_design', 'TBD')}")
                    print()
                
                phase += 1
        
        # Summary statistics
        total_sheets = sum(len(sheets) for sheets in priority_groups.values())
        total_records = sum(
            sheet_info['estimated_records'] 
            for sheets in priority_groups.values() 
            for sheet_info in sheets
        )
        
        print(f"MIGRATION SUMMARY:")
        print(f"  Sheets to migrate: {total_sheets}")
        print(f"  Estimated total records: {total_records:,}")
        print(f"  High priority sheets: {len(priority_groups['HIGH'])}")
        print(f"  Medium priority sheets: {len(priority_groups['MEDIUM'])}")
        print(f"  Low priority sheets: {len(priority_groups['LOW'])}")

def main():
    """Execute comprehensive worksheet analysis"""
    
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        print(f"Excel file not found: {excel_file}")
        return
    
    analyzer = WorksheetAnalyzer(excel_file)
    
    # Generate analysis
    results = analyzer.generate_comprehensive_analysis()
    
    # Create migration roadmap
    analyzer.create_migration_roadmap()
    
    # Save results
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results_file = f"worksheet_analysis_{timestamp}.json"
    
    with open(results_file, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    
    print(f"\n\nDetailed analysis saved to: {results_file}")

if __name__ == "__main__":
    main()