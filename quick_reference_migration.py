#!/usr/bin/env python3
"""
Quick Reference Table Migration - Focus on efficiency
"""

import logging
import sys
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal
import json

import openpyxl
from sqlalchemy import create_engine, Column, Integer, String, DECIMAL, Date, DateTime, JSON as SQLJSON, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

Base = declarative_base()

class ReferenceData(Base):
    """Simple table for reference data"""
    __tablename__ = 'reference_data'
    
    id = Column(Integer, primary_key=True)
    section_name = Column(String(100))
    row_number = Column(Integer)
    correlation_date = Column(Date)
    raw_data = Column(SQLJSON)
    created_at = Column(DateTime, default=datetime.now)

def quick_migration():
    """Quick migration focusing on the main data section"""
    
    excel_file = "TradeHypoPrelimv32.xlsm"
    database_url = "sqlite:///clo_reference_quick.db"
    
    if not Path(excel_file).exists():
        print(f"Excel file not found: {excel_file}")
        return
    
    print("QUICK REFERENCE TABLE MIGRATION")
    print("=" * 50)
    
    # Setup database
    engine = create_engine(database_url)
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    session = SessionLocal()
    
    try:
        # Load workbook
        print("Loading Excel workbook...")
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = workbook['Reference Table']
        
        print(f"Sheet dimensions: {sheet.max_column} columns × {sheet.max_row} rows")
        
        # Focus on the main data section (rows 12-7751 based on previous analysis)
        start_row = 12
        end_row = min(7751, sheet.max_row)
        total_rows = end_row - start_row + 1
        
        print(f"Processing {total_rows} rows from S&P Rating Migration Correlation section")
        
        # Process in batches
        batch_size = 500
        processed_count = 0
        
        for batch_start in range(start_row, end_row + 1, batch_size):
            batch_end = min(batch_start + batch_size - 1, end_row)
            
            print(f"Processing batch: rows {batch_start}-{batch_end}")
            
            batch_records = []
            
            for row_num in range(batch_start, batch_end + 1):
                # Extract first 10 columns for efficiency
                row_data = {}
                has_data = False
                
                for col in range(1, min(11, sheet.max_column + 1)):
                    value = sheet.cell(row=row_num, column=col).value
                    if value is not None:
                        has_data = True
                        if isinstance(value, (int, float)):
                            row_data[f'col_{col}'] = float(value)
                        elif isinstance(value, (date, datetime)):
                            row_data[f'col_{col}'] = value.isoformat()
                        else:
                            row_data[f'col_{col}'] = str(value)
                
                if has_data:
                    # Try to parse date from column 2
                    correlation_date = None
                    col2_value = sheet.cell(row=row_num, column=2).value
                    if isinstance(col2_value, (date, datetime)):
                        correlation_date = col2_value.date() if isinstance(col2_value, datetime) else col2_value
                    
                    record = ReferenceData(
                        section_name='SP_Rating_Migration_Correlation',
                        row_number=row_num,
                        correlation_date=correlation_date,
                        raw_data=row_data
                    )
                    batch_records.append(record)
                
                processed_count += 1
                
                if processed_count % 1000 == 0:
                    print(f"  Processed {processed_count}/{total_rows} rows...")
            
            # Save batch to database
            if batch_records:
                session.add_all(batch_records)
                session.commit()
                print(f"  Saved {len(batch_records)} records to database")
        
        # Verify results
        total_records = session.query(ReferenceData).count()
        records_with_dates = session.query(ReferenceData).filter(
            ReferenceData.correlation_date.isnot(None)
        ).count()
        
        print(f"\nMIGRATION COMPLETE:")
        print(f"  Total records: {total_records}")
        print(f"  Records with dates: {records_with_dates}")
        print(f"  Database file: clo_reference_quick.db")
        
        # Sample some data
        print(f"\nSample records:")
        sample_records = session.query(ReferenceData).limit(5).all()
        for record in sample_records:
            print(f"  Row {record.row_number}: Date {record.correlation_date}, Data keys: {list(record.raw_data.keys()) if record.raw_data else 'None'}")
        
        return total_records
        
    except Exception as e:
        print(f"Error during migration: {e}")
        logger.error(f"Migration error: {e}")
        return 0
        
    finally:
        session.close()

if __name__ == "__main__":
    result = quick_migration()
    print(f"\nMigration result: {result} records processed")