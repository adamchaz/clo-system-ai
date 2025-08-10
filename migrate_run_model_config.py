#!/usr/bin/env python3
"""
Run Model Configuration Migration
Migrate Run Model and Run Model_old sheets to model_parameters table
"""

import logging
import sys
import traceback
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import json

# Excel processing
import openpyxl

# Database
from sqlalchemy import create_engine, Column, Integer, String, Text, DateTime, Index, text
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.declarative import declarative_base

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('run_model_migration.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

Base = declarative_base()

class ModelParameter(Base):
    """Model configuration parameters storage"""
    __tablename__ = 'model_parameters'
    
    id = Column(Integer, primary_key=True)
    config_name = Column(String(50), nullable=False)  # 'Run Model' or 'Run Model_old'
    section_name = Column(String(100))  # Section within configuration
    parameter_name = Column(String(200), nullable=False)
    parameter_value = Column(Text)  # Store as text for flexibility
    parameter_type = Column(String(20))  # 'path', 'numeric', 'text', 'boolean'
    description = Column(Text)  # Parameter description if available
    row_number = Column(Integer)
    column_number = Column(Integer)
    is_active = Column(String(10), default='active')  # 'active', 'legacy'
    data_source = Column(String(50), default='Run_Model')
    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)
    
    # Indexes for efficient lookups
    __table_args__ = (
        Index('idx_model_parameters_config', 'config_name'),
        Index('idx_model_parameters_parameter', 'parameter_name'),
        Index('idx_model_parameters_active', 'is_active'),
    )

class RunModelMigrator:
    """Migrator for Run Model configuration sheets"""
    
    def __init__(self, excel_path: str, database_url: str):
        self.excel_path = Path(excel_path)
        self.database_url = database_url
        
        # Configuration sheets to migrate
        self.config_sheets = [
            {'name': 'Run Model', 'status': 'active'},
            {'name': 'Run Model_old', 'status': 'legacy'}
        ]
        
        # Migration statistics
        self.stats = {
            'start_time': datetime.now(),
            'end_time': None,
            'sheets_processed': 0,
            'total_parameters': 0,
            'successful_loads': 0,
            'failed_loads': 0
        }
        
        # Error tracking
        self.errors = {
            'sheet_errors': [],
            'extraction_errors': [],
            'loading_errors': []
        }
        
        # Initialize database
        self.engine = create_engine(database_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        self.db_session = SessionLocal()
    
    def setup_database(self):
        """Create database tables"""
        try:
            Base.metadata.create_all(bind=self.engine)
            logger.info("Model parameter tables created successfully")
            return True
        except Exception as e:
            logger.error(f"Error creating database tables: {e}")
            return False
    
    def extract_model_config(self, sheet_info: Dict[str, str]) -> List[Dict[str, Any]]:
        """Extract model configuration from a Run Model sheet"""
        sheet_name = sheet_info['name']
        status = sheet_info['status']
        
        logger.info(f"Processing configuration sheet: {sheet_name}")
        
        try:
            workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            if sheet_name not in workbook.sheetnames:
                self.errors['sheet_errors'].append(f"Sheet not found: {sheet_name}")
                return []
            
            sheet = workbook[sheet_name]
            config_data = []
            
            current_section = None
            
            # Process all cells in the sheet
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    
                    if cell.value is not None:
                        value = cell.value
                        
                        # Determine if this is a section header
                        if self._is_section_header(row, col, value, sheet):
                            current_section = str(value).strip()
                            logger.debug(f"Found section header: {current_section}")
                            continue
                        
                        # Determine parameter type and clean value
                        param_type, cleaned_value = self._classify_and_clean_config_value(value)
                        
                        if cleaned_value is not None:
                            # Generate parameter name and description
                            param_name, description = self._generate_config_parameter_name(
                                row, col, value, current_section, sheet
                            )
                            
                            parameter = {
                                'config_name': sheet_name,
                                'section_name': current_section,
                                'parameter_name': param_name,
                                'parameter_value': cleaned_value,
                                'parameter_type': param_type,
                                'description': description,
                                'row_number': row,
                                'column_number': col,
                                'is_active': status,
                                'data_source': sheet_name.replace(' ', '_')
                            }
                            
                            config_data.append(parameter)
            
            logger.info(f"Extracted {len(config_data)} parameters from {sheet_name}")
            return config_data
            
        except Exception as e:
            error_msg = f"Error processing {sheet_name}: {e}"
            logger.error(error_msg)
            self.errors['extraction_errors'].append(error_msg)
            return []
    
    def _is_section_header(self, row: int, col: int, value: Any, sheet) -> bool:
        """Determine if a cell contains a section header"""
        if not isinstance(value, str):
            return False
        
        value_str = str(value).strip().lower()
        
        # Check if it spans multiple columns (merged cell indicator)
        # Check if it's followed by empty cells or different content structure
        # Check for common configuration section patterns
        
        section_patterns = [
            'clo reinvestment model',
            'table of contents',
            'model inputs',
            'parameters',
            'configuration',
            'settings',
            'paths',
            'file locations'
        ]
        
        # Strong section header indicators
        if any(pattern in value_str for pattern in section_patterns):
            return True
        
        # Check if this appears to be a standalone descriptive text
        if (len(value_str) > 20 and 
            col == 1 and 
            not value_str.replace('.', '').replace('\\', '').replace(':', '').isdigit()):
            # Check if the next cell is empty or very different
            next_cell = sheet.cell(row=row, column=col + 1).value
            if next_cell is None or str(next_cell).strip() == '':
                return True
        
        return False
    
    def _classify_and_clean_config_value(self, value: Any) -> tuple[str, Optional[str]]:
        """Classify parameter type and clean value for model configuration"""
        if value is None:
            return 'empty', None
        
        try:
            # Handle datetime values
            if hasattr(value, 'date'):
                return 'date', value.strftime('%Y-%m-%d')
            
            # Handle numeric values
            if isinstance(value, (int, float)):
                return 'numeric', str(value)
            
            # Handle string values
            elif isinstance(value, str):
                cleaned = str(value).strip()
                
                if not cleaned:
                    return 'empty', None
                
                # Check for file paths
                if ('\\' in cleaned or '/' in cleaned) and ('.' in cleaned or 'Users' in cleaned or 'Dropbox' in cleaned):
                    return 'path', cleaned
                
                # Check for boolean values
                if cleaned.lower() in ['true', 'false', 'yes', 'no', 'on', 'off']:
                    return 'boolean', cleaned.lower()
                
                # Check if it's numeric string
                try:
                    float(cleaned)
                    return 'numeric', cleaned
                except ValueError:
                    pass
                
                # Default to text
                return 'text', cleaned
            
            # Default case
            else:
                return 'text', str(value)
                
        except Exception as e:
            logger.warning(f"Error classifying config value {value}: {e}")
            return 'text', str(value) if value is not None else None
    
    def _generate_config_parameter_name(self, row: int, col: int, value: Any, section: Optional[str], sheet) -> tuple[str, Optional[str]]:
        """Generate descriptive parameter name and description"""
        base_name = f"R{row}C{col}"
        description = None
        
        # Look for parameter labels in adjacent cells
        label_cell = None
        
        # Check left cell for parameter label (common pattern: Label | Value)
        if col > 1:
            left_cell = sheet.cell(row=row, column=col - 1).value
            if left_cell and isinstance(left_cell, str) and len(str(left_cell).strip()) > 0:
                label_cell = str(left_cell).strip()
        
        # Check cell above for parameter label (common pattern: Label on top, Value below)
        if not label_cell and row > 1:
            above_cell = sheet.cell(row=row - 1, column=col).value
            if above_cell and isinstance(above_cell, str) and len(str(above_cell).strip()) > 0:
                # Make sure it's not a section header itself
                if not self._is_section_header(row - 1, col, above_cell, sheet):
                    label_cell = str(above_cell).strip()
        
        # Generate parameter name
        if label_cell:
            # Clean the label for parameter name
            clean_label = label_cell.replace(' ', '_').replace(':', '').replace(',', '').replace('(', '').replace(')', '')
            clean_label = ''.join(c for c in clean_label if c.isalnum() or c in ['_', '-']).strip('_')
            
            if clean_label:
                if section:
                    section_clean = section.replace(' ', '_').replace(':', '').replace(',', '')
                    section_clean = ''.join(c for c in section_clean if c.isalnum() or c in ['_', '-']).strip('_')
                    param_name = f"{section_clean}_{clean_label}_{base_name}"
                else:
                    param_name = f"{clean_label}_{base_name}"
                
                description = f"Parameter: {label_cell}"
            else:
                param_name = base_name
        else:
            # Use the value itself to generate name if it's descriptive text
            if isinstance(value, str) and len(str(value).strip()) > 0:
                value_clean = str(value)[:30].replace(' ', '_').replace(':', '').replace(',', '')
                value_clean = ''.join(c for c in value_clean if c.isalnum() or c in ['_', '-']).strip('_')
                
                if value_clean and not value_clean.replace('.', '').replace('\\', '').isdigit():
                    param_name = f"{value_clean}_{base_name}"
                    description = f"Value: {str(value)[:100]}"
                else:
                    param_name = base_name
            else:
                param_name = base_name
        
        # Add section prefix if available
        if section and not param_name.startswith(section.replace(' ', '_')):
            section_clean = section.replace(' ', '_').replace(':', '').replace(',', '')
            section_clean = ''.join(c for c in section_clean if c.isalnum() or c in ['_', '-']).strip('_')
            if section_clean:
                param_name = f"{section_clean}_{param_name}"
        
        return param_name[:200], description  # Limit to database column size
    
    def load_config_data(self, config_data: List[Dict[str, Any]]) -> Dict[str, int]:
        """Load configuration data to database"""
        if not config_data:
            return {'successful': 0, 'failed': 0}
        
        logger.info(f"Loading {len(config_data)} model parameters to database...")
        
        batch_size = 100
        successful_loads = 0
        failed_loads = 0
        
        # Load in batches
        for i in range(0, len(config_data), batch_size):
            batch = config_data[i:i + batch_size]
            batch_num = i // batch_size + 1
            total_batches = (len(config_data) + batch_size - 1) // batch_size
            
            logger.info(f"Loading batch {batch_num}/{total_batches}: {len(batch)} parameters")
            
            try:
                # Create parameter objects
                parameter_objects = []
                for param_data in batch:
                    parameter = ModelParameter(
                        config_name=param_data['config_name'],
                        section_name=param_data['section_name'],
                        parameter_name=param_data['parameter_name'],
                        parameter_value=param_data['parameter_value'],
                        parameter_type=param_data['parameter_type'],
                        description=param_data.get('description'),
                        row_number=param_data['row_number'],
                        column_number=param_data['column_number'],
                        is_active=param_data['is_active'],
                        data_source=param_data['data_source']
                    )
                    parameter_objects.append(parameter)
                
                # Bulk insert batch
                self.db_session.add_all(parameter_objects)
                self.db_session.commit()
                successful_loads += len(parameter_objects)
                logger.info(f"Successfully loaded batch {batch_num}")
                
            except SQLAlchemyError as e:
                logger.error(f"Database error loading batch {batch_num}: {e}")
                self.db_session.rollback()
                failed_loads += len(batch)
                self.errors['loading_errors'].append(f"Batch {batch_num}: {e}")
        
        return {
            'successful': successful_loads,
            'failed': failed_loads,
            'total_processed': successful_loads + failed_loads
        }
    
    def validate_loaded_data(self) -> Dict[str, Any]:
        """Validate loaded configuration data"""
        logger.info("Validating loaded model configuration data...")
        
        try:
            # Count total parameters
            total_count = self.db_session.query(ModelParameter).count()
            
            # Count by configuration
            config_counts = self.db_session.execute(text("""
                SELECT config_name, is_active, COUNT(*) as count
                FROM model_parameters
                GROUP BY config_name, is_active
                ORDER BY config_name, is_active
            """)).fetchall()
            
            # Count by parameter type
            type_counts = self.db_session.execute(text("""
                SELECT parameter_type, COUNT(*) as count
                FROM model_parameters
                GROUP BY parameter_type
                ORDER BY count DESC
            """)).fetchall()
            
            # Sample parameters
            sample_params = self.db_session.execute(text("""
                SELECT config_name, section_name, parameter_name, parameter_value, parameter_type, description
                FROM model_parameters
                WHERE description IS NOT NULL
                LIMIT 10
            """)).fetchall()
            
            validation_results = {
                'total_parameters': total_count,
                'config_counts': [{'config': row[0], 'status': row[1], 'count': row[2]} for row in config_counts],
                'type_counts': [{'type': row[0], 'count': row[1]} for row in type_counts],
                'sample_parameters': [
                    {
                        'config': row[0],
                        'section': row[1],
                        'parameter': row[2],
                        'value': row[3],
                        'type': row[4],
                        'description': row[5]
                    }
                    for row in sample_params
                ]
            }
            
            logger.info("Validation results:")
            logger.info(f"  Total parameters: {validation_results['total_parameters']}")
            
            for config in validation_results['config_counts']:
                logger.info(f"    {config['config']} ({config['status']}): {config['count']} parameters")
            
            return validation_results
            
        except Exception as e:
            logger.error(f"Error validating data: {e}")
            return {'error': str(e)}
    
    def execute_full_migration(self) -> Dict[str, Any]:
        """Execute complete Run Model configuration migration"""
        logger.info("Starting Run Model Configuration Migration")
        
        try:
            # Setup database
            if not self.setup_database():
                return {'success': False, 'error': 'Database setup failed'}
            
            # Clear existing configuration data
            try:
                self.db_session.execute(text("DELETE FROM model_parameters"))
                self.db_session.commit()
                logger.info("Cleared existing model parameter data")
            except Exception as e:
                logger.warning(f"Could not clear existing data: {e}")
            
            # Process each configuration sheet
            all_config_data = []
            
            for sheet_info in self.config_sheets:
                logger.info(f"Processing configuration sheet: {sheet_info['name']}")
                sheet_data = self.extract_model_config(sheet_info)
                
                if sheet_data:
                    all_config_data.extend(sheet_data)
                    self.stats['sheets_processed'] += 1
            
            self.stats['total_parameters'] = len(all_config_data)
            
            if not all_config_data:
                return {'success': False, 'error': 'No configuration data extracted'}
            
            # Load to database
            load_results = self.load_config_data(all_config_data)
            self.stats['successful_loads'] = load_results['successful']
            self.stats['failed_loads'] = load_results['failed']
            
            # Validate loaded data
            validation_results = self.validate_loaded_data()
            
            # Generate report
            report = self.generate_migration_report()
            
            return {
                'success': load_results['successful'] > 0,
                'statistics': self.stats,
                'load_results': load_results,
                'validation_results': validation_results,
                'errors': self.errors,
                'report': report
            }
            
        except Exception as e:
            logger.error(f"Fatal error during migration: {e}")
            logger.error(traceback.format_exc())
            
            return {
                'success': False,
                'error': str(e),
                'statistics': self.stats,
                'errors': self.errors
            }
        
        finally:
            self.db_session.close()
    
    def generate_migration_report(self) -> str:
        """Generate comprehensive migration report"""
        self.stats['end_time'] = datetime.now()
        duration = (self.stats['end_time'] - self.stats['start_time']).total_seconds()
        
        report = []
        report.append("=" * 80)
        report.append("RUN MODEL CONFIGURATION MIGRATION REPORT")
        report.append("=" * 80)
        
        report.append(f"\nMigration Summary:")
        report.append(f"  Start Time: {self.stats['start_time']}")
        report.append(f"  End Time: {self.stats['end_time']}")
        report.append(f"  Duration: {duration:.1f} seconds")
        
        report.append(f"\nProcessing Statistics:")
        report.append(f"  Sheets Processed: {self.stats['sheets_processed']}/{len(self.config_sheets)}")
        report.append(f"  Total Parameters Extracted: {self.stats['total_parameters']}")
        report.append(f"  Successfully Loaded: {self.stats['successful_loads']}")
        report.append(f"  Failed Loads: {self.stats['failed_loads']}")
        
        # Success rate
        if self.stats['total_parameters'] > 0:
            success_rate = (self.stats['successful_loads'] / self.stats['total_parameters']) * 100
            report.append(f"  Success Rate: {success_rate:.1f}%")
        
        # Error summary
        total_errors = sum(len(errors) for errors in self.errors.values())
        report.append(f"\nError Summary:")
        report.append(f"  Total Errors: {total_errors}")
        
        # Migration status
        report.append(f"\nMigration Status:")
        if self.stats['successful_loads'] > 0:
            report.append(f"  SUCCESS - Model configuration migrated successfully")
            report.append(f"  {self.stats['sheets_processed']} configuration sheets processed")
            report.append(f"  {self.stats['successful_loads']} parameters available for model execution")
        else:
            report.append(f"  FAILED - No parameters migrated successfully")
        
        report.append("=" * 80)
        
        return "\n".join(report)

def main():
    """Execute Run Model configuration migration"""
    
    excel_file = "TradeHypoPrelimv32.xlsm"
    database_url = "sqlite:///clo_model_config.db"
    
    if not Path(excel_file).exists():
        print(f"Error: Excel file not found: {excel_file}")
        return 1
    
    print("Starting Run Model Configuration Migration")
    print("=" * 60)
    
    # Execute migration
    migrator = RunModelMigrator(excel_file, database_url)
    results = migrator.execute_full_migration()
    
    # Display results
    if results['success']:
        print("Migration completed successfully!")
        print(f"Parameters migrated: {results['statistics']['successful_loads']}")
        print(f"Sheets processed: {results['statistics']['sheets_processed']}")
    else:
        print("Migration failed!")
        if 'error' in results:
            print(f"Error: {results['error']}")
    
    # Display report
    if 'report' in results:
        print("\n" + results['report'])
    
    # Save detailed results
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results_file = f"run_model_migration_results_{timestamp}.json"
    
    with open(results_file, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    
    print(f"\nDetailed results saved to: {results_file}")
    print(f"Database created: clo_model_config.db")
    
    return 0 if results['success'] else 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)