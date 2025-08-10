#!/usr/bin/env python3
"""
Identify and analyze different sections within the Reference Table
"""

import openpyxl
from pathlib import Path
from datetime import datetime
import re

def analyze_reference_sections():
    """Identify distinct sections within the Reference Table"""
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        print(f"Excel file not found: {excel_file}")
        return
    
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook['Reference Table']
    
    print("REFERENCE TABLE SECTION ANALYSIS")
    print("=" * 60)
    
    # Strategy: Look for section identifiers in column 1 and analyze the data structure that follows
    sections = {}
    current_section = None
    section_order = []
    
    # First pass: identify section boundaries
    for row_num in range(1, sheet.max_row + 1):
        col1_value = sheet.cell(row=row_num, column=1).value
        
        if col1_value and str(col1_value).strip():
            section_name = str(col1_value).strip()
            
            # Skip generic headers
            if section_name not in ['Column_1', 'Table Of Contents']:
                # This looks like a section identifier
                if current_section:
                    # End the previous section
                    sections[current_section]['end_row'] = row_num - 1
                
                # Start new section
                current_section = section_name
                sections[current_section] = {
                    'name': section_name,
                    'start_row': row_num,
                    'end_row': None,
                    'data_rows': [],
                    'structure': 'unknown'
                }
                section_order.append(current_section)
    
    # Close the last section
    if current_section:
        sections[current_section]['end_row'] = sheet.max_row
    
    print(f"IDENTIFIED SECTIONS: {len(sections)}")
    for i, section_name in enumerate(section_order, 1):
        section = sections[section_name]
        row_count = section['end_row'] - section['start_row'] + 1
        print(f"{i:2d}. {section_name[:50]:50} | Rows {section['start_row']}-{section['end_row']} ({row_count} rows)")
    
    # Second pass: analyze each section's structure
    print(f"\nSECTION STRUCTURE ANALYSIS:")
    print("-" * 60)
    
    for section_name in section_order:
        section = sections[section_name]
        print(f"\n{section_name}:")
        
        start_row = section['start_row'] + 1  # Skip the section header
        end_row = min(section['end_row'], section['start_row'] + 20)  # Analyze first 20 rows
        
        # Analyze data pattern
        data_rows = []
        headers_identified = False
        potential_headers = []
        
        for row_num in range(start_row, end_row + 1):
            row_data = []
            non_empty_count = 0
            
            for col in range(1, min(20, sheet.max_column + 1)):  # Check first 20 columns
                value = sheet.cell(row=row_num, column=col).value
                row_data.append(value)
                if value is not None and str(value).strip():
                    non_empty_count += 1
            
            if non_empty_count > 0:
                if not headers_identified and non_empty_count >= 3:
                    # This might be headers
                    potential_headers = [str(v) if v else f"Col{i+1}" for i, v in enumerate(row_data)]
                    headers_identified = True
                    print(f"  Potential Headers: {', '.join([h for h in potential_headers[:5] if h])}")
                
                data_rows.append({
                    'row': row_num,
                    'data': row_data,
                    'non_empty': non_empty_count
                })
        
        # Determine section type based on content analysis
        section_type = "unknown"
        
        if 'libor' in section_name.lower():
            section_type = "interest_rates"
        elif 'transition' in section_name.lower() or 'migration' in section_name.lower():
            section_type = "rating_transitions"
        elif 'holiday' in section_name.lower():
            section_type = "business_calendar"
        elif 'yield' in section_name.lower() or 'curve' in section_name.lower():
            section_type = "yield_curve"
        elif 'cashflow' in section_name.lower():
            section_type = "cashflow_data"
        elif 'deal' in section_name.lower():
            section_type = "deal_reference"
        elif 'sink' in section_name.lower():
            section_type = "amortization_schedule"
        
        sections[section_name]['type'] = section_type
        sections[section_name]['data_sample'] = data_rows[:5]
        
        print(f"  Type: {section_type}")
        print(f"  Data rows analyzed: {len(data_rows)}")
        
        if data_rows:
            print(f"  Sample data (first 3 rows):")
            for data_row in data_rows[:3]:
                sample_values = [str(v)[:15] if v else '' for v in data_row['data'][:4]]
                print(f"    Row {data_row['row']}: {' | '.join(sample_values)}")
    
    # Migration strategy for each section type
    print(f"\nMIGRATION STRATEGY BY SECTION TYPE:")
    print("=" * 60)
    
    type_sections = {}
    for section_name, section_data in sections.items():
        section_type = section_data['type']
        if section_type not in type_sections:
            type_sections[section_type] = []
        type_sections[section_type].append(section_name)
    
    for section_type, section_names in type_sections.items():
        print(f"\n{section_type.upper()} ({len(section_names)} sections):")
        
        if section_type == "interest_rates":
            print(f"  Database Table: historical_rates")
            print(f"  Schema: rate_date, rate_type, rate_value, source")
            print(f"  Purpose: Store LIBOR and other benchmark rates")
            print(f"  Sections: {', '.join(section_names)}")
        
        elif section_type == "rating_transitions":
            print(f"  Database Table: rating_transition_matrices")
            print(f"  Schema: agency, from_rating, to_rating, probability, time_period")
            print(f"  Purpose: Credit migration modeling")
            print(f"  Sections: {', '.join(section_names)}")
        
        elif section_type == "business_calendar":
            print(f"  Database Table: business_holidays")
            print(f"  Schema: holiday_date, holiday_name, country, market")
            print(f"  Purpose: Business day calculations")
            print(f"  Sections: {', '.join(section_names)}")
        
        elif section_type == "yield_curve":
            print(f"  Database Table: yield_curves")
            print(f"  Schema: curve_date, tenor, rate, curve_type")
            print(f"  Purpose: Interest rate curve data for pricing")
            print(f"  Sections: {', '.join(section_names)}")
        
        elif section_type == "cashflow_data":
            print(f"  Database Table: reference_cashflows")
            print(f"  Schema: asset_id, payment_date, principal, interest, type")
            print(f"  Purpose: Expected cashflow projections")
            print(f"  Sections: {', '.join(section_names)}")
        
        elif section_type == "deal_reference":
            print(f"  Database Table: deal_parameters")
            print(f"  Schema: deal_name, parameter_name, parameter_value, effective_date")
            print(f"  Purpose: Deal-specific reference parameters")
            print(f"  Sections: {', '.join(section_names)}")
        
        elif section_type == "amortization_schedule":
            print(f"  Database Table: amortization_schedules")
            print(f"  Schema: schedule_id, payment_date, principal_amount, balance")
            print(f"  Purpose: Scheduled principal payments")
            print(f"  Sections: {', '.join(section_names)}")
        
        else:
            print(f"  Database Table: {section_type}_reference")
            print(f"  Schema: [To be determined based on detailed analysis]")
            print(f"  Purpose: General reference data")
            print(f"  Sections: {', '.join(section_names)}")
    
    # Overall migration approach
    print(f"\nOVERALL MIGRATION APPROACH:")
    print("-" * 40)
    print(f"1. Create specialized tables for each reference data type")
    print(f"2. Normalize data to eliminate redundancy")
    print(f"3. Establish proper indexing for lookup performance")
    print(f"4. Create foreign key relationships where applicable")
    print(f"5. Implement data validation and integrity constraints")
    print(f"6. Consider temporal aspects (effective dates, versioning)")
    
    return sections

if __name__ == "__main__":
    sections = analyze_reference_sections()