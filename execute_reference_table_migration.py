#!/usr/bin/env python3
"""
Execute Reference Table Migration
Focus on S&P Rating Migration Correlation data (7,740 rows)
"""

import json
import logging
import sys
import traceback
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from typing import Dict, List, Any, Optional, Union
import pandas as pd
import numpy as np

# Excel processing
import openpyxl

# Database
from sqlalchemy import create_engine, text, MetaData, Table, Column, Integer, String, DECIMAL, Date, DateTime, Boolean, JSON as SQLJSON
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
from sqlalchemy.ext.declarative import declarative_base

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('reference_table_migration.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Create base for models
Base = declarative_base()

class RatingTransitionMatrix(Base):
    """S&P Rating Transition Matrix data"""
    __tablename__ = 'rating_transition_matrices'
    
    id = Column(Integer, primary_key=True)
    agency = Column(String(10), default='SP')
    correlation_date = Column(Date)
    from_rating = Column(String(20))
    to_rating = Column(String(20))
    correlation_value = Column(DECIMAL(10, 6))
    time_period = Column(String(20), default='1_year')
    data_source = Column(String(50), default='Reference_Table')
    raw_data = Column(SQLJSON)  # Store original row data
    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)

class HistoricalRates(Base):
    """Historical interest rates"""
    __tablename__ = 'historical_rates'
    
    id = Column(Integer, primary_key=True)
    rate_date = Column(Date, nullable=False)
    rate_type = Column(String(50), nullable=False)
    rate_value = Column(DECIMAL(10, 6), nullable=False)
    source = Column(String(20))
    created_at = Column(DateTime, default=datetime.now)

class BusinessHolidays(Base):
    """Business holidays calendar"""
    __tablename__ = 'business_holidays'
    
    id = Column(Integer, primary_key=True)
    holiday_date = Column(Date, nullable=False)
    holiday_name = Column(String(100))
    country = Column(String(10), default='US')
    market = Column(String(20), default='USD')
    created_at = Column(DateTime, default=datetime.now)

class ReferenceTableMigrator:
    """Migrator for Reference Table data"""
    
    def __init__(self, excel_path: str, database_url: str):
        self.excel_path = Path(excel_path)
        self.database_url = database_url
        
        # Migration statistics
        self.stats = {
            'start_time': datetime.now(),
            'end_time': None,
            'extracted_count': 0,
            'transformed_count': 0,
            'loaded_count': 0
        }
        
        # Error tracking
        self.errors = {
            'extraction_errors': [],
            'transformation_errors': [],
            'loading_errors': []
        }
        
        # Initialize database
        self.engine = create_engine(database_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        self.db_session = SessionLocal()
        
        # Section configurations
        self.sections = {
            'S&P Rating Migration Correlation': {
                'start_row': 12,
                'model_class': RatingTransitionMatrix,
                'processor': self._process_rating_correlation_data
            }
        }
    
    def setup_database(self):
        """Create database tables"""
        try:
            Base.metadata.create_all(bind=self.engine)
            logger.info("Database tables created successfully")
            return True
        except Exception as e:
            logger.error(f"Error creating database tables: {e}")
            return False
    
    def extract_section_data(self, section_name: str) -> List[Dict[str, Any]]:
        """Extract data from a specific section"""
        logger.info(f"Extracting data from section: {section_name}")
        
        try:
            workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            sheet = workbook['Reference Table']
            
            if section_name not in self.sections:
                raise ValueError(f"Section {section_name} not configured")
            
            section_config = self.sections[section_name]
            start_row = section_config['start_row']
            
            # Extract data from start_row to end of meaningful data
            extracted_data = []
            
            # First, determine the actual data range
            last_data_row = start_row
            for row_num in range(start_row, sheet.max_row + 1):
                has_data = False
                for col in range(1, min(20, sheet.max_column + 1)):
                    value = sheet.cell(row=row_num, column=col).value
                    if value is not None and str(value).strip():
                        has_data = True
                        break
                if has_data:
                    last_data_row = row_num
                else:
                    # Check if we've had 10 consecutive empty rows
                    empty_count = 0
                    for check_row in range(row_num, min(row_num + 10, sheet.max_row + 1)):
                        row_empty = True
                        for col in range(1, min(20, sheet.max_column + 1)):
                            if sheet.cell(row=check_row, column=col).value:
                                row_empty = False
                                break
                        if not row_empty:
                            break
                        empty_count += 1
                    if empty_count >= 10:
                        break
            
            logger.info(f"Data range: rows {start_row} to {last_data_row}")
            
            # Extract the actual data
            for row_num in range(start_row, last_data_row + 1):
                row_data = {}
                has_data = False
                
                # Extract all columns for this row
                for col in range(1, sheet.max_column + 1):
                    value = sheet.cell(row=row_num, column=col).value
                    row_data[f'col_{col}'] = value
                    if value is not None and str(value).strip():
                        has_data = True
                
                if has_data:
                    row_data['_source_row'] = row_num
                    extracted_data.append(row_data)
            
            self.stats['extracted_count'] = len(extracted_data)
            logger.info(f"Extracted {len(extracted_data)} rows from {section_name}")
            
            return extracted_data
            
        except Exception as e:
            logger.error(f"Error extracting section {section_name}: {e}")
            logger.error(traceback.format_exc())
            self.errors['extraction_errors'].append(f"{section_name}: {e}")
            return []
    
    def _process_rating_correlation_data(self, raw_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Process S&P Rating Migration Correlation data"""
        logger.info(f"Processing {len(raw_data)} rating correlation rows")
        
        processed_data = []
        
        for row_data in raw_data:
            try:
                # Extract key fields (adjust based on actual column structure)
                correlation_date = self._parse_date_value(row_data.get('col_2'))
                day_of_week = row_data.get('col_3')
                rate_info = row_data.get('col_4')
                
                # Create processed record
                processed_record = {
                    'agency': 'SP',
                    'correlation_date': correlation_date,
                    'data_source': 'Reference_Table',
                    'time_period': '1_year',
                    'raw_data': {
                        'day_of_week': day_of_week,
                        'rate_info': str(rate_info) if rate_info else None,
                        'source_row': row_data.get('_source_row')
                    }
                }
                
                # Try to extract correlation values from numeric columns
                correlation_values = {}
                for col_key, value in row_data.items():
                    if col_key.startswith('col_') and isinstance(value, (int, float)):
                        col_num = col_key.replace('col_', '')
                        correlation_values[f'column_{col_num}'] = float(value)
                
                if correlation_values:
                    processed_record['raw_data']['correlations'] = correlation_values
                
                processed_data.append(processed_record)
                
            except Exception as e:
                logger.warning(f"Error processing row {row_data.get('_source_row', 'unknown')}: {e}")
                self.errors['transformation_errors'].append(f"Row processing: {e}")
        
        logger.info(f"Processed {len(processed_data)} correlation records")
        return processed_data
    
    def _parse_date_value(self, value: Any) -> Optional[date]:
        """Parse date value from various Excel formats"""
        if pd.isna(value) or value in [None, '', ' ']:
            return None
        
        try:
            # Handle datetime objects
            if isinstance(value, datetime):
                return value.date()
            
            # Handle date objects
            if isinstance(value, date):
                return value
            
            # Handle numeric Excel dates
            if isinstance(value, (int, float)):
                try:
                    # Excel date serial number (days since 1900-01-01)
                    excel_epoch = datetime(1900, 1, 1)
                    parsed_date = excel_epoch + pd.Timedelta(days=value - 2)  # Excel quirk
                    return parsed_date.date()
                except:
                    pass
            
            return None
            
        except Exception:
            return None
    
    def load_section_data(self, section_name: str, processed_data: List[Dict[str, Any]]) -> Dict[str, int]:
        """Load processed data to database"""
        logger.info(f"Loading {len(processed_data)} records to {section_name} table")
        
        section_config = self.sections[section_name]
        model_class = section_config['model_class']
        
        successful_loads = 0
        failed_loads = 0
        
        batch_size = 100
        
        for i in range(0, len(processed_data), batch_size):
            batch = processed_data[i:i + batch_size]
            batch_num = i // batch_size + 1
            total_batches = (len(processed_data) + batch_size - 1) // batch_size
            
            logger.info(f"Loading batch {batch_num}/{total_batches}: {len(batch)} records")
            
            try:
                # Create model objects
                model_objects = []
                for record_data in batch:
                    try:
                        record = model_class(**record_data)
                        model_objects.append(record)
                    except Exception as e:
                        logger.error(f"Error creating model object: {e}")
                        failed_loads += 1
                        self.errors['loading_errors'].append(f"Model creation: {e}")
                
                # Bulk insert batch
                if model_objects:
                    self.db_session.add_all(model_objects)
                    self.db_session.commit()
                    successful_loads += len(model_objects)
                    logger.info(f"Successfully loaded batch {batch_num}")
                
            except SQLAlchemyError as e:
                logger.error(f"Database error loading batch {batch_num}: {e}")
                self.db_session.rollback()
                failed_loads += len(batch)
                self.errors['loading_errors'].append(f"Batch {batch_num}: {e}")
        
        self.stats['loaded_count'] = successful_loads
        
        return {
            'successful': successful_loads,
            'failed': failed_loads,
            'total_processed': successful_loads + failed_loads
        }
    
    def migrate_section(self, section_name: str):
        """Migrate a specific section"""
        logger.info(f"Starting migration for section: {section_name}")
        
        try:
            # Extract data
            raw_data = self.extract_section_data(section_name)
            if not raw_data:
                return {'success': False, 'error': 'No data extracted'}
            
            # Process data
            section_config = self.sections[section_name]
            processor = section_config['processor']
            processed_data = processor(raw_data)
            
            self.stats['transformed_count'] = len(processed_data)
            
            # Load data
            load_results = self.load_section_data(section_name, processed_data)
            
            return {
                'success': load_results['successful'] > 0,
                'extracted': len(raw_data),
                'processed': len(processed_data),
                'loaded': load_results['successful'],
                'failed': load_results['failed']
            }
            
        except Exception as e:
            logger.error(f"Error migrating section {section_name}: {e}")
            return {'success': False, 'error': str(e)}
    
    def generate_migration_report(self) -> str:
        """Generate migration report"""
        self.stats['end_time'] = datetime.now()
        duration = (self.stats['end_time'] - self.stats['start_time']).total_seconds() / 60
        
        report = []
        report.append("=" * 80)
        report.append("REFERENCE TABLE MIGRATION REPORT")
        report.append("=" * 80)
        
        report.append(f"\nMigration Summary:")
        report.append(f"  Start Time: {self.stats['start_time']}")
        report.append(f"  End Time: {self.stats['end_time']}")
        report.append(f"  Duration: {duration:.1f} minutes")
        report.append(f"  Excel File: {self.excel_path}")
        
        report.append(f"\nProcessing Statistics:")
        report.append(f"  Records Extracted: {self.stats['extracted_count']}")
        report.append(f"  Records Transformed: {self.stats['transformed_count']}")
        report.append(f"  Records Loaded: {self.stats['loaded_count']}")
        
        # Success rate
        if self.stats['extracted_count'] > 0:
            success_rate = (self.stats['loaded_count'] / self.stats['extracted_count']) * 100
            report.append(f"  Success Rate: {success_rate:.1f}%")
        
        # Error summary
        total_errors = sum(len(errors) for errors in self.errors.values())
        report.append(f"\nError Summary:")
        report.append(f"  Total Errors: {total_errors}")
        
        report.append("=" * 80)
        
        return "\n".join(report)

def main():
    """Execute Reference Table migration"""
    
    excel_file = "TradeHypoPrelimv32.xlsm"
    database_url = "sqlite:///clo_reference_data.db"
    
    if not Path(excel_file).exists():
        print(f"Error: Excel file not found: {excel_file}")
        return 1
    
    print("Starting Reference Table Migration")
    print("=" * 50)
    
    migrator = ReferenceTableMigrator(excel_file, database_url)
    
    # Setup database
    if not migrator.setup_database():
        print("Failed to setup database")
        return 1
    
    # Migrate main section with data
    results = migrator.migrate_section('S&P Rating Migration Correlation')
    
    # Generate report
    report = migrator.generate_migration_report()
    
    # Display results
    if results['success']:
        print("Migration completed successfully!")
        print(f"Records loaded: {results['loaded']}")
    else:
        print("Migration failed!")
        if 'error' in results:
            print(f"Error: {results['error']}")
    
    print("\n" + report)
    
    # Save results
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results_file = f"reference_table_migration_results_{timestamp}.json"
    
    with open(results_file, 'w') as f:
        json.dump({
            'migration_results': results,
            'statistics': migrator.stats,
            'errors': migrator.errors
        }, f, indent=2, default=str)
    
    print(f"\nDetailed results saved to: {results_file}")
    
    return 0 if results['success'] else 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)