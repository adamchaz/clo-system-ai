#!/usr/bin/env python3
"""
Debug script to identify the B2 issue
"""

import openpyxl
from pathlib import Path

def debug_excel_data():
    """Find the B2 value in the Excel file"""
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        print(f"Excel file not found: {excel_file}")
        return
    
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook['All Assets']
    
    print(f"Searching for 'B2' value in All Assets sheet...")
    
    # Extract headers from row 1
    headers = []
    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=col).value
        headers.append(str(header_value) if header_value else f"Column_{col}")
    
    print(f"\nHeaders: {headers}")
    
    # Search for B2 value
    found_b2 = []
    
    for row_num in range(2, min(50, sheet.max_row + 1)):  # Check first 50 rows
        for col_idx, header in enumerate(headers, 1):
            cell_value = sheet.cell(row=row_num, column=col_idx).value
            if cell_value == 'B2':
                found_b2.append({
                    'row': row_num,
                    'column': col_idx,
                    'header': header,
                    'value': cell_value
                })
                print(f"Found 'B2' at row {row_num}, column {col_idx} ({header})")
    
    # Show context around B2 findings
    if found_b2:
        for finding in found_b2[:3]:  # Show first 3 occurrences
            row = finding['row']
            col = finding['column']
            print(f"\nContext for B2 at row {row}, column {col} ({finding['header']}):")
            
            # Show a few cells around this position
            for c in range(max(1, col-2), min(sheet.max_column + 1, col+3)):
                value = sheet.cell(row=row, column=c).value
                header = headers[c-1]
                print(f"  {header}: {value}")
    else:
        print("No 'B2' values found in first 50 rows")
    
    # Check some sample data types in common numeric fields
    numeric_fields = ['Par Amount', 'Market Value', 'Coupon', 'WAL', 'Facility Size']
    
    print(f"\nSample values in numeric fields:")
    for field_name in numeric_fields:
        if field_name in headers:
            col_idx = headers.index(field_name) + 1
            print(f"\n{field_name} (Column {col_idx}):")
            for row in range(2, min(12, sheet.max_row + 1)):  # Check first 10 data rows
                value = sheet.cell(row=row, column=col_idx).value
                print(f"  Row {row}: {value} (type: {type(value)})")

if __name__ == "__main__":
    debug_excel_data()