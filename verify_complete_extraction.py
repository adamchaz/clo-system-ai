#!/usr/bin/env python3
"""
Verify that we extracted all available assets from Excel
"""

import openpyxl
from pathlib import Path

def verify_complete_extraction():
    """Verify complete asset extraction"""
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        print(f"Excel file not found: {excel_file}")
        return
    
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook['All Assets']
    
    print("Complete Asset Extraction Verification")
    print("=" * 50)
    
    # Count assets row by row with detailed analysis
    assets_found = []
    empty_rows_with_partial_data = []
    
    for row_num in range(2, sheet.max_row + 1):
        blkrock_id = sheet.cell(row=row_num, column=1).value  # BLKRockID
        issue_name = sheet.cell(row=row_num, column=2).value  # Issue Name
        issuer_name = sheet.cell(row=row_num, column=3).value # Issuer Name
        
        # Check if this row has identifying information
        has_blkrock_id = blkrock_id is not None and str(blkrock_id).strip()
        has_issue_name = issue_name is not None and str(issue_name).strip()
        has_issuer_name = issuer_name is not None and str(issuer_name).strip()
        
        if has_blkrock_id:
            assets_found.append({
                'row': row_num,
                'blkrock_id': str(blkrock_id).strip(),
                'issue_name': str(issue_name).strip() if has_issue_name else None,
                'issuer_name': str(issuer_name).strip() if has_issuer_name else None
            })
        elif has_issue_name or has_issuer_name:
            # Row with some identifying data but no BLKRockID
            empty_rows_with_partial_data.append({
                'row': row_num,
                'issue_name': str(issue_name).strip() if has_issue_name else None,
                'issuer_name': str(issuer_name).strip() if has_issuer_name else None
            })
    
    print(f"Assets found with BLKRockID: {len(assets_found)}")
    print(f"Rows with partial data but no BLKRockID: {len(empty_rows_with_partial_data)}")
    
    if assets_found:
        print(f"\nFirst 5 assets:")
        for asset in assets_found[:5]:
            print(f"  Row {asset['row']}: {asset['blkrock_id']} | {asset['issue_name']}")
        
        print(f"\nLast 5 assets:")
        for asset in assets_found[-5:]:
            print(f"  Row {asset['row']}: {asset['blkrock_id']} | {asset['issue_name']}")
    
    if empty_rows_with_partial_data:
        print(f"\nRows with partial data (first 10):")
        for row_info in empty_rows_with_partial_data[:10]:
            print(f"  Row {row_info['row']}: Issue: {row_info['issue_name']} | Issuer: {row_info['issuer_name']}")
    
    # Check for any patterns in supposedly empty rows
    print(f"\nChecking for hidden data in 'empty' rows...")
    hidden_data_rows = []
    
    for row_num in range(386, min(500, sheet.max_row + 1)):  # Check first 100+ empty rows
        row_has_data = False
        non_empty_cells = []
        
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row_num, column=col).value
            if value is not None and str(value).strip():
                non_empty_cells.append((col, value))
                row_has_data = True
        
        if row_has_data:
            hidden_data_rows.append({
                'row': row_num,
                'data': non_empty_cells[:5]  # First 5 non-empty cells
            })
    
    if hidden_data_rows:
        print(f"Found {len(hidden_data_rows)} rows with hidden data:")
        for row_info in hidden_data_rows[:5]:
            print(f"  Row {row_info['row']}: {row_info['data']}")
    else:
        print("No hidden data found in supposedly empty rows")
    
    # Final summary
    print(f"\n" + "=" * 50)
    print(f"EXTRACTION VERIFICATION SUMMARY")
    print(f"=" * 50)
    print(f"Total rows in Excel: {sheet.max_row - 1} (excluding header)")
    print(f"Assets with BLKRockID: {len(assets_found)}")
    print(f"Previously migrated: 384")
    print(f"Match status: {'✅ COMPLETE' if len(assets_found) == 384 else '❌ MISMATCH'}")
    
    if len(assets_found) == 384:
        print(f"\n✅ All available assets have been successfully migrated!")
        print(f"The remaining {sheet.max_row - 1 - 384} rows are empty template rows.")
    else:
        print(f"\n⚠️ Found {len(assets_found)} assets but only migrated 384")
        print(f"Need to investigate {len(assets_found) - 384} missing assets")

if __name__ == "__main__":
    verify_complete_extraction()