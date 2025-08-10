#!/usr/bin/env python3
"""
Asset Correlation Matrix Migration
Migrate the 489x489 correlation matrix from Excel to database
"""

import logging
import sys
import traceback
from pathlib import Path
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from typing import Dict, List, Any, Optional
import json

# Excel processing
import openpyxl

# Database
from sqlalchemy import create_engine, Column, Integer, String, DECIMAL, DateTime, Index, text
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.declarative import declarative_base

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('asset_correlation_migration.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

Base = declarative_base()

class AssetCorrelation(Base):
    """Asset correlation matrix storage"""
    __tablename__ = 'asset_correlations'
    
    id = Column(Integer, primary_key=True)
    asset1_id = Column(String(50), nullable=False)
    asset2_id = Column(String(50), nullable=False)
    correlation_value = Column(DECIMAL(10, 8), nullable=False)
    correlation_type = Column(String(20), default='return')
    data_source = Column(String(50), default='Asset_Correlation_Matrix')
    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)
    
    # Indexes for efficient lookups
    __table_args__ = (
        Index('idx_asset_correlations_asset1', 'asset1_id'),
        Index('idx_asset_correlations_asset2', 'asset2_id'),
        Index('idx_asset_correlations_pair', 'asset1_id', 'asset2_id'),
    )

class AssetCorrelationMigrator:
    """Specialized migrator for Asset Correlation Matrix"""
    
    def __init__(self, excel_path: str, database_url: str):
        self.excel_path = Path(excel_path)
        self.database_url = database_url
        
        # Migration statistics
        self.stats = {
            'start_time': datetime.now(),
            'end_time': None,
            'matrix_size': 0,
            'extracted_pairs': 0,
            'valid_correlations': 0,
            'loaded_count': 0,
            'diagonal_pairs': 0,
            'symmetric_pairs': 0
        }
        
        # Error tracking
        self.errors = {
            'extraction_errors': [],
            'transformation_errors': [],
            'loading_errors': []
        }
        
        # Initialize database
        self.engine = create_engine(database_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        self.db_session = SessionLocal()
    
    def setup_database(self):
        """Create database tables"""
        try:
            Base.metadata.create_all(bind=self.engine)
            logger.info("Asset correlation tables created successfully")
            return True
        except Exception as e:
            logger.error(f"Error creating database tables: {e}")
            return False
    
    def extract_correlation_matrix(self) -> Dict[str, Any]:
        """Extract the correlation matrix from Asset Correlation worksheet"""
        logger.info("Starting correlation matrix extraction...")
        
        try:
            workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            if 'Asset Correlation' not in workbook.sheetnames:
                raise ValueError("Asset Correlation worksheet not found")
            
            sheet = workbook['Asset Correlation']
            logger.info(f"Loaded Asset Correlation sheet: {sheet.max_column} columns × {sheet.max_row} rows")
            
            # Extract asset IDs from row 1 (headers) and column 1
            asset_ids_row = []
            asset_ids_col = []
            
            # Extract asset IDs from first row (headers)
            for col in range(2, sheet.max_column + 1):  # Skip first column
                asset_id = sheet.cell(row=1, column=col).value
                if asset_id and str(asset_id).strip():
                    asset_ids_row.append(str(asset_id).strip())
                else:
                    break  # Stop at first empty header
            
            # Extract asset IDs from first column
            for row in range(2, sheet.max_row + 1):  # Skip header row
                asset_id = sheet.cell(row=row, column=1).value
                if asset_id and str(asset_id).strip():
                    asset_ids_col.append(str(asset_id).strip())
                else:
                    break  # Stop at first empty row header
            
            logger.info(f"Found {len(asset_ids_row)} asset IDs in headers")
            logger.info(f"Found {len(asset_ids_col)} asset IDs in row headers")
            
            # Verify matrix is square and IDs match
            matrix_size = min(len(asset_ids_row), len(asset_ids_col))
            self.stats['matrix_size'] = matrix_size
            
            logger.info(f"Using {matrix_size}x{matrix_size} correlation matrix")
            
            # Extract correlation values
            correlations = []
            
            for row_idx in range(matrix_size):
                asset1_id = asset_ids_col[row_idx]
                
                for col_idx in range(matrix_size):
                    asset2_id = asset_ids_row[col_idx]
                    
                    # Get correlation value from Excel
                    excel_row = row_idx + 2  # +2 because Excel is 1-indexed and we skip header
                    excel_col = col_idx + 2  # +2 because Excel is 1-indexed and we skip first column
                    
                    correlation_value = sheet.cell(row=excel_row, column=excel_col).value
                    
                    if correlation_value is not None:
                        try:
                            correlation_decimal = self._clean_correlation_value(correlation_value)
                            if correlation_decimal is not None:
                                correlations.append({
                                    'asset1_id': asset1_id,
                                    'asset2_id': asset2_id,
                                    'correlation_value': correlation_decimal,
                                    'is_diagonal': asset1_id == asset2_id,
                                    'excel_row': excel_row,
                                    'excel_col': excel_col
                                })
                                
                                if asset1_id == asset2_id:
                                    self.stats['diagonal_pairs'] += 1
                                
                        except Exception as e:
                            logger.warning(f"Error processing correlation [{asset1_id}, {asset2_id}]: {e}")
                            self.errors['extraction_errors'].append(f"[{asset1_id}, {asset2_id}]: {e}")
            
            self.stats['extracted_pairs'] = len(correlations)
            logger.info(f"Extracted {len(correlations)} correlation pairs from matrix")
            
            return {
                'asset_ids_row': asset_ids_row[:matrix_size],
                'asset_ids_col': asset_ids_col[:matrix_size],
                'correlations': correlations,
                'matrix_size': matrix_size
            }
            
        except Exception as e:
            logger.error(f"Fatal error during matrix extraction: {e}")
            logger.error(traceback.format_exc())
            raise
    
    def _clean_correlation_value(self, value: Any) -> Optional[Decimal]:
        """Clean correlation value to proper decimal format"""
        if value is None:
            return None
        
        try:
            # Handle numeric types
            if isinstance(value, (int, float)):
                decimal_value = Decimal(str(value))
                
                # Validate correlation range (-1 to 1)
                if -1 <= decimal_value <= 1:
                    return decimal_value.quantize(Decimal('0.00000001'), rounding=ROUND_HALF_UP)
                else:
                    logger.warning(f"Correlation value out of range [-1,1]: {value}")
                    return None
            
            # Handle string values
            elif isinstance(value, str):
                cleaned = value.replace('%', '').strip()
                if cleaned in ['', '-', 'N/A', 'n/a', '#N/A']:
                    return None
                
                decimal_value = Decimal(cleaned)
                if -1 <= decimal_value <= 1:
                    return decimal_value.quantize(Decimal('0.00000001'), rounding=ROUND_HALF_UP)
                else:
                    logger.warning(f"Correlation value out of range [-1,1]: {value}")
                    return None
            
            return None
            
        except (ValueError, InvalidOperation):
            logger.warning(f"Could not parse correlation value: {value}")
            return None
    
    def validate_correlations(self, correlation_data: Dict[str, Any]) -> Dict[str, Any]:
        """Validate extracted correlation data"""
        logger.info("Validating correlation matrix...")
        
        correlations = correlation_data['correlations']
        matrix_size = correlation_data['matrix_size']
        
        validation_results = {
            'total_pairs': len(correlations),
            'diagonal_ones': 0,
            'symmetric_pairs': 0,
            'valid_range_pairs': 0,
            'invalid_correlations': [],
            'missing_diagonal': [],
            'asymmetric_pairs': []
        }
        
        # Check diagonal values (should be 1.0 or close)
        diagonal_correlations = [c for c in correlations if c['is_diagonal']]
        for corr in diagonal_correlations:
            if abs(float(corr['correlation_value']) - 1.0) < 0.001:  # Allow small floating point errors
                validation_results['diagonal_ones'] += 1
            else:
                validation_results['invalid_correlations'].append({
                    'asset1': corr['asset1_id'],
                    'asset2': corr['asset2_id'],
                    'value': corr['correlation_value'],
                    'issue': 'diagonal_not_one'
                })
        
        # Check for valid correlation range
        for corr in correlations:
            value = float(corr['correlation_value'])
            if -1 <= value <= 1:
                validation_results['valid_range_pairs'] += 1
            else:
                validation_results['invalid_correlations'].append({
                    'asset1': corr['asset1_id'],
                    'asset2': corr['asset2_id'],
                    'value': corr['correlation_value'],
                    'issue': 'out_of_range'
                })
        
        # Check matrix symmetry (sample check for performance)
        correlation_dict = {}
        for corr in correlations:
            key = (corr['asset1_id'], corr['asset2_id'])
            correlation_dict[key] = float(corr['correlation_value'])
        
        symmetric_count = 0
        asymmetric_pairs = []
        
        # Sample check for symmetry (check first 100 pairs)
        sample_pairs = correlations[:min(100, len(correlations))]
        for corr in sample_pairs:
            if not corr['is_diagonal']:
                asset1, asset2 = corr['asset1_id'], corr['asset2_id']
                value1 = correlation_dict.get((asset1, asset2))
                value2 = correlation_dict.get((asset2, asset1))
                
                if value1 is not None and value2 is not None:
                    if abs(value1 - value2) < 0.0001:  # Very small tolerance
                        symmetric_count += 1
                    else:
                        asymmetric_pairs.append({
                            'asset1': asset1,
                            'asset2': asset2,
                            'value1': value1,
                            'value2': value2,
                            'difference': abs(value1 - value2)
                        })
        
        validation_results['symmetric_pairs'] = symmetric_count
        validation_results['asymmetric_pairs'] = asymmetric_pairs[:10]  # Limit output
        
        # Log validation results
        logger.info(f"Validation Results:")
        logger.info(f"  Total pairs: {validation_results['total_pairs']}")
        logger.info(f"  Diagonal ones: {validation_results['diagonal_ones']}/{len(diagonal_correlations)}")
        logger.info(f"  Valid range: {validation_results['valid_range_pairs']}/{len(correlations)}")
        logger.info(f"  Symmetric pairs (sample): {validation_results['symmetric_pairs']}")
        
        if validation_results['invalid_correlations']:
            logger.warning(f"Found {len(validation_results['invalid_correlations'])} invalid correlations")
        
        return validation_results
    
    def load_correlations_to_database(self, correlations: List[Dict[str, Any]]) -> Dict[str, int]:
        """Load correlation data to database"""
        logger.info(f"Loading {len(correlations)} correlations to database...")
        
        # Filter out only valid correlations
        valid_correlations = [c for c in correlations if c['correlation_value'] is not None]
        logger.info(f"Filtered to {len(valid_correlations)} valid correlations")
        
        self.stats['valid_correlations'] = len(valid_correlations)
        
        batch_size = 1000  # Larger batches for correlation data
        successful_loads = 0
        failed_loads = 0
        
        # Clear existing correlations
        try:
            self.db_session.execute(text("DELETE FROM asset_correlations"))
            self.db_session.commit()
            logger.info("Cleared existing correlation data")
        except Exception as e:
            logger.warning(f"Could not clear existing data: {e}")
        
        # Load in batches
        for i in range(0, len(valid_correlations), batch_size):
            batch = valid_correlations[i:i + batch_size]
            batch_num = i // batch_size + 1
            total_batches = (len(valid_correlations) + batch_size - 1) // batch_size
            
            logger.info(f"Loading batch {batch_num}/{total_batches}: {len(batch)} correlations")
            
            try:
                # Create correlation objects
                correlation_objects = []
                for corr_data in batch:
                    correlation = AssetCorrelation(
                        asset1_id=corr_data['asset1_id'],
                        asset2_id=corr_data['asset2_id'],
                        correlation_value=corr_data['correlation_value'],
                        correlation_type='return',
                        data_source='Asset_Correlation_Matrix'
                    )
                    correlation_objects.append(correlation)
                
                # Bulk insert batch
                self.db_session.add_all(correlation_objects)
                self.db_session.commit()
                successful_loads += len(correlation_objects)
                logger.info(f"Successfully loaded batch {batch_num}")
                
            except SQLAlchemyError as e:
                logger.error(f"Database error loading batch {batch_num}: {e}")
                self.db_session.rollback()
                failed_loads += len(batch)
                self.errors['loading_errors'].append(f"Batch {batch_num}: {e}")
        
        self.stats['loaded_count'] = successful_loads
        
        return {
            'successful': successful_loads,
            'failed': failed_loads,
            'total_processed': successful_loads + failed_loads
        }
    
    def validate_database_data(self) -> Dict[str, Any]:
        """Validate loaded correlation data in database"""
        logger.info("Validating database correlation data...")
        
        try:
            # Count total correlations
            total_count = self.db_session.query(AssetCorrelation).count()
            
            # Count unique assets
            asset1_count = self.db_session.execute(text(
                "SELECT COUNT(DISTINCT asset1_id) FROM asset_correlations"
            )).scalar()
            
            asset2_count = self.db_session.execute(text(
                "SELECT COUNT(DISTINCT asset2_id) FROM asset_correlations"
            )).scalar()
            
            # Check diagonal correlations
            diagonal_count = self.db_session.execute(text(
                "SELECT COUNT(*) FROM asset_correlations WHERE asset1_id = asset2_id"
            )).scalar()
            
            # Check correlation value ranges
            min_correlation = self.db_session.execute(text(
                "SELECT MIN(correlation_value) FROM asset_correlations"
            )).scalar()
            
            max_correlation = self.db_session.execute(text(
                "SELECT MAX(correlation_value) FROM asset_correlations"
            )).scalar()
            
            # Sample correlations
            sample_correlations = self.db_session.execute(text("""
                SELECT asset1_id, asset2_id, correlation_value 
                FROM asset_correlations 
                LIMIT 5
            """)).fetchall()
            
            validation_results = {
                'total_correlations': total_count,
                'unique_asset1_ids': asset1_count,
                'unique_asset2_ids': asset2_count,
                'diagonal_correlations': diagonal_count,
                'min_correlation': float(min_correlation) if min_correlation else None,
                'max_correlation': float(max_correlation) if max_correlation else None,
                'sample_correlations': [
                    {
                        'asset1': row[0],
                        'asset2': row[1], 
                        'correlation': float(row[2])
                    }
                    for row in sample_correlations
                ]
            }
            
            logger.info("Database validation results:")
            logger.info(f"  Total correlations: {validation_results['total_correlations']}")
            logger.info(f"  Unique assets: {validation_results['unique_asset1_ids']}")
            logger.info(f"  Diagonal pairs: {validation_results['diagonal_correlations']}")
            logger.info(f"  Correlation range: [{validation_results['min_correlation']:.4f}, {validation_results['max_correlation']:.4f}]")
            
            return validation_results
            
        except Exception as e:
            logger.error(f"Error validating database data: {e}")
            return {'error': str(e)}
    
    def generate_migration_report(self) -> str:
        """Generate comprehensive migration report"""
        self.stats['end_time'] = datetime.now()
        duration = (self.stats['end_time'] - self.stats['start_time']).total_seconds() / 60
        
        report = []
        report.append("=" * 80)
        report.append("ASSET CORRELATION MATRIX MIGRATION REPORT")
        report.append("=" * 80)
        
        report.append(f"\nMatrix Information:")
        report.append(f"  Matrix Size: {self.stats['matrix_size']}x{self.stats['matrix_size']}")
        report.append(f"  Excel Sheet: Asset Correlation")
        report.append(f"  Total Cells: {self.stats['matrix_size'] ** 2:,}")
        
        report.append(f"\nMigration Summary:")
        report.append(f"  Start Time: {self.stats['start_time']}")
        report.append(f"  End Time: {self.stats['end_time']}")
        report.append(f"  Duration: {duration:.1f} minutes")
        
        report.append(f"\nProcessing Statistics:")
        report.append(f"  Extracted Pairs: {self.stats['extracted_pairs']:,}")
        report.append(f"  Valid Correlations: {self.stats['valid_correlations']:,}")
        report.append(f"  Loaded to Database: {self.stats['loaded_count']:,}")
        report.append(f"  Diagonal Pairs: {self.stats['diagonal_pairs']:,}")
        
        # Success rate
        if self.stats['extracted_pairs'] > 0:
            success_rate = (self.stats['loaded_count'] / self.stats['extracted_pairs']) * 100
            report.append(f"  Success Rate: {success_rate:.1f}%")
        
        # Error summary
        total_errors = sum(len(errors) for errors in self.errors.values())
        report.append(f"\nError Summary:")
        report.append(f"  Total Errors: {total_errors}")
        
        # Migration status
        report.append(f"\nMigration Status:")
        if self.stats['loaded_count'] > 0 and total_errors < self.stats['extracted_pairs'] * 0.1:
            report.append(f"  SUCCESS - Correlation matrix migrated successfully")
        else:
            report.append(f"  ISSUES - Check error details")
        
        report.append("=" * 80)
        
        return "\n".join(report)
    
    def execute_full_migration(self) -> Dict[str, Any]:
        """Execute complete correlation matrix migration"""
        logger.info("Starting Asset Correlation Matrix Migration")
        
        try:
            # Setup database
            if not self.setup_database():
                return {'success': False, 'error': 'Database setup failed'}
            
            # Extract correlation matrix
            matrix_data = self.extract_correlation_matrix()
            
            # Validate correlations
            validation_results = self.validate_correlations(matrix_data)
            
            # Load to database
            load_results = self.load_correlations_to_database(matrix_data['correlations'])
            
            # Validate database data
            db_validation = self.validate_database_data()
            
            # Generate report
            report = self.generate_migration_report()
            
            return {
                'success': load_results['successful'] > 0,
                'statistics': self.stats,
                'load_results': load_results,
                'validation_results': validation_results,
                'db_validation': db_validation,
                'errors': self.errors,
                'report': report
            }
            
        except Exception as e:
            logger.error(f"Fatal error during migration: {e}")
            logger.error(traceback.format_exc())
            
            return {
                'success': False,
                'error': str(e),
                'statistics': self.stats,
                'errors': self.errors
            }
        
        finally:
            self.db_session.close()

def main():
    """Execute Asset Correlation Matrix migration"""
    
    excel_file = "TradeHypoPrelimv32.xlsm"
    database_url = "sqlite:///clo_correlations.db"
    
    if not Path(excel_file).exists():
        print(f"Error: Excel file not found: {excel_file}")
        return 1
    
    print("Starting Asset Correlation Matrix Migration")
    print("=" * 60)
    
    # Execute migration
    migrator = AssetCorrelationMigrator(excel_file, database_url)
    results = migrator.execute_full_migration()
    
    # Display results
    if results['success']:
        print("Migration completed successfully!")
        print(f"Correlations migrated: {results['statistics']['loaded_count']:,}")
        print(f"Matrix size: {results['statistics']['matrix_size']}x{results['statistics']['matrix_size']}")
    else:
        print("Migration failed!")
        if 'error' in results:
            print(f"Error: {results['error']}")
    
    # Display report
    if 'report' in results:
        print("\n" + results['report'])
    
    # Save detailed results
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results_file = f"correlation_matrix_migration_results_{timestamp}.json"
    
    with open(results_file, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    
    print(f"\nDetailed results saved to: {results_file}")
    print(f"Database created: clo_correlations.db")
    
    return 0 if results['success'] else 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)