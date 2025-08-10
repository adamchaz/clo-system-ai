import pandas as pd
import openpyxl
from pathlib import Path
import json
from typing import Dict, List, Any

class ExcelAnalyzer:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.analysis_results = {}
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            print(f"Successfully loaded workbook: {self.file_path}")
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
        return True
    
    def analyze_sheet_structure(self) -> Dict[str, Any]:
        """Analyze the structure of all sheets in the workbook"""
        if not self.workbook:
            return {}
            
        sheet_analysis = {}
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Basic sheet info
            sheet_info = {
                'name': sheet_name,
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'dimensions': f"{sheet.max_column} columns x {sheet.max_row} rows",
                'has_data': sheet.max_row > 1,
                'column_headers': [],
                'data_types': {},
                'named_ranges': [],
                'formulas': []
            }
            
            # Get column headers (first row)
            if sheet.max_row > 0:
                headers = []
                for col in range(1, min(sheet.max_column + 1, 26)):  # Limit to first 25 columns
                    cell_value = sheet.cell(row=1, column=col).value
                    if cell_value is not None:
                        headers.append(str(cell_value))
                    else:
                        headers.append(f"Column_{col}")
                sheet_info['column_headers'] = headers
            
            # Sample first few data rows to understand data types
            sample_data = []
            for row in range(2, min(sheet.max_row + 1, 6)):  # Sample first 5 data rows
                row_data = []
                for col in range(1, min(len(sheet_info['column_headers']) + 1, 26)):
                    cell = sheet.cell(row=row, column=col)
                    row_data.append(cell.value)
                sample_data.append(row_data)
            
            sheet_info['sample_data'] = sample_data
            
            # Find formulas in the sheet
            formulas = []
            for row in sheet.iter_rows(max_row=min(100, sheet.max_row), max_col=min(20, sheet.max_column)):
                for cell in row:
                    if cell.data_type == 'f':  # Formula cell
                        formulas.append({
                            'cell': cell.coordinate,
                            'formula': cell.value
                        })
            sheet_info['formulas'] = formulas[:10]  # Limit to first 10 formulas
            
            sheet_analysis[sheet_name] = sheet_info
            
        return sheet_analysis
    
    def identify_data_relationships(self, sheet_analysis: Dict[str, Any]) -> Dict[str, List[str]]:
        """Identify potential relationships between sheets based on column names"""
        relationships = {}
        all_headers = {}
        
        # Collect all headers by sheet
        for sheet_name, info in sheet_analysis.items():
            all_headers[sheet_name] = info.get('column_headers', [])
        
        # Find common columns between sheets
        for sheet1, headers1 in all_headers.items():
            relationships[sheet1] = []
            for sheet2, headers2 in all_headers.items():
                if sheet1 != sheet2:
                    common_headers = set(headers1) & set(headers2)
                    if common_headers:
                        relationships[sheet1].append({
                            'related_sheet': sheet2,
                            'common_columns': list(common_headers)
                        })
        
        return relationships
    
    def generate_extraction_plan(self, sheet_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Generate a comprehensive data extraction plan"""
        plan = {
            'overview': {
                'total_sheets': len(sheet_analysis),
                'sheets_with_data': len([s for s in sheet_analysis.values() if s['has_data']]),
                'recommended_extraction_order': []
            },
            'extraction_strategies': {},
            'data_validation_checks': [],
            'output_formats': ['csv', 'json', 'parquet']
        }
        
        # Categorize sheets by data size and complexity
        data_sheets = []
        lookup_sheets = []
        summary_sheets = []
        
        for sheet_name, info in sheet_analysis.items():
            if info['has_data']:
                if info['max_row'] > 1000:
                    data_sheets.append(sheet_name)
                elif len(info['formulas']) > 5:
                    summary_sheets.append(sheet_name)
                else:
                    lookup_sheets.append(sheet_name)
        
        # Create extraction strategies for each category
        plan['extraction_strategies']['large_data_sheets'] = {
            'sheets': data_sheets,
            'method': 'chunk_processing',
            'chunk_size': 10000,
            'memory_optimization': True
        }
        
        plan['extraction_strategies']['lookup_tables'] = {
            'sheets': lookup_sheets,
            'method': 'full_load',
            'cache_results': True
        }
        
        plan['extraction_strategies']['calculated_sheets'] = {
            'sheets': summary_sheets,
            'method': 'preserve_formulas',
            'extract_both_values_and_formulas': True
        }
        
        # Recommended extraction order
        plan['overview']['recommended_extraction_order'] = lookup_sheets + data_sheets + summary_sheets
        
        # Data validation checks
        plan['data_validation_checks'] = [
            'Check for missing values in key columns',
            'Validate date formats and ranges',
            'Check numeric columns for outliers',
            'Verify referential integrity between sheets',
            'Validate formula calculations'
        ]
        
        return plan
    
    def run_full_analysis(self) -> Dict[str, Any]:
        """Run complete analysis of the Excel file"""
        if not self.load_workbook():
            return {}
        
        print("Analyzing sheet structures...")
        sheet_analysis = self.analyze_sheet_structure()
        
        print("Identifying data relationships...")
        relationships = self.identify_data_relationships(sheet_analysis)
        
        print("Generating extraction plan...")
        extraction_plan = self.generate_extraction_plan(sheet_analysis)
        
        self.analysis_results = {
            'file_info': {
                'filename': self.file_path.name,
                'file_size_mb': round(self.file_path.stat().st_size / (1024 * 1024), 2),
                'total_sheets': len(sheet_analysis)
            },
            'sheet_analysis': sheet_analysis,
            'relationships': relationships,
            'extraction_plan': extraction_plan
        }
        
        return self.analysis_results
    
    def save_analysis(self, output_path: str = None):
        """Save analysis results to JSON file"""
        if not output_path:
            output_path = self.file_path.parent / f"{self.file_path.stem}_analysis.json"
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.analysis_results, f, indent=2, default=str)
        
        print(f"Analysis saved to: {output_path}")

if __name__ == "__main__":
    # Analyze the Excel file
    analyzer = ExcelAnalyzer("TradeHypoPrelimv32.xlsm")
    results = analyzer.run_full_analysis()
    
    if results:
        # Save detailed analysis
        analyzer.save_analysis()
        
        # Print summary
        print("\n" + "="*50)
        print("EXCEL FILE ANALYSIS SUMMARY")
        print("="*50)
        
        print(f"File: {results['file_info']['filename']}")
        print(f"Size: {results['file_info']['file_size_mb']} MB")
        print(f"Total Sheets: {results['file_info']['total_sheets']}")
        
        print("\nSHEET OVERVIEW:")
        print("-" * 30)
        for sheet_name, info in results['sheet_analysis'].items():
            print(f"• {sheet_name}: {info['dimensions']} - {'Has Data' if info['has_data'] else 'No Data'}")
            if info['column_headers'][:5]:  # Show first 5 headers
                print(f"  Headers: {', '.join(info['column_headers'][:5])}...")
        
        print(f"\nRECOMMENDED EXTRACTION ORDER:")
        print("-" * 30)
        for i, sheet in enumerate(results['extraction_plan']['overview']['recommended_extraction_order'], 1):
            print(f"{i}. {sheet}")
        
        print(f"\nEXTRACTION STRATEGIES:")
        print("-" * 30)
        for strategy, details in results['extraction_plan']['extraction_strategies'].items():
            if details['sheets']:
                print(f"• {strategy}: {len(details['sheets'])} sheets")
                print(f"  Method: {details['method']}")
                print(f"  Sheets: {', '.join(details['sheets'])}")
    else:
        print("Failed to analyze the Excel file.")