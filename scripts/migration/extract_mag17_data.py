#!/usr/bin/env python3
"""
MAG17 CLO Data Extraction Script
Extracts comprehensive data from TradeHypoPrelimv32.xlsm for CLO management system
"""

import json
import pandas as pd
import numpy as np
from datetime import datetime, date
from pathlib import Path
import logging
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class MAG17DataExtractor:
    def __init__(self, excel_path: str, output_dir: str = "."):
        self.excel_path = Path(excel_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Analysis date for this extraction
        self.analysis_date = datetime(2016, 3, 23).date()
        
        # Initialize data containers
        self.extracted_data = {
            'assets': [],
            'tranches': [],
            'triggers': [],
            'accounts': [],
            'parameters': {},
            'metadata': {
                'extraction_date': datetime.now().isoformat(),
                'analysis_date': self.analysis_date.isoformat(),
                'source_file': str(self.excel_path),
                'deal_name': 'MAG17',
                'deal_full_name': 'Magnetar MAG17 CLO'
            }
        }
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            logger.info(f"Loading workbook: {self.excel_path}")
            self.wb = load_workbook(self.excel_path, read_only=True, data_only=True)
            logger.info(f"Available worksheets: {self.wb.sheetnames}")
            return True
        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            return False
    
    def safe_cell_value(self, cell):
        """Safely extract cell value with type conversion"""
        if cell is None:
            return None
        
        value = cell
        if hasattr(cell, 'value'):
            value = cell.value
            
        if value is None:
            return None
        elif isinstance(value, (int, float)):
            return value
        elif isinstance(value, str):
            value = value.strip()
            return value if value else None
        elif isinstance(value, datetime):
            return value.date().isoformat()
        elif isinstance(value, date):
            return value.isoformat()
        else:
            return str(value)
    
    def extract_asset_data(self, worksheet):
        """Extract comprehensive asset-level data from MAG17 worksheet"""
        logger.info("Extracting asset-level data...")
        
        try:
            # Find the asset data section by looking for common headers
            asset_start_row = None
            asset_headers = []
            
            # Scan for asset data headers (typically around row 20-40)
            for row_idx in range(10, 100):
                row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
                row_values = [self.safe_cell_value(cell) for cell in row]
                
                # Look for typical asset headers
                if any(header and ('obligor' in str(header).lower() or 
                                 'balance' in str(header).lower() or
                                 'spread' in str(header).lower() or
                                 'cusip' in str(header).lower()) for header in row_values):
                    asset_start_row = row_idx
                    asset_headers = row_values
                    logger.info(f"Found asset headers at row {asset_start_row}")
                    break
            
            if not asset_start_row:
                logger.error("Could not find asset data section")
                return []
            
            # Clean and map headers
            clean_headers = []
            for i, header in enumerate(asset_headers):
                if header:
                    clean_header = str(header).strip().replace('\n', ' ').replace('  ', ' ')
                    clean_headers.append((i, clean_header))
            
            logger.info(f"Found {len(clean_headers)} asset columns: {[h[1] for h in clean_headers[:10]]}")
            
            # Extract asset rows
            assets = []
            for row_idx in range(asset_start_row + 1, asset_start_row + 300):  # Assume max 300 assets
                try:
                    row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
                    row_values = [self.safe_cell_value(cell) for cell in row]
                    
                    # Check if this row has asset data (first few columns should have values)
                    if not any(row_values[:5]):  # If first 5 columns are empty, likely end of data
                        break
                    
                    # Create asset record
                    asset = {
                        'asset_id': f"MAG17_ASSET_{row_idx - asset_start_row:03d}",
                        'deal_id': 'MAG17',
                        'row_number': row_idx,
                        'extraction_date': datetime.now().isoformat()
                    }
                    
                    # Map all available columns
                    for col_idx, header in clean_headers:
                        if col_idx < len(row_values):
                            value = row_values[col_idx]
                            if value is not None:
                                # Map common field names
                                field_name = self.normalize_field_name(header)
                                asset[field_name] = value
                                
                                # Also keep original header as backup
                                asset[f"original_{col_idx}"] = value
                    
                    # Only add if we have substantial data
                    non_null_fields = sum(1 for v in asset.values() if v is not None)
                    if non_null_fields > 5:  # At least 5 non-null fields
                        assets.append(asset)
                
                except Exception as e:
                    logger.warning(f"Error processing asset row {row_idx}: {e}")
                    continue
            
            logger.info(f"Extracted {len(assets)} assets")
            return assets
            
        except Exception as e:
            logger.error(f"Error extracting asset data: {e}")
            return []
    
    def normalize_field_name(self, header: str) -> str:
        """Normalize field names for database compatibility"""
        if not header:
            return "unknown_field"
        
        # Common field mappings
        field_mappings = {
            'obligor': 'obligor_name',
            'balance': 'outstanding_balance',
            'spread': 'spread_bps',
            'coupon': 'coupon_rate',
            'price': 'current_price',
            'market value': 'market_value',
            'maturity': 'maturity_date',
            'issue': 'issue_date',
            'rating': 'rating',
            'industry': 'industry_classification',
            'country': 'country',
            'currency': 'currency_code',
            'seniority': 'seniority_level'
        }
        
        # Normalize the header
        normalized = str(header).lower().strip()
        normalized = normalized.replace(' ', '_').replace('-', '_')
        normalized = ''.join(c for c in normalized if c.isalnum() or c == '_')
        
        # Apply mappings
        for key, value in field_mappings.items():
            if key in normalized:
                return value
                
        return normalized
    
    def extract_tranche_data(self, worksheet):
        """Extract tranche/liability structure data"""
        logger.info("Extracting tranche data...")
        
        try:
            tranches = []
            
            # Look for tranche data section (usually has "Class" or "Note" headers)
            for row_idx in range(10, 200):
                row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
                row_values = [self.safe_cell_value(cell) for cell in row]
                
                # Look for tranche identifiers
                first_cell = row_values[0] if row_values else None
                if first_cell and isinstance(first_cell, str):
                    cell_text = first_cell.strip().upper()
                    
                    # Check if this looks like a tranche (Class A, Class B, etc.)
                    if (cell_text.startswith('CLASS ') or 
                        cell_text.startswith('NOTE ') or
                        cell_text in ['AAA', 'AA', 'A', 'BBB', 'BB', 'B', 'EQUITY']):
                        
                        tranche = {
                            'tranche_id': f"MAG17_{cell_text.replace(' ', '_')}",
                            'deal_id': 'MAG17',
                            'class_name': cell_text,
                            'row_number': row_idx
                        }
                        
                        # Extract additional tranche data from the row
                        if len(row_values) > 1:
                            tranche['principal_amount'] = row_values[1] if row_values[1] else None
                        if len(row_values) > 2:
                            tranche['coupon_rate'] = row_values[2] if row_values[2] else None
                        if len(row_values) > 3:
                            tranche['rating'] = row_values[3] if row_values[3] else None
                        
                        tranches.append(tranche)
            
            logger.info(f"Extracted {len(tranches)} tranches")
            return tranches
            
        except Exception as e:
            logger.error(f"Error extracting tranche data: {e}")
            return []
    
    def extract_trigger_tests(self, worksheet):
        """Extract OC/IC trigger test data"""
        logger.info("Extracting OC/IC trigger tests...")
        
        try:
            triggers = []
            
            # Look for OC/IC test data
            for row_idx in range(10, 300):
                row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
                row_values = [self.safe_cell_value(cell) for cell in row]
                
                first_cell = row_values[0] if row_values else None
                if first_cell and isinstance(first_cell, str):
                    cell_text = first_cell.strip().lower()
                    
                    # Look for OC/IC test indicators
                    if ('oc' in cell_text and ('test' in cell_text or 'ratio' in cell_text)) or \
                       ('ic' in cell_text and ('test' in cell_text or 'ratio' in cell_text)) or \
                       'overcollateralization' in cell_text or \
                       'interest coverage' in cell_text:
                        
                        trigger = {
                            'trigger_id': f"MAG17_TRIGGER_{len(triggers) + 1}",
                            'deal_id': 'MAG17',
                            'test_name': first_cell.strip(),
                            'row_number': row_idx
                        }
                        
                        # Extract trigger levels and current values
                        if len(row_values) > 1:
                            trigger['trigger_level'] = row_values[1]
                        if len(row_values) > 2:
                            trigger['current_value'] = row_values[2]
                        if len(row_values) > 3:
                            trigger['pass_fail_status'] = row_values[3]
                        
                        triggers.append(trigger)
            
            logger.info(f"Extracted {len(triggers)} trigger tests")
            return triggers
            
        except Exception as e:
            logger.error(f"Error extracting trigger data: {e}")
            return []
    
    def extract_account_data(self, worksheet):
        """Extract account balance and structure data"""
        logger.info("Extracting account data...")
        
        try:
            accounts = []
            
            # Look for account data
            for row_idx in range(10, 200):
                row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
                row_values = [self.safe_cell_value(cell) for cell in row]
                
                first_cell = row_values[0] if row_values else None
                if first_cell and isinstance(first_cell, str):
                    cell_text = first_cell.strip().lower()
                    
                    # Look for account indicators
                    if 'account' in cell_text or 'reserve' in cell_text or 'collection' in cell_text:
                        account = {
                            'account_id': f"MAG17_ACCOUNT_{len(accounts) + 1}",
                            'deal_id': 'MAG17',
                            'account_name': first_cell.strip(),
                            'row_number': row_idx
                        }
                        
                        # Extract account balance and details
                        if len(row_values) > 1:
                            account['balance'] = row_values[1]
                        if len(row_values) > 2:
                            account['minimum_required'] = row_values[2]
                        
                        accounts.append(account)
            
            logger.info(f"Extracted {len(accounts)} accounts")
            return accounts
            
        except Exception as e:
            logger.error(f"Error extracting account data: {e}")
            return []
    
    def extract_deal_parameters(self, worksheet):
        """Extract key deal parameters and constraints"""
        logger.info("Extracting deal parameters...")
        
        try:
            parameters = {
                'deal_id': 'MAG17',
                'deal_name': 'Magnetar MAG17 CLO',
                'extraction_date': datetime.now().isoformat(),
                'analysis_date': self.analysis_date.isoformat()
            }
            
            # Scan for key parameters
            for row_idx in range(1, 500):
                row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
                row_values = [self.safe_cell_value(cell) for cell in row]
                
                first_cell = row_values[0] if row_values else None
                if first_cell and isinstance(first_cell, str):
                    cell_text = first_cell.strip().lower()
                    
                    # Look for key parameters
                    param_indicators = [
                        'reinvestment', 'diversity score', 'warf', 'weighted average',
                        'portfolio', 'concentration', 'effective date', 'legal final',
                        'payment date', 'frequency', 'manager', 'trustee'
                    ]
                    
                    for indicator in param_indicators:
                        if indicator in cell_text:
                            param_key = cell_text.replace(' ', '_').replace('-', '_')
                            param_value = row_values[1] if len(row_values) > 1 else None
                            if param_value is not None:
                                parameters[param_key] = param_value
                            break
            
            logger.info(f"Extracted {len(parameters)} parameters")
            return parameters
            
        except Exception as e:
            logger.error(f"Error extracting parameters: {e}")
            return {}
    
    def process_mag17_worksheet(self):
        """Process the MAG 17 Inputs worksheet"""
        try:
            # Look for MAG 17 worksheet
            mag17_sheet = None
            for sheet_name in self.wb.sheetnames:
                if 'mag' in sheet_name.lower() and '17' in sheet_name:
                    mag17_sheet = self.wb[sheet_name]
                    logger.info(f"Found MAG17 worksheet: {sheet_name}")
                    break
            
            if not mag17_sheet:
                # Try alternative names
                for sheet_name in ['Mag 17 Inputs', 'MAG17', 'Mag17', 'MAG 17']:
                    if sheet_name in self.wb.sheetnames:
                        mag17_sheet = self.wb[sheet_name]
                        logger.info(f"Found MAG17 worksheet: {sheet_name}")
                        break
            
            if not mag17_sheet:
                logger.error(f"Could not find MAG17 worksheet. Available sheets: {self.wb.sheetnames}")
                return False
            
            # Extract all data types
            self.extracted_data['assets'] = self.extract_asset_data(mag17_sheet)
            self.extracted_data['tranches'] = self.extract_tranche_data(mag17_sheet)
            self.extracted_data['triggers'] = self.extract_trigger_tests(mag17_sheet)
            self.extracted_data['accounts'] = self.extract_account_data(mag17_sheet)
            self.extracted_data['parameters'] = self.extract_deal_parameters(mag17_sheet)
            
            return True
            
        except Exception as e:
            logger.error(f"Error processing MAG17 worksheet: {e}")
            return False
    
    def generate_sql_inserts(self):
        """Generate SQL INSERT statements for database migration"""
        sql_statements = []
        
        # Asset INSERT statements
        if self.extracted_data['assets']:
            sql_statements.append("-- MAG17 Asset Data")
            sql_statements.append("INSERT INTO assets (")
            
            # Get all unique field names from assets
            all_fields = set()
            for asset in self.extracted_data['assets']:
                all_fields.update(asset.keys())
            
            field_list = sorted(list(all_fields))
            sql_statements.append("  " + ", ".join(field_list))
            sql_statements.append(") VALUES")
            
            for i, asset in enumerate(self.extracted_data['assets']):
                values = []
                for field in field_list:
                    value = asset.get(field)
                    if value is None:
                        values.append("NULL")
                    elif isinstance(value, str):
                        escaped_value = value.replace("'", "''")
                        values.append(f"'{escaped_value}'")
                    else:
                        values.append(str(value))
                
                comma = "," if i < len(self.extracted_data['assets']) - 1 else ";"
                sql_statements.append(f"  ({', '.join(values)}){comma}")
            
            sql_statements.append("")
        
        # Similar SQL generation for other data types...
        
        return "\n".join(sql_statements)
    
    def generate_validation_report(self):
        """Generate data validation summary"""
        report = []
        report.append("MAG17 Data Extraction Validation Report")
        report.append("=" * 50)
        report.append(f"Extraction Date: {datetime.now().isoformat()}")
        report.append(f"Analysis Date: {self.analysis_date.isoformat()}")
        report.append(f"Source File: {self.excel_path}")
        report.append("")
        
        # Data counts
        report.append("Data Extraction Summary:")
        report.append("-" * 25)
        report.append(f"Assets Extracted: {len(self.extracted_data['assets'])}")
        report.append(f"Tranches Extracted: {len(self.extracted_data['tranches'])}")
        report.append(f"Trigger Tests: {len(self.extracted_data['triggers'])}")
        report.append(f"Accounts: {len(self.extracted_data['accounts'])}")
        report.append(f"Parameters: {len(self.extracted_data['parameters'])}")
        report.append("")
        
        # Asset validation
        if self.extracted_data['assets']:
            report.append("Asset Data Validation:")
            report.append("-" * 20)
            
            # Calculate totals where possible
            total_balance = 0
            total_market_value = 0
            assets_with_balance = 0
            
            for asset in self.extracted_data['assets']:
                balance = asset.get('outstanding_balance')
                if balance and isinstance(balance, (int, float)):
                    total_balance += balance
                    assets_with_balance += 1
                
                mv = asset.get('market_value')
                if mv and isinstance(mv, (int, float)):
                    total_market_value += mv
            
            report.append(f"Total Outstanding Balance: ${total_balance:,.2f}")
            report.append(f"Total Market Value: ${total_market_value:,.2f}")
            report.append(f"Assets with Balance Data: {assets_with_balance}")
            report.append("")
        
        # Field coverage analysis
        if self.extracted_data['assets']:
            report.append("Asset Field Coverage:")
            report.append("-" * 20)
            
            field_counts = {}
            for asset in self.extracted_data['assets']:
                for field, value in asset.items():
                    if value is not None:
                        field_counts[field] = field_counts.get(field, 0) + 1
            
            for field, count in sorted(field_counts.items()):
                coverage = count / len(self.extracted_data['assets']) * 100
                report.append(f"{field}: {count} assets ({coverage:.1f}%)")
            
            report.append("")
        
        return "\n".join(report)
    
    def save_extracted_data(self):
        """Save all extracted data to files"""
        try:
            # Save individual JSON files
            output_files = []
            
            # Assets data
            assets_file = self.output_dir / "mag17_assets_complete.json"
            with open(assets_file, 'w') as f:
                json.dump(self.extracted_data['assets'], f, indent=2, default=str)
            output_files.append(assets_file)
            
            # Tranches data
            tranches_file = self.output_dir / "mag17_tranches.json"
            with open(tranches_file, 'w') as f:
                json.dump(self.extracted_data['tranches'], f, indent=2, default=str)
            output_files.append(tranches_file)
            
            # Triggers data
            triggers_file = self.output_dir / "mag17_triggers.json"
            with open(triggers_file, 'w') as f:
                json.dump(self.extracted_data['triggers'], f, indent=2, default=str)
            output_files.append(triggers_file)
            
            # Accounts data
            accounts_file = self.output_dir / "mag17_accounts.json"
            with open(accounts_file, 'w') as f:
                json.dump(self.extracted_data['accounts'], f, indent=2, default=str)
            output_files.append(accounts_file)
            
            # Parameters data
            parameters_file = self.output_dir / "mag17_parameters.json"
            with open(parameters_file, 'w') as f:
                json.dump(self.extracted_data['parameters'], f, indent=2, default=str)
            output_files.append(parameters_file)
            
            # Complete dataset
            complete_file = self.output_dir / "mag17_complete_dataset.json"
            with open(complete_file, 'w') as f:
                json.dump(self.extracted_data, f, indent=2, default=str)
            output_files.append(complete_file)
            
            # SQL migration script
            sql_file = self.output_dir / "mag17_migration.sql"
            with open(sql_file, 'w') as f:
                f.write(self.generate_sql_inserts())
            output_files.append(sql_file)
            
            # Validation report
            report_file = self.output_dir / "mag17_validation_report.txt"
            with open(report_file, 'w') as f:
                f.write(self.generate_validation_report())
            output_files.append(report_file)
            
            logger.info(f"Successfully saved {len(output_files)} output files:")
            for file in output_files:
                logger.info(f"  - {file}")
            
            return output_files
            
        except Exception as e:
            logger.error(f"Error saving extracted data: {e}")
            return []
    
    def extract_all_data(self):
        """Main extraction method"""
        logger.info("Starting MAG17 data extraction...")
        
        # Load workbook
        if not self.load_workbook():
            return False
        
        # Process MAG17 worksheet
        if not self.process_mag17_worksheet():
            return False
        
        # Save extracted data
        output_files = self.save_extracted_data()
        
        if output_files:
            logger.info("MAG17 data extraction completed successfully!")
            return True
        else:
            logger.error("Failed to save extracted data")
            return False

def main():
    """Main execution function"""
    excel_path = r"C:\Users\adamc\OneDrive\Development\CLO System AI\TradeHypoPrelimv32.xlsm"
    output_dir = r"C:\Users\adamc\OneDrive\Development\CLO System AI"
    
    extractor = MAG17DataExtractor(excel_path, output_dir)
    success = extractor.extract_all_data()
    
    if success:
        print("\n" + "="*60)
        print("MAG17 DATA EXTRACTION COMPLETED SUCCESSFULLY")
        print("="*60)
        print("\nOutput files generated:")
        print("- mag17_assets_complete.json (Asset-level data)")
        print("- mag17_tranches.json (Tranche structure)")
        print("- mag17_triggers.json (OC/IC tests)")
        print("- mag17_accounts.json (Account balances)")
        print("- mag17_parameters.json (Deal parameters)")
        print("- mag17_migration.sql (Database insertion scripts)")
        print("- mag17_validation_report.txt (Data summary)")
        print("- mag17_complete_dataset.json (Full dataset)")
        print("\nData extraction ready for CLO management system integration!")
    else:
        print("\nData extraction failed. Check logs for details.")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())