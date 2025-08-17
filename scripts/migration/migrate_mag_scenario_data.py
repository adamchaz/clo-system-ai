#!/usr/bin/env python3
"""
MAG Scenario Data Migration
Migrate all MAG scenario input sheets (Mag 6-17) to scenario_inputs table
"""

import logging
import sys
import traceback
from pathlib import Path
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from typing import Dict, List, Any, Optional
import json

# Excel processing
import openpyxl

# Database
from sqlalchemy import create_engine, Column, Integer, String, DECIMAL, DateTime, Text, Index, text
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.declarative import declarative_base

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('mag_scenario_migration.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

Base = declarative_base()

class ScenarioInput(Base):
    """MAG scenario input data storage"""
    __tablename__ = 'scenario_inputs'
    
    id = Column(Integer, primary_key=True)
    scenario_name = Column(String(50), nullable=False)  # e.g., 'Mag 6', 'Mag 11'
    scenario_type = Column(String(20), default='MAG')
    section_name = Column(String(100))  # Section within scenario
    parameter_name = Column(String(150), nullable=False)
    parameter_value = Column(Text)  # Store as text for flexibility
    parameter_type = Column(String(20))  # 'numeric', 'text', 'date', 'percentage'
    row_number = Column(Integer)
    column_number = Column(Integer)
    data_source = Column(String(50), default='MAG_Inputs')
    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)
    
    # Indexes for efficient lookups
    __table_args__ = (
        Index('idx_scenario_inputs_scenario', 'scenario_name'),
        Index('idx_scenario_inputs_parameter', 'parameter_name'),
        Index('idx_scenario_inputs_type', 'scenario_type'),
    )

class MAGScenarioMigrator:
    """Migrator for MAG scenario input sheets"""
    
    def __init__(self, excel_path: str, database_url: str):
        self.excel_path = Path(excel_path)
        self.database_url = database_url
        
        # MAG scenario sheets to migrate
        self.mag_sheets = [
            'Mag 6 Inputs', 'Mag 7 Inputs', 'Mag 8 Inputs', 'Mag 9 Inputs',
            'Mag 11 Inputs', 'Mag 12 Inputs', 'Mag 14 Inputs', 'Mag 15 Inputs',
            'Mag 16 Inputs', 'Mag 17 Inputs'
        ]
        
        # Migration statistics
        self.stats = {
            'start_time': datetime.now(),
            'end_time': None,
            'sheets_processed': 0,
            'total_parameters': 0,
            'successful_loads': 0,
            'failed_loads': 0,
            'skipped_empty': 0
        }
        
        # Error tracking
        self.errors = {
            'sheet_errors': [],
            'extraction_errors': [],
            'loading_errors': []
        }
        
        # Initialize database
        self.engine = create_engine(database_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        self.db_session = SessionLocal()
    
    def setup_database(self):
        """Create database tables"""
        try:
            Base.metadata.create_all(bind=self.engine)
            logger.info("Scenario input tables created successfully")
            return True
        except Exception as e:
            logger.error(f"Error creating database tables: {e}")
            return False
    
    def extract_scenario_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract scenario data from a MAG input sheet"""
        logger.info(f"Processing sheet: {sheet_name}")
        
        try:
            workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            if sheet_name not in workbook.sheetnames:
                self.errors['sheet_errors'].append(f"Sheet not found: {sheet_name}")
                return []
            
            sheet = workbook[sheet_name]
            scenario_data = []
            
            # Extract scenario name from sheet name
            scenario_name = sheet_name.replace(' Inputs', '').strip()
            
            # Process all cells in the sheet
            current_section = None
            
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    
                    if cell.value is not None:
                        value = cell.value
                        
                        # Determine parameter type and clean value
                        param_type, cleaned_value = self._classify_and_clean_value(value)
                        
                        if cleaned_value is not None:
                            # Determine section (look for section headers)
                            if self._is_section_header(value):
                                current_section = str(value).strip()
                            
                            # Create parameter name from position and content
                            param_name = self._generate_parameter_name(row, col, value, current_section)
                            
                            parameter = {
                                'scenario_name': scenario_name,
                                'scenario_type': 'MAG',
                                'section_name': current_section,
                                'parameter_name': param_name,
                                'parameter_value': cleaned_value,
                                'parameter_type': param_type,
                                'row_number': row,
                                'column_number': col,
                                'data_source': sheet_name
                            }
                            
                            scenario_data.append(parameter)
            
            logger.info(f"Extracted {len(scenario_data)} parameters from {sheet_name}")
            return scenario_data
            
        except Exception as e:
            error_msg = f"Error processing {sheet_name}: {e}"
            logger.error(error_msg)
            self.errors['extraction_errors'].append(error_msg)
            return []
    
    def _classify_and_clean_value(self, value: Any) -> tuple[str, Optional[str]]:
        """Classify parameter type and clean value"""
        if value is None:
            return 'empty', None
        
        try:
            # Handle numeric values
            if isinstance(value, (int, float)):
                if isinstance(value, float) and (value == int(value)):
                    return 'numeric', str(int(value))
                return 'numeric', str(value)
            
            # Handle datetime values
            elif hasattr(value, 'date'):
                return 'date', value.strftime('%Y-%m-%d')
            
            # Handle string values
            elif isinstance(value, str):
                cleaned = str(value).strip()
                
                if not cleaned:
                    return 'empty', None
                
                # Check if it's a percentage
                if cleaned.endswith('%'):
                    try:
                        percent_val = float(cleaned.replace('%', ''))
                        return 'percentage', str(percent_val)
                    except ValueError:
                        return 'text', cleaned
                
                # Check if it's numeric string
                try:
                    float_val = float(cleaned)
                    return 'numeric', str(float_val)
                except ValueError:
                    return 'text', cleaned
            
            # Default case
            else:
                return 'text', str(value)
                
        except Exception as e:
            logger.warning(f"Error classifying value {value}: {e}")
            return 'text', str(value) if value is not None else None
    
    def _is_section_header(self, value: Any) -> bool:
        """Determine if a value looks like a section header"""
        if not isinstance(value, str):
            return False
        
        value_str = str(value).strip().lower()
        
        # Common section header patterns
        section_indicators = [
            'table of contents', 'assumptions', 'parameters', 'inputs',
            'scenario', 'configuration', 'settings', 'model',
            'economic', 'market', 'credit', 'portfolio'
        ]
        
        return any(indicator in value_str for indicator in section_indicators)
    
    def _generate_parameter_name(self, row: int, col: int, value: Any, section: Optional[str]) -> str:
        """Generate descriptive parameter name"""
        base_name = f"R{row}C{col}"
        
        if section and not self._is_section_header(value):
            section_clean = section.replace(' ', '_').replace(',', '').replace(':', '')
            base_name = f"{section_clean}_{base_name}"
        
        # Add value context if it's descriptive text
        if isinstance(value, str) and len(str(value).strip()) > 0:
            value_clean = str(value)[:50].replace(' ', '_').replace(',', '').replace(':', '')
            if value_clean and not value_clean.replace('.', '').replace('-', '').isdigit():
                base_name = f"{base_name}_{value_clean}"
        
        return base_name[:150]  # Limit to database column size
    
    def load_scenario_data(self, scenario_data: List[Dict[str, Any]]) -> Dict[str, int]:
        """Load scenario data to database"""
        if not scenario_data:
            return {'successful': 0, 'failed': 0}
        
        logger.info(f"Loading {len(scenario_data)} scenario parameters to database...")
        
        batch_size = 500
        successful_loads = 0
        failed_loads = 0
        
        # Load in batches
        for i in range(0, len(scenario_data), batch_size):
            batch = scenario_data[i:i + batch_size]
            batch_num = i // batch_size + 1
            total_batches = (len(scenario_data) + batch_size - 1) // batch_size
            
            logger.info(f"Loading batch {batch_num}/{total_batches}: {len(batch)} parameters")
            
            try:
                # Create parameter objects
                parameter_objects = []
                for param_data in batch:
                    parameter = ScenarioInput(
                        scenario_name=param_data['scenario_name'],
                        scenario_type=param_data['scenario_type'],
                        section_name=param_data['section_name'],
                        parameter_name=param_data['parameter_name'],
                        parameter_value=param_data['parameter_value'],
                        parameter_type=param_data['parameter_type'],
                        row_number=param_data['row_number'],
                        column_number=param_data['column_number'],
                        data_source=param_data['data_source']
                    )
                    parameter_objects.append(parameter)
                
                # Bulk insert batch
                self.db_session.add_all(parameter_objects)
                self.db_session.commit()
                successful_loads += len(parameter_objects)
                logger.info(f"Successfully loaded batch {batch_num}")
                
            except SQLAlchemyError as e:
                logger.error(f"Database error loading batch {batch_num}: {e}")
                self.db_session.rollback()
                failed_loads += len(batch)
                self.errors['loading_errors'].append(f"Batch {batch_num}: {e}")
        
        return {
            'successful': successful_loads,
            'failed': failed_loads,
            'total_processed': successful_loads + failed_loads
        }
    
    def validate_loaded_data(self) -> Dict[str, Any]:
        """Validate loaded scenario data"""
        logger.info("Validating loaded scenario data...")
        
        try:
            # Count total parameters
            total_count = self.db_session.query(ScenarioInput).count()
            
            # Count by scenario
            scenario_counts = self.db_session.execute(text("""
                SELECT scenario_name, COUNT(*) as count
                FROM scenario_inputs
                GROUP BY scenario_name
                ORDER BY scenario_name
            """)).fetchall()
            
            # Count by parameter type
            type_counts = self.db_session.execute(text("""
                SELECT parameter_type, COUNT(*) as count
                FROM scenario_inputs
                GROUP BY parameter_type
                ORDER BY count DESC
            """)).fetchall()
            
            # Sample parameters
            sample_params = self.db_session.execute(text("""
                SELECT scenario_name, section_name, parameter_name, parameter_value, parameter_type
                FROM scenario_inputs
                LIMIT 10
            """)).fetchall()
            
            validation_results = {
                'total_parameters': total_count,
                'scenario_counts': [{'scenario': row[0], 'count': row[1]} for row in scenario_counts],
                'type_counts': [{'type': row[0], 'count': row[1]} for row in type_counts],
                'sample_parameters': [
                    {
                        'scenario': row[0],
                        'section': row[1],
                        'parameter': row[2],
                        'value': row[3],
                        'type': row[4]
                    }
                    for row in sample_params
                ]
            }
            
            logger.info("Validation results:")
            logger.info(f"  Total parameters: {validation_results['total_parameters']}")
            logger.info(f"  Scenarios: {len(validation_results['scenario_counts'])}")
            
            for scenario in validation_results['scenario_counts']:
                logger.info(f"    {scenario['scenario']}: {scenario['count']} parameters")
            
            return validation_results
            
        except Exception as e:
            logger.error(f"Error validating data: {e}")
            return {'error': str(e)}
    
    def execute_full_migration(self) -> Dict[str, Any]:
        """Execute complete MAG scenario migration"""
        logger.info("Starting MAG Scenario Data Migration")
        
        try:
            # Setup database
            if not self.setup_database():
                return {'success': False, 'error': 'Database setup failed'}
            
            # Clear existing scenario data
            try:
                self.db_session.execute(text("DELETE FROM scenario_inputs WHERE scenario_type = 'MAG'"))
                self.db_session.commit()
                logger.info("Cleared existing MAG scenario data")
            except Exception as e:
                logger.warning(f"Could not clear existing data: {e}")
            
            # Process each MAG sheet
            all_scenario_data = []
            
            for sheet_name in self.mag_sheets:
                logger.info(f"Processing sheet: {sheet_name}")
                sheet_data = self.extract_scenario_data(sheet_name)
                
                if sheet_data:
                    all_scenario_data.extend(sheet_data)
                    self.stats['sheets_processed'] += 1
            
            self.stats['total_parameters'] = len(all_scenario_data)
            
            if not all_scenario_data:
                return {'success': False, 'error': 'No scenario data extracted'}
            
            # Load to database
            load_results = self.load_scenario_data(all_scenario_data)
            self.stats['successful_loads'] = load_results['successful']
            self.stats['failed_loads'] = load_results['failed']
            
            # Validate loaded data
            validation_results = self.validate_loaded_data()
            
            # Generate report
            report = self.generate_migration_report()
            
            return {
                'success': load_results['successful'] > 0,
                'statistics': self.stats,
                'load_results': load_results,
                'validation_results': validation_results,
                'errors': self.errors,
                'report': report
            }
            
        except Exception as e:
            logger.error(f"Fatal error during migration: {e}")
            logger.error(traceback.format_exc())
            
            return {
                'success': False,
                'error': str(e),
                'statistics': self.stats,
                'errors': self.errors
            }
        
        finally:
            self.db_session.close()
    
    def generate_migration_report(self) -> str:
        """Generate comprehensive migration report"""
        self.stats['end_time'] = datetime.now()
        duration = (self.stats['end_time'] - self.stats['start_time']).total_seconds() / 60
        
        report = []
        report.append("=" * 80)
        report.append("MAG SCENARIO DATA MIGRATION REPORT")
        report.append("=" * 80)
        
        report.append(f"\nMigration Summary:")
        report.append(f"  Start Time: {self.stats['start_time']}")
        report.append(f"  End Time: {self.stats['end_time']}")
        report.append(f"  Duration: {duration:.1f} minutes")
        
        report.append(f"\nProcessing Statistics:")
        report.append(f"  Sheets Processed: {self.stats['sheets_processed']}/{len(self.mag_sheets)}")
        report.append(f"  Total Parameters Extracted: {self.stats['total_parameters']:,}")
        report.append(f"  Successfully Loaded: {self.stats['successful_loads']:,}")
        report.append(f"  Failed Loads: {self.stats['failed_loads']:,}")
        
        # Success rate
        if self.stats['total_parameters'] > 0:
            success_rate = (self.stats['successful_loads'] / self.stats['total_parameters']) * 100
            report.append(f"  Success Rate: {success_rate:.1f}%")
        
        # Error summary
        total_errors = sum(len(errors) for errors in self.errors.values())
        report.append(f"\nError Summary:")
        report.append(f"  Total Errors: {total_errors}")
        
        # Migration status
        report.append(f"\nMigration Status:")
        if self.stats['successful_loads'] > 0 and self.stats['sheets_processed'] >= len(self.mag_sheets) * 0.8:
            report.append(f"  SUCCESS - MAG scenario data migrated successfully")
            report.append(f"  {self.stats['sheets_processed']} MAG sheets processed")
            report.append(f"  {self.stats['successful_loads']:,} parameters available for modeling")
        else:
            report.append(f"  PARTIAL - Some issues occurred during migration")
        
        report.append("=" * 80)
        
        return "\n".join(report)

def main():
    """Execute MAG scenario data migration"""
    
    excel_file = "TradeHypoPrelimv32.xlsm"
    database_url = "sqlite:///clo_mag_scenarios.db"
    
    if not Path(excel_file).exists():
        print(f"Error: Excel file not found: {excel_file}")
        return 1
    
    print("Starting MAG Scenario Data Migration")
    print("=" * 60)
    
    # Execute migration
    migrator = MAGScenarioMigrator(excel_file, database_url)
    results = migrator.execute_full_migration()
    
    # Display results
    if results['success']:
        print("Migration completed successfully!")
        print(f"Parameters migrated: {results['statistics']['successful_loads']:,}")
        print(f"Sheets processed: {results['statistics']['sheets_processed']}")
    else:
        print("Migration failed!")
        if 'error' in results:
            print(f"Error: {results['error']}")
    
    # Display report
    if 'report' in results:
        print("\n" + results['report'])
    
    # Save detailed results
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results_file = f"mag_scenario_migration_results_{timestamp}.json"
    
    with open(results_file, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    
    print(f"\nDetailed results saved to: {results_file}")
    print(f"Database created: clo_mag_scenarios.db")
    
    return 0 if results['success'] else 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)