#!/usr/bin/env python3
"""
Reference Table Migration Plan and Implementation
"""

import openpyxl
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, List, Any, Optional
import json

class ReferenceTableMigrator:
    """Migrator for Reference Table with multiple sections"""
    
    def __init__(self, excel_path: str, database_url: str = None):
        self.excel_path = Path(excel_path)
        self.database_url = database_url or "sqlite:///clo_reference_data.db"
        
        # Define section mappings and their structures
        self.section_configs = {
            'Actual LIBOR Rates': {
                'type': 'interest_rates',
                'table_name': 'historical_rates',
                'expected_columns': ['rate_date', 'rate_type', 'rate_value', 'source'],
                'description': 'Historical LIBOR and benchmark interest rates'
            },
            'Holidays': {
                'type': 'business_calendar',
                'table_name': 'business_holidays',
                'expected_columns': ['holiday_date', 'holiday_name', 'country', 'market'],
                'description': 'Business holidays for day count calculations'
            },
            'Yield Curve': {
                'type': 'yield_curves',
                'table_name': 'yield_curves',
                'expected_columns': ['curve_date', 'tenor', 'rate', 'curve_type'],
                'description': 'Interest rate curves for asset valuation'
            },
            'S&P 1 year Transition Matrix': {
                'type': 'rating_transitions',
                'table_name': 'rating_transition_matrices',
                'expected_columns': ['agency', 'from_rating', 'to_rating', 'probability', 'time_period'],
                'description': 'Credit rating migration probabilities'
            },
            'S&P Rating Migration Correlation': {
                'type': 'rating_transitions',
                'table_name': 'rating_transition_matrices',
                'expected_columns': ['agency', 'from_rating', 'to_rating', 'probability', 'time_period'],
                'description': 'Extended rating migration correlation data'
            },
            'Cashflow Vectors': {
                'type': 'cashflow_projections',
                'table_name': 'reference_cashflows',
                'expected_columns': ['asset_id', 'payment_date', 'principal', 'interest', 'cashflow_type'],
                'description': 'Asset cashflow projections and vectors'
            },
            'Collateral Cashflows': {
                'type': 'cashflow_projections',
                'table_name': 'reference_cashflows',
                'expected_columns': ['asset_id', 'payment_date', 'principal', 'interest', 'cashflow_type'],
                'description': 'Portfolio-level cashflow projections'
            },
            'Sink Schedule': {
                'type': 'amortization',
                'table_name': 'amortization_schedules',
                'expected_columns': ['schedule_id', 'payment_date', 'principal_amount', 'remaining_balance'],
                'description': 'Scheduled principal amortization'
            },
            'Deals': {
                'type': 'deal_parameters',
                'table_name': 'deal_parameters',
                'expected_columns': ['deal_name', 'parameter_name', 'parameter_value', 'effective_date'],
                'description': 'Deal-specific configuration parameters'
            }
        }
    
    def analyze_reference_structure(self):
        """Analyze the actual structure of the Reference Table"""
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        sheet = workbook['Reference Table']
        
        print("REFERENCE TABLE MIGRATION ANALYSIS")
        print("=" * 60)
        print(f"Total dimensions: {sheet.max_column} columns × {sheet.max_row} rows")
        
        # Identify section boundaries and data ranges
        sections_found = {}
        
        # Find section headers in column 1
        for row_num in range(1, sheet.max_row + 1):
            col1_value = sheet.cell(row=row_num, column=1).value
            
            if col1_value and str(col1_value).strip() in self.section_configs:
                section_name = str(col1_value).strip()
                
                # Find the data range for this section
                data_start_row = row_num + 1
                data_end_row = data_start_row
                
                # Look for data until we hit another section or end of sheet
                for check_row in range(data_start_row, sheet.max_row + 1):
                    next_col1 = sheet.cell(row=check_row, column=1).value
                    if next_col1 and str(next_col1).strip() in self.section_configs:
                        data_end_row = check_row - 1
                        break
                    # Check if row has data
                    has_data = False
                    for col in range(1, min(20, sheet.max_column + 1)):
                        if sheet.cell(row=check_row, column=col).value:
                            has_data = True
                            break
                    if has_data:
                        data_end_row = check_row
                
                sections_found[section_name] = {
                    'header_row': row_num,
                    'data_start': data_start_row,
                    'data_end': data_end_row,
                    'data_rows': max(0, data_end_row - data_start_row + 1),
                    'config': self.section_configs[section_name]
                }
        
        print(f"\nSECTIONS FOUND FOR MIGRATION:")
        total_data_rows = 0
        
        for section_name, section_info in sections_found.items():
            data_count = section_info['data_rows']
            total_data_rows += data_count
            print(f"- {section_name}")
            print(f"  Type: {section_info['config']['type']}")
            print(f"  Table: {section_info['config']['table_name']}")
            print(f"  Rows: {data_count} (rows {section_info['data_start']}-{section_info['data_end']})")
            print(f"  Description: {section_info['config']['description']}")
        
        print(f"\nTOTAL DATA ROWS TO MIGRATE: {total_data_rows}")
        
        return sections_found
    
    def create_database_schema(self):
        """Generate SQL schema for reference tables"""
        
        schema_sql = """
-- Reference Data Database Schema for CLO System
-- Generated from Reference Table analysis

-- Historical Interest Rates
CREATE TABLE historical_rates (
    id INTEGER PRIMARY KEY,
    rate_date DATE NOT NULL,
    rate_type VARCHAR(50) NOT NULL,
    rate_value DECIMAL(10, 6) NOT NULL,
    source VARCHAR(20),
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    INDEX idx_rates_date_type (rate_date, rate_type)
);

-- Business Holidays Calendar
CREATE TABLE business_holidays (
    id INTEGER PRIMARY KEY,
    holiday_date DATE NOT NULL,
    holiday_name VARCHAR(100),
    country VARCHAR(10) DEFAULT 'US',
    market VARCHAR(20) DEFAULT 'USD',
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    INDEX idx_holidays_date (holiday_date)
);

-- Yield Curves
CREATE TABLE yield_curves (
    id INTEGER PRIMARY KEY,
    curve_date DATE NOT NULL,
    tenor VARCHAR(10) NOT NULL,
    rate DECIMAL(10, 6) NOT NULL,
    curve_type VARCHAR(50) DEFAULT 'Treasury',
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    INDEX idx_curves_date_tenor (curve_date, tenor)
);

-- Rating Transition Matrices
CREATE TABLE rating_transition_matrices (
    id INTEGER PRIMARY KEY,
    agency VARCHAR(10) NOT NULL,
    from_rating VARCHAR(20) NOT NULL,
    to_rating VARCHAR(20) NOT NULL,
    probability DECIMAL(8, 6) NOT NULL,
    time_period VARCHAR(20) DEFAULT '1_year',
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    INDEX idx_transitions_agency_from (agency, from_rating)
);

-- Reference Cashflows
CREATE TABLE reference_cashflows (
    id INTEGER PRIMARY KEY,
    asset_id VARCHAR(50),
    payment_date DATE NOT NULL,
    principal_amount DECIMAL(20, 2) DEFAULT 0,
    interest_amount DECIMAL(20, 2) DEFAULT 0,
    cashflow_type VARCHAR(50),
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    INDEX idx_cashflows_asset_date (asset_id, payment_date)
);

-- Amortization Schedules
CREATE TABLE amortization_schedules (
    id INTEGER PRIMARY KEY,
    schedule_id VARCHAR(50) NOT NULL,
    payment_date DATE NOT NULL,
    principal_amount DECIMAL(20, 2) NOT NULL,
    remaining_balance DECIMAL(20, 2),
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    INDEX idx_amortization_schedule_date (schedule_id, payment_date)
);

-- Deal Parameters
CREATE TABLE deal_parameters (
    id INTEGER PRIMARY KEY,
    deal_name VARCHAR(100) NOT NULL,
    parameter_name VARCHAR(100) NOT NULL,
    parameter_value TEXT,
    parameter_type VARCHAR(50),
    effective_date DATE,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    INDEX idx_deal_params (deal_name, parameter_name)
);
"""
        
        return schema_sql
    
    def generate_migration_plan(self):
        """Generate comprehensive migration plan"""
        
        sections = self.analyze_reference_structure()
        schema = self.create_database_schema()
        
        migration_plan = {
            'total_sections': len(sections),
            'total_data_rows': sum(s['data_rows'] for s in sections.values()),
            'database_tables': len(set(s['config']['table_name'] for s in sections.values())),
            'migration_phases': []
        }
        
        # Group sections by table for efficient migration
        tables_to_migrate = {}
        for section_name, section_info in sections.items():
            table_name = section_info['config']['table_name']
            if table_name not in tables_to_migrate:
                tables_to_migrate[table_name] = []
            tables_to_migrate[table_name].append({
                'section': section_name,
                'rows': section_info['data_rows'],
                'range': f"{section_info['data_start']}-{section_info['data_end']}"
            })
        
        # Create migration phases
        for table_name, table_sections in tables_to_migrate.items():
            phase = {
                'table_name': table_name,
                'sections': table_sections,
                'total_rows': sum(s['rows'] for s in table_sections),
                'priority': self._get_table_priority(table_name)
            }
            migration_plan['migration_phases'].append(phase)
        
        # Sort by priority
        migration_plan['migration_phases'].sort(key=lambda x: x['priority'])
        
        return migration_plan, schema
    
    def _get_table_priority(self, table_name):
        """Assign migration priority to tables"""
        priority_map = {
            'business_holidays': 1,        # Required for date calculations
            'historical_rates': 2,         # Core pricing data
            'yield_curves': 3,             # Asset valuation
            'deal_parameters': 4,          # Deal configuration
            'rating_transition_matrices': 5, # Risk modeling
            'amortization_schedules': 6,   # Payment schedules
            'reference_cashflows': 7       # Projections
        }
        return priority_map.get(table_name, 8)

def main():
    """Generate Reference Table migration analysis and plan"""
    
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        print(f"Excel file not found: {excel_file}")
        return
    
    migrator = ReferenceTableMigrator(excel_file)
    
    # Generate migration plan
    migration_plan, schema = migrator.generate_migration_plan()
    
    print("\nMIGRATION EXECUTION PLAN:")
    print("=" * 60)
    
    for i, phase in enumerate(migration_plan['migration_phases'], 1):
        print(f"\nPhase {i}: {phase['table_name']} (Priority {phase['priority']})")
        print(f"  Total rows: {phase['total_rows']}")
        print(f"  Sections:")
        for section in phase['sections']:
            print(f"    - {section['section']}: {section['rows']} rows (Excel {section['range']})")
    
    print(f"\nMIGRATION SUMMARY:")
    print(f"  Total sections: {migration_plan['total_sections']}")
    print(f"  Total data rows: {migration_plan['total_data_rows']}")
    print(f"  Database tables: {migration_plan['database_tables']}")
    print(f"  Migration phases: {len(migration_plan['migration_phases'])}")
    
    # Save schema to file
    with open('reference_table_schema.sql', 'w') as f:
        f.write(schema)
    
    # Save migration plan
    with open('reference_migration_plan.json', 'w') as f:
        json.dump(migration_plan, f, indent=2, default=str)
    
    print(f"\nFiles generated:")
    print(f"  - reference_table_schema.sql")
    print(f"  - reference_migration_plan.json")
    
    print(f"\nRECOMMENDATIONS:")
    print(f"1. Start with high-priority tables (holidays, rates, curves)")
    print(f"2. Validate data quality before migration")
    print(f"3. Create foreign key relationships after data loading")
    print(f"4. Implement proper indexing for lookup performance")
    print(f"5. Consider data versioning for temporal reference data")

if __name__ == "__main__":
    main()