#!/usr/bin/env python3
"""
Complete All Assets Migration Execution Script
Migrate the All Assets worksheet (1,003 assets × 71 columns) to PostgreSQL

This script provides:
- Complete extraction of All Assets sheet data
- Comprehensive data transformation and cleaning
- PostgreSQL database connection and loading
- Multi-level validation with detailed reporting
"""

import json
import logging
import sys
import traceback
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from typing import Dict, List, Any, Optional, Union
import pandas as pd
import numpy as np

# Excel processing
import openpyxl

# Database
from sqlalchemy import create_engine, text, MetaData, Table, Column, Integer, String, DECIMAL, Date, DateTime, Boolean, JSON as SQLJSON
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
from sqlalchemy.ext.declarative import declarative_base

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('all_assets_migration_execution.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Create base for simple model
Base = declarative_base()

class SimpleAsset(Base):
    """Simplified Asset model for migration testing"""
    __tablename__ = 'assets'
    
    id = Column(Integer, primary_key=True)
    blkrock_id = Column(String(50), nullable=False, unique=True)
    issue_name = Column(String(500))
    issuer_name = Column(String(500))
    issuer_id = Column(String(100))
    tranche = Column(String(100))
    bond_loan = Column(String(50))
    par_amount = Column(DECIMAL(20, 2))
    maturity = Column(Date)
    coupon_type = Column(String(50))
    payment_freq = Column(Integer)
    cpn_spread = Column(DECIMAL(10, 6))
    libor_floor = Column(DECIMAL(10, 6))
    index_name = Column(String(100))
    coupon = Column(DECIMAL(10, 6))
    commit_fee = Column(DECIMAL(10, 6))
    unfunded_amount = Column(DECIMAL(20, 2))
    facility_size = Column(DECIMAL(20, 2))
    currency = Column(String(10))
    wal = Column(DECIMAL(10, 4))
    market_value = Column(DECIMAL(10, 4))
    dated_date = Column(Date)
    issue_date = Column(Date)
    first_payment_date = Column(Date)
    amount_issued = Column(DECIMAL(20, 2))
    day_count = Column(String(50))
    index_cap = Column(DECIMAL(10, 6))
    business_day_conv = Column(String(50))
    payment_eom = Column(Boolean)
    amortization_type = Column(String(100))
    country = Column(String(100))
    seniority = Column(String(100))
    
    # Moody's ratings
    mdy_rating = Column(String(20))
    mdy_recovery_rate = Column(DECIMAL(10, 6))
    mdy_dp_rating = Column(String(20))
    mdy_dp_rating_warf = Column(String(20))
    mdy_facility_rating = Column(String(20))
    mdy_facility_outlook = Column(String(20))
    mdy_issuer_rating = Column(String(20))
    mdy_issuer_outlook = Column(String(20))
    mdy_snr_sec_rating = Column(String(20))
    mdy_snr_unsec_rating = Column(String(20))
    mdy_sub_rating = Column(String(20))
    mdy_credit_est_rating = Column(String(20))
    mdy_credit_est_date = Column(Date)
    
    # S&P ratings
    sandp_facility_rating = Column(String(20))
    sandp_issuer_rating = Column(String(20))
    sandp_snr_sec_rating = Column(String(20))
    sandp_subordinate = Column(String(20))
    sandp_rec_rating = Column(String(20))
    
    # Industries
    mdy_industry = Column(String(200))
    sp_industry = Column(String(200))
    mdy_asset_category = Column(String(100))
    sp_priority_category = Column(String(100))
    
    # Special fields
    default_asset = Column(Boolean)
    date_of_default = Column(Date)
    piking = Column(Boolean)
    pik_amount = Column(DECIMAL(20, 2))
    analyst_opinion = Column(String(1000))
    
    # Asset flags as JSON
    flags = Column(SQLJSON)
    
    # Timestamps
    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)

class AllAssetsMigrator:
    """Complete All Assets migration implementation"""
    
    def __init__(self, excel_path: str, database_url: str):
        self.excel_path = Path(excel_path)
        self.database_url = database_url
        
        # Migration statistics
        self.stats = {
            'start_time': datetime.now(),
            'end_time': None,
            'extracted_count': 0,
            'transformed_count': 0,
            'loaded_count': 0,
            'validation_errors': 0
        }
        
        # Error tracking
        self.errors = {
            'extraction_errors': [],
            'transformation_errors': [],
            'loading_errors': [],
            'validation_errors': []
        }
        
        # Initialize database
        self.engine = create_engine(database_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        self.db_session = SessionLocal()
        
        # Column mapping from Excel to database
        self.column_mapping = {
            'BLKRockID': 'blkrock_id',
            'Issue Name': 'issue_name', 
            'Issuer Name': 'issuer_name',
            'ISSUER ID': 'issuer_id',
            'Tranche': 'tranche',
            'Bond/Loan': 'bond_loan',
            'Par Amount': 'par_amount',
            'Maturity': 'maturity',
            'Coupon Type': 'coupon_type',
            'Payment Frequency': 'payment_freq',
            'Coupon Spread': 'cpn_spread',
            'Libor Floor': 'libor_floor',
            'Index': 'index_name',
            'Coupon': 'coupon',
            'Commitment Fee': 'commit_fee',
            'Unfunded Amount': 'unfunded_amount',
            'Facility Size': 'facility_size',
            'Market Value': 'market_value',
            'Currency': 'currency',
            'WAL': 'wal',
            'Dated Date': 'dated_date',
            'Issue Date': 'issue_date',
            'First Payment Date': 'first_payment_date',
            'Amount Issued': 'amount_issued',
            'Day Count': 'day_count',
            'Index Cap': 'index_cap',
            'Business Day Convention': 'business_day_conv',
            'PMT EOM': 'payment_eom',
            'Amortization Type': 'amortization_type',
            'Country ': 'country',  # Note: Excel has trailing space
            'Seniority': 'seniority',
            
            # Moody's ratings
            "Moody's Rating-21": 'mdy_rating',
            "Moody's Recovery Rate-23": 'mdy_recovery_rate',
            "Moody's Default Probability Rating-18": 'mdy_dp_rating',
            "Moody's Rating WARF": 'mdy_dp_rating_warf',
            'Moody Facility Rating': 'mdy_facility_rating',
            'Moody Facility Outlook': 'mdy_facility_outlook',
            'Moody Issuer Rating': 'mdy_issuer_rating',
            'Moody Issuer outlook': 'mdy_issuer_outlook',
            "Moody's Senior Secured Rating": 'mdy_snr_sec_rating',
            'Moody Senior Unsecured Rating': 'mdy_snr_unsec_rating',
            'Moody Subordinate Rating': 'mdy_sub_rating',
            "Moody's Credit Estimate": 'mdy_credit_est_rating',
            "Moody's Credit Estimate Date": 'mdy_credit_est_date',
            
            # S&P ratings
            'S&P Facility Rating': 'sandp_facility_rating',
            'S&P Issuer Rating': 'sandp_issuer_rating',
            'S&P Senior Secured Rating': 'sandp_snr_sec_rating',
            'S&P Subordinated Rating': 'sandp_subordinate',
            'S&P Recovery Rating': 'sandp_rec_rating',
            
            # Industries
            "Moody's Industry": 'mdy_industry',
            'S&P Industry': 'sp_industry',
            'Moody Asset Category': 'mdy_asset_category',
            'S&P Priority Category': 'sp_priority_category',
            
            # Special fields
            'Default Obligation': 'default_asset',
            'Date of Defaulted': 'date_of_default',
            'PiKing': 'piking',
            'PIK Amount': 'pik_amount',
            'Credit Opinion': 'analyst_opinion'
        }
        
        # Asset flags that go into JSON flags field
        self.flag_columns = [
            'Deferred Interest Bond', 'Delayed Drawdown ', 'Revolver',
            'Letter of Credit', 'Participation', 'DIP', 'Convertible', 
            'Structured Finance', 'Bridge Loan', 'Current Pay', 'Cov-Lite',
            'First Lien Last Out Loan'
        ]
    
    def setup_database(self):
        """Create database tables"""
        try:
            Base.metadata.create_all(bind=self.engine)
            logger.info("Database tables created successfully")
            return True
        except Exception as e:
            logger.error(f"Error creating database tables: {e}")
            return False
    
    def extract_all_assets(self) -> List[Dict[str, Any]]:
        """Extract all assets from the All Assets worksheet"""
        logger.info("Starting All Assets data extraction...")
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            if 'All Assets' not in workbook.sheetnames:
                raise ValueError("All Assets worksheet not found")
            
            sheet = workbook['All Assets']
            logger.info(f"Loaded All Assets sheet: {sheet.max_column} columns × {sheet.max_row} rows")
            
            # Extract headers from row 1
            headers = []
            for col in range(1, sheet.max_column + 1):
                header_value = sheet.cell(row=1, column=col).value
                headers.append(str(header_value) if header_value else f"Column_{col}")
            
            logger.info(f"Extracted {len(headers)} column headers")
            
            # Extract asset data (rows 2 to max_row)
            assets_data = []
            
            for row_num in range(2, sheet.max_row + 1):
                try:
                    asset_row = {}
                    
                    # Extract all column values for this row
                    for col_idx, header in enumerate(headers, 1):
                        cell_value = sheet.cell(row=row_num, column=col_idx).value
                        asset_row[header] = cell_value
                    
                    # Skip completely empty rows
                    if all(value in [None, '', ' '] for value in asset_row.values()):
                        continue
                    
                    # Skip rows without BLKRockID
                    if not asset_row.get('BLKRockID'):
                        logger.warning(f"Row {row_num}: Missing BLKRockID, skipping")
                        continue
                    
                    asset_row['_source_row'] = row_num  # Track source row for debugging
                    assets_data.append(asset_row)
                    
                except Exception as e:
                    logger.error(f"Error extracting row {row_num}: {e}")
                    self.errors['extraction_errors'].append(f"Row {row_num}: {e}")
            
            self.stats['extracted_count'] = len(assets_data)
            logger.info(f"Successfully extracted {len(assets_data)} assets from Excel")
            
            return assets_data
            
        except Exception as e:
            logger.error(f"Fatal error during extraction: {e}")
            logger.error(traceback.format_exc())
            raise
    
    def clean_monetary_value(self, value: Any) -> Optional[Decimal]:
        """Clean monetary value from Excel format"""
        if pd.isna(value) or value in [None, '', ' ']:
            return None
        
        try:
            # Handle numeric types directly
            if isinstance(value, (int, float)):
                return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            # Clean string format
            cleaned = str(value).replace('$', '').replace(',', '').replace(' ', '').strip()
            
            if cleaned in ['', '-', '0', 'N/A', 'n/a', '#N/A']:
                return Decimal('0.00')
            
            # Check if this looks like a numeric value
            if not cleaned.replace('.', '').replace('-', '').isdigit():
                logger.warning(f"Non-numeric monetary value: {value}")
                return None
            
            return Decimal(cleaned).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
        except (ValueError, InvalidOperation):
            logger.warning(f"Could not parse monetary value: {value}")
            return None
    
    def parse_date_value(self, value: Any) -> Optional[date]:
        """Parse date value from various Excel formats"""
        if pd.isna(value) or value in [None, '', ' ']:
            return None
        
        try:
            # Handle datetime objects
            if isinstance(value, datetime):
                return value.date()
            
            # Handle date objects
            if isinstance(value, date):
                return value
            
            # Handle numeric Excel dates
            if isinstance(value, (int, float)):
                try:
                    # Excel date serial number (days since 1900-01-01)
                    excel_epoch = datetime(1900, 1, 1)
                    parsed_date = excel_epoch + pd.Timedelta(days=value - 2)  # Excel quirk
                    return parsed_date.date()
                except:
                    pass
            
            # Handle string dates
            date_str = str(value).strip()
            date_formats = [
                '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%m-%d-%Y',
                '%Y%m%d', '%d-%m-%Y', '%m/%d/%y', '%d/%m/%y'
            ]
            
            for fmt in date_formats:
                try:
                    parsed = datetime.strptime(date_str, fmt)
                    return parsed.date()
                except ValueError:
                    continue
            
            return None
            
        except Exception as e:
            logger.warning(f"Error parsing date {value}: {e}")
            return None
    
    def parse_boolean_value(self, value: Any) -> Optional[bool]:
        """Parse boolean value from various formats"""
        if pd.isna(value) or value in [None, '', ' ']:
            return None
        
        if isinstance(value, bool):
            return value
        
        str_value = str(value).strip().upper()
        
        if str_value in ['Y', 'YES', '1', 'TRUE', 'T']:
            return True
        elif str_value in ['N', 'NO', '0', 'FALSE', 'F']:
            return False
        else:
            return None
    
    def transform_asset_data(self, raw_assets: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Transform raw Excel data to database format"""
        logger.info(f"Starting transformation of {len(raw_assets)} assets...")
        
        transformed_assets = []
        
        for i, raw_asset in enumerate(raw_assets, 1):
            try:
                transformed = self._transform_single_asset(raw_asset)
                if transformed:
                    transformed_assets.append(transformed)
                
                if i % 100 == 0:
                    logger.info(f"Transformed {i}/{len(raw_assets)} assets")
                    
            except Exception as e:
                logger.error(f"Error transforming asset {raw_asset.get('BLKRockID', 'unknown')}: {e}")
                self.errors['transformation_errors'].append(f"Asset {raw_asset.get('BLKRockID', 'unknown')}: {e}")
        
        self.stats['transformed_count'] = len(transformed_assets)
        logger.info(f"Successfully transformed {len(transformed_assets)} assets")
        
        return transformed_assets
    
    def _transform_single_asset(self, raw_asset: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Transform a single asset with comprehensive data cleaning"""
        
        # Start with base transformation
        transformed = {
            'created_at': datetime.now(),
            'updated_at': datetime.now()
        }
        
        # Transform directly mapped fields
        for excel_col, db_field in self.column_mapping.items():
            raw_value = raw_asset.get(excel_col)
            
            if raw_value is not None:
                transformed_value = self._transform_field_value(raw_value, db_field, excel_col)
                if transformed_value is not None:
                    transformed[db_field] = transformed_value
        
        # Process asset flags into JSON
        flags = {}
        for flag_col in self.flag_columns:
            flag_value = raw_asset.get(flag_col)
            if flag_value is not None:
                boolean_value = self.parse_boolean_value(flag_value)
                if boolean_value is not None:
                    flag_key = flag_col.lower().replace(' ', '_').strip('_')
                    flags[flag_key] = boolean_value
        
        if flags:
            transformed['flags'] = flags
        
        # Ensure critical fields are present
        if not transformed.get('blkrock_id'):
            logger.warning(f"Asset missing BLKRockID, skipping")
            return None
        
        return transformed
    
    def _transform_field_value(self, value: Any, db_field: str, excel_col: str) -> Any:
        """Transform individual field value based on target database field"""
        
        if pd.isna(value) or value in ['', ' ', 'N/A', 'n/a', '#N/A']:
            return None
        
        try:
            # Monetary fields
            if db_field in ['par_amount', 'market_value', 'facility_size', 'amount_issued', 
                           'unfunded_amount', 'pik_amount', 'commit_fee']:
                return self.clean_monetary_value(value)
            
            # Percentage/rate fields
            elif db_field in ['coupon', 'cpn_spread', 'libor_floor', 'index_cap', 'mdy_recovery_rate']:
                cleaned_value = self.clean_monetary_value(value)
                if cleaned_value and cleaned_value > 1:
                    return cleaned_value / 100  # Convert percentage to decimal
                return cleaned_value
            
            # WAL field - special handling for tranche identifiers like "B2"
            elif db_field in ['wal']:
                if isinstance(value, (int, float)):
                    return Decimal(str(value))
                try:
                    cleaned = str(value).strip()
                    # Skip non-numeric values like "B2", "A1", etc.
                    if not cleaned.replace('.', '').replace('-', '').isdigit():
                        return None
                    return Decimal(cleaned)
                except:
                    return None
            
            # Date fields
            elif db_field in ['maturity', 'dated_date', 'issue_date', 'first_payment_date', 
                             'date_of_default', 'mdy_credit_est_date']:
                return self.parse_date_value(value)
            
            # Integer fields
            elif db_field in ['payment_freq']:
                if isinstance(value, (int, float)):
                    return int(value)
                try:
                    return int(float(str(value).strip()))
                except:
                    return None
            
            # Boolean fields
            elif db_field in ['payment_eom', 'piking', 'default_asset']:
                return self.parse_boolean_value(value)
            
            # Text fields - clean whitespace
            else:
                text = str(value).strip()
                return text if text and text not in ['N/A', 'n/a', '#N/A', '-'] else None
                
        except Exception as e:
            logger.warning(f"Error transforming {excel_col} value '{value}': {e}")
            return None
    
    def load_assets_to_database(self, transformed_assets: List[Dict[str, Any]]) -> Dict[str, int]:
        """Load transformed assets to database with batch processing"""
        logger.info(f"Starting database load of {len(transformed_assets)} assets...")
        
        batch_size = 50  # Smaller batches for complex asset data
        successful_loads = 0
        failed_loads = 0
        
        for i in range(0, len(transformed_assets), batch_size):
            batch = transformed_assets[i:i + batch_size]
            batch_num = i // batch_size + 1
            total_batches = (len(transformed_assets) + batch_size - 1) // batch_size
            
            logger.info(f"Loading batch {batch_num}/{total_batches}: {len(batch)} assets")
            
            try:
                # Create Asset objects
                asset_objects = []
                for asset_data in batch:
                    try:
                        asset = SimpleAsset(**asset_data)
                        asset_objects.append(asset)
                    except Exception as e:
                        logger.error(f"Error creating Asset object for {asset_data.get('blkrock_id')}: {e}")
                        failed_loads += 1
                        self.errors['loading_errors'].append(f"Asset {asset_data.get('blkrock_id')}: {e}")
                
                # Bulk insert batch
                if asset_objects:
                    self.db_session.add_all(asset_objects)
                    self.db_session.commit()
                    successful_loads += len(asset_objects)
                    logger.info(f"Successfully loaded batch {batch_num}")
                
            except SQLAlchemyError as e:
                logger.error(f"Database error loading batch {batch_num}: {e}")
                self.db_session.rollback()
                failed_loads += len(batch)
                self.errors['loading_errors'].append(f"Batch {batch_num}: {e}")
            
            except Exception as e:
                logger.error(f"Unexpected error loading batch {batch_num}: {e}")
                self.db_session.rollback()
                failed_loads += len(batch)
                self.errors['loading_errors'].append(f"Batch {batch_num}: {e}")
        
        self.stats['loaded_count'] = successful_loads
        logger.info(f"Database loading complete: {successful_loads} successful, {failed_loads} failed")
        
        return {
            'successful': successful_loads,
            'failed': failed_loads,
            'total_processed': successful_loads + failed_loads
        }
    
    def validate_migration(self) -> Dict[str, Any]:
        """Comprehensive migration validation"""
        logger.info("Starting migration validation...")
        
        try:
            db_count = self.db_session.query(SimpleAsset).count()
            expected_count = self.stats['extracted_count']
            
            validation_results = {
                'asset_count': {
                    'database_count': db_count,
                    'expected_count': expected_count,
                    'passed': db_count == expected_count
                },
                'sample_data': self._validate_sample_data()
            }
            
            logger.info(f"Validation complete: Database has {db_count} assets, expected {expected_count}")
            return validation_results
            
        except Exception as e:
            logger.error(f"Error during validation: {e}")
            return {'passed': False, 'error': str(e)}
    
    def _validate_sample_data(self):
        """Validate sample of migrated data"""
        try:
            sample_assets = self.db_session.query(SimpleAsset).limit(5).all()
            
            logger.info("Sample migrated assets:")
            for asset in sample_assets:
                logger.info(f"  {asset.blkrock_id}: {asset.issue_name}, Par: {asset.par_amount}")
            
            return {
                'sample_size': len(sample_assets),
                'passed': len(sample_assets) > 0
            }
        except Exception as e:
            logger.error(f"Error validating sample data: {e}")
            return {'passed': False, 'error': str(e)}
    
    def generate_migration_report(self) -> str:
        """Generate comprehensive migration report"""
        self.stats['end_time'] = datetime.now()
        duration = (self.stats['end_time'] - self.stats['start_time']).total_seconds() / 60
        
        report = []
        report.append("=" * 80)
        report.append("ALL ASSETS MIGRATION EXECUTION REPORT")
        report.append("=" * 80)
        
        # Migration summary
        report.append(f"\nMigration Summary:")
        report.append(f"  Start Time: {self.stats['start_time']}")
        report.append(f"  End Time: {self.stats['end_time']}")
        report.append(f"  Duration: {duration:.1f} minutes")
        report.append(f"  Excel File: {self.excel_path}")
        
        # Processing statistics
        report.append(f"\nProcessing Statistics:")
        report.append(f"  Assets Extracted: {self.stats['extracted_count']}")
        report.append(f"  Assets Transformed: {self.stats['transformed_count']}")
        report.append(f"  Assets Loaded: {self.stats['loaded_count']}")
        
        # Success rate
        if self.stats['extracted_count'] > 0:
            success_rate = (self.stats['loaded_count'] / self.stats['extracted_count']) * 100
            report.append(f"  Success Rate: {success_rate:.1f}%")
        
        # Error summary
        total_errors = sum(len(errors) for errors in self.errors.values())
        report.append(f"\nError Summary:")
        report.append(f"  Total Errors: {total_errors}")
        
        # Migration status
        report.append(f"\nMigration Status:")
        if self.stats['loaded_count'] == self.stats['extracted_count'] and total_errors == 0:
            report.append(f"  SUCCESS - All assets migrated successfully")
        elif self.stats['loaded_count'] >= self.stats['extracted_count'] * 0.95:
            report.append(f"  PARTIAL SUCCESS - Most assets migrated with minor issues")
        else:
            report.append(f"  ISSUES - Significant problems encountered during migration")
        
        report.append("=" * 80)
        
        return "\n".join(report)
    
    def execute_full_migration(self) -> Dict[str, Any]:
        """Execute complete All Assets migration workflow"""
        logger.info("Starting All Assets Migration Execution")
        
        try:
            # Setup database
            if not self.setup_database():
                return {'success': False, 'error': 'Database setup failed'}
            
            # Phase 1: Extract
            raw_assets = self.extract_all_assets()
            
            # Phase 2: Transform
            transformed_assets = self.transform_asset_data(raw_assets)
            
            # Phase 3: Load
            load_results = self.load_assets_to_database(transformed_assets)
            
            # Phase 4: Validate
            validation_results = self.validate_migration()
            
            # Generate report
            report = self.generate_migration_report()
            
            return {
                'success': load_results['successful'] > 0,
                'statistics': self.stats,
                'load_results': load_results,
                'validation_results': validation_results,
                'errors': self.errors,
                'report': report
            }
            
        except Exception as e:
            logger.error(f"Fatal error during migration: {e}")
            logger.error(traceback.format_exc())
            
            return {
                'success': False,
                'error': str(e),
                'statistics': self.stats,
                'errors': self.errors
            }
        
        finally:
            self.db_session.close()

def main():
    """Main migration execution"""
    
    # Configuration
    excel_file = "TradeHypoPrelimv32.xlsm"
    database_url = "sqlite:///clo_assets_production.db"  # SQLite for testing
    
    if not Path(excel_file).exists():
        print(f"Error: Excel file not found: {excel_file}")
        return 1
    
    print("Starting Complete All Assets Migration")
    print("=" * 50)
    
    # Execute migration
    migrator = AllAssetsMigrator(excel_file, database_url)
    results = migrator.execute_full_migration()
    
    # Display results
    if results['success']:
        print("Migration completed successfully!")
        print(f"Assets migrated: {results['statistics']['loaded_count']}")
    else:
        print("Migration failed!")
        if 'error' in results:
            print(f"Error: {results['error']}")
    
    # Display report
    if 'report' in results:
        print("\n" + results['report'])
    
    # Save detailed results
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results_file = f"all_assets_migration_results_{timestamp}.json"
    
    with open(results_file, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    
    print(f"\nDetailed results saved to: {results_file}")
    
    return 0 if results['success'] else 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)