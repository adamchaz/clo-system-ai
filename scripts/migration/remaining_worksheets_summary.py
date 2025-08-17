#!/usr/bin/env python3
"""
Simple analysis of remaining worksheets without Unicode characters
"""

import openpyxl
from pathlib import Path
from datetime import datetime

def analyze_remaining_worksheets():
    """Analyze remaining worksheets for migration planning"""
    
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        print(f"Excel file not found: {excel_file}")
        return
    
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    migrated_sheets = {'All Assets', 'Reference Table'}
    
    print("REMAINING WORKSHEETS ANALYSIS")
    print("=" * 60)
    print(f"Total worksheets: {len(workbook.sheetnames)}")
    print(f"Already migrated: {len(migrated_sheets)}")
    
    # Analyze each remaining sheet
    remaining_sheets = []
    
    for sheet_name in workbook.sheetnames:
        if sheet_name in migrated_sheets:
            continue
        
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Sample first few cells to understand content
        sample_data = []
        for row in range(1, min(4, max_row + 1)):
            row_data = []
            for col in range(1, min(6, max_col + 1)):
                value = sheet.cell(row=row, column=col).value
                if value:
                    row_data.append(str(value)[:30])
            if row_data:
                sample_data.extend(row_data)
        
        # Classify sheet type
        sheet_lower = sheet_name.lower()
        sheet_type = "unknown"
        priority = "LOW"
        
        if "run" in sheet_lower and "model" in sheet_lower:
            sheet_type = "model_config"
            priority = "HIGH"
        elif "mag" in sheet_lower and "input" in sheet_lower:
            sheet_type = "scenario_data"
            priority = "HIGH"
        elif "correlation" in sheet_lower:
            sheet_type = "correlation_data"
            priority = "HIGH"
        elif "deal" in sheet_lower and "asset" in sheet_lower:
            sheet_type = "portfolio_data"
            priority = "MEDIUM"
        elif "output" in sheet_lower:
            sheet_type = "output_data"
            priority = "LOW"
        elif "filter" in sheet_lower:
            sheet_type = "filter_criteria"
            priority = "MEDIUM"
        elif "ranking" in sheet_lower or "rebalance" in sheet_lower:
            sheet_type = "output_data"
            priority = "LOW"
        
        # Calculate data density (rough estimate)
        sample_size = min(100, max_row * max_col)
        non_empty = 0
        
        for row in range(1, min(11, max_row + 1)):
            for col in range(1, min(11, max_col + 1)):
                if sheet.cell(row=row, column=col).value:
                    non_empty += 1
        
        density = (non_empty / min(100, max_row * max_col)) * 100
        
        remaining_sheets.append({
            'name': sheet_name,
            'dimensions': f"{max_col} cols x {max_row} rows",
            'type': sheet_type,
            'priority': priority,
            'density': round(density, 1),
            'estimated_records': int(max_row * density / 100) if density > 5 else 0,
            'sample': sample_data[:5]
        })
    
    # Sort by priority
    priority_order = {'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
    remaining_sheets.sort(key=lambda x: (priority_order.get(x['priority'], 4), x['name']))
    
    print(f"\nREMAINING SHEETS TO MIGRATE: {len(remaining_sheets)}")
    print()
    
    current_priority = None
    for sheet in remaining_sheets:
        if sheet['priority'] != current_priority:
            current_priority = sheet['priority']
            print(f"{current_priority} PRIORITY SHEETS:")
            print("-" * 30)
        
        print(f"  {sheet['name']}")
        print(f"    Type: {sheet['type']}")
        print(f"    Size: {sheet['dimensions']}")
        print(f"    Data Density: {sheet['density']}%")
        print(f"    Est. Records: {sheet['estimated_records']}")
        if sheet['sample']:
            print(f"    Sample: {', '.join(sheet['sample'][:3])}")
        print()
    
    # Create migration recommendations
    print("MIGRATION RECOMMENDATIONS:")
    print("=" * 40)
    
    high_priority = [s for s in remaining_sheets if s['priority'] == 'HIGH']
    medium_priority = [s for s in remaining_sheets if s['priority'] == 'MEDIUM']
    low_priority = [s for s in remaining_sheets if s['priority'] == 'LOW']
    
    print(f"\nPHASE 1 (HIGH PRIORITY): {len(high_priority)} sheets")
    for sheet in high_priority:
        print(f"  - {sheet['name']}: {sheet['type']} ({sheet['estimated_records']} records)")
        
        if sheet['type'] == 'model_config':
            print(f"    -> Create 'model_parameters' table")
            print(f"    -> Store configuration as key-value pairs")
        elif sheet['type'] == 'scenario_data':
            print(f"    -> Create 'scenario_inputs' table")
            print(f"    -> Normalize MAG scenario parameters")
        elif sheet['type'] == 'correlation_data':
            print(f"    -> Create 'asset_correlations' table")
            print(f"    -> Store correlation matrices")
    
    print(f"\nPHASE 2 (MEDIUM PRIORITY): {len(medium_priority)} sheets")
    for sheet in medium_priority:
        print(f"  - {sheet['name']}: {sheet['type']} ({sheet['estimated_records']} records)")
    
    print(f"\nPHASE 3 (LOW PRIORITY): {len(low_priority)} sheets")
    for sheet in low_priority:
        print(f"  - {sheet['name']}: {sheet['type']} ({sheet['estimated_records']} records)")
    
    # Calculate totals
    total_records = sum(s['estimated_records'] for s in remaining_sheets)
    high_priority_records = sum(s['estimated_records'] for s in high_priority)
    
    print(f"\nMIGRATION SUMMARY:")
    print(f"  Total sheets remaining: {len(remaining_sheets)}")
    print(f"  Estimated total records: {total_records:,}")
    print(f"  High priority records: {high_priority_records:,}")
    print(f"  Recommended next step: Start with Phase 1 (HIGH priority)")
    
    return remaining_sheets

if __name__ == "__main__":
    results = analyze_remaining_worksheets()