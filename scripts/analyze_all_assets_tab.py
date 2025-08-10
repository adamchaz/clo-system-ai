#!/usr/bin/env python3
"""
All Assets Tab Analysis Script
Analyze the "All Assets" worksheet in TradeHypoPrelimv32.xlsm for migration planning

This script provides detailed analysis of:
- Column structure and data types
- Data quality assessment
- Mapping to database schema
- Transformation requirements
- Sample data extraction
"""

import json
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Any, Optional
import numpy as np

class AllAssetsAnalyzer:
    """Analyze the All Assets worksheet for migration planning"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.worksheet = None
        self.analysis = {
            'sheet_info': {},
            'column_analysis': {},
            'data_quality': {},
            'database_mapping': {},
            'migration_plan': {},
            'sample_data': {}
        }
    
    def load_workbook(self):
        """Load the Excel workbook and find All Assets sheet"""
        try:
            print(f"Loading Excel file: {self.excel_path}")
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            # List all worksheets to find the correct one
            sheet_names = self.workbook.sheetnames
            print(f"Available worksheets: {sheet_names}")
            
            # Look for asset-related sheets
            asset_sheets = [name for name in sheet_names if 'asset' in name.lower()]
            print(f"Asset-related sheets: {asset_sheets}")
            
            # Try to find the All Assets sheet or similar
            target_sheet = None
            for name in sheet_names:
                if 'all assets' in name.lower() or name.lower() == 'assets':
                    target_sheet = name
                    break
            
            if not target_sheet and asset_sheets:
                target_sheet = asset_sheets[0]  # Take the first asset sheet
            
            if target_sheet:
                self.worksheet = self.workbook[target_sheet]
                print(f"Using worksheet: {target_sheet}")
                return True
            else:
                print("No asset-related worksheet found. Using first sheet for analysis.")
                self.worksheet = self.workbook.worksheets[0]
                target_sheet = self.worksheet.title
                
            self.analysis['sheet_info']['name'] = target_sheet
            return True
            
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def analyze_sheet_structure(self):
        """Analyze the structure of the All Assets sheet"""
        if not self.worksheet:
            return
        
        print("\nAnalyzing sheet structure...")
        
        # Get sheet dimensions
        max_row = self.worksheet.max_row
        max_col = self.worksheet.max_column
        
        print(f"Sheet dimensions: {max_col} columns × {max_row} rows")
        
        self.analysis['sheet_info'].update({
            'max_row': max_row,
            'max_col': max_col,
            'total_cells': max_row * max_col
        })
        
        # Extract headers (assuming first row contains headers)
        headers = []
        for col in range(1, max_col + 1):
            cell_value = self.worksheet.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value else f"Column_{col}")
        
        print(f"Detected {len(headers)} columns")
        print("Column headers:")
        for i, header in enumerate(headers, 1):
            print(f"  {i:2d}. {header}")
        
        self.analysis['sheet_info']['headers'] = headers
        self.analysis['sheet_info']['column_count'] = len(headers)
        
        return headers
    
    def analyze_columns(self):
        """Analyze each column in detail"""
        if not self.worksheet:
            return
        
        print("\nAnalyzing column data...")
        
        headers = self.analysis['sheet_info']['headers']
        max_row = self.analysis['sheet_info']['max_row']
        
        column_analysis = {}
        
        for col_idx, header in enumerate(headers, 1):
            print(f"Analyzing column {col_idx}: {header}")
            
            # Extract all values for this column (skip header row)
            values = []
            for row in range(2, min(max_row + 1, 1000)):  # Limit to 1000 rows for performance
                cell_value = self.worksheet.cell(row=row, column=col_idx).value
                if cell_value is not None:
                    values.append(cell_value)
            
            # Analyze data types and patterns
            analysis = self._analyze_column_data(header, values)
            column_analysis[header] = analysis
        
        self.analysis['column_analysis'] = column_analysis
        return column_analysis
    
    def _analyze_column_data(self, header: str, values: List[Any]) -> Dict[str, Any]:
        """Analyze individual column data"""
        if not values:
            return {
                'data_type': 'empty',
                'sample_values': [],
                'unique_count': 0,
                'null_count': 0,
                'total_count': 0
            }
        
        analysis = {
            'total_count': len(values),
            'null_count': 0,
            'unique_count': len(set(str(v) for v in values)),
            'sample_values': values[:5],  # First 5 values
            'data_type': 'unknown',
            'inferred_purpose': 'unknown'
        }
        
        # Infer data type
        numeric_count = 0
        date_count = 0
        text_count = 0
        
        for value in values:
            if isinstance(value, (int, float, Decimal)) and value is not None:
                numeric_count += 1
            elif isinstance(value, datetime):
                date_count += 1
            elif isinstance(value, str) and value.strip():
                text_count += 1
                
                # Try to parse as number
                try:
                    float(value.replace('$', '').replace(',', '').strip())
                    numeric_count += 1
                except ValueError:
                    pass
                    
                # Try to parse as date
                try:
                    pd.to_datetime(value)
                    date_count += 1
                except:
                    pass
        
        # Determine primary data type
        total_valid = len(values)
        if numeric_count > total_valid * 0.7:
            analysis['data_type'] = 'numeric'
        elif date_count > total_valid * 0.7:
            analysis['data_type'] = 'date'
        else:
            analysis['data_type'] = 'text'
        
        # Infer purpose based on header name
        header_lower = header.lower()
        if any(keyword in header_lower for keyword in ['id', 'blkrock', 'cusip', 'isin']):
            analysis['inferred_purpose'] = 'identifier'
        elif any(keyword in header_lower for keyword in ['name', 'issuer', 'issue']):
            analysis['inferred_purpose'] = 'name'
        elif any(keyword in header_lower for keyword in ['amount', 'par', 'balance', 'value', 'price']):
            analysis['inferred_purpose'] = 'monetary'
        elif any(keyword in header_lower for keyword in ['date', 'maturity']):
            analysis['inferred_purpose'] = 'date'
        elif any(keyword in header_lower for keyword in ['rating', 'sp', 'moody']):
            analysis['inferred_purpose'] = 'rating'
        elif any(keyword in header_lower for keyword in ['country', 'geography']):
            analysis['inferred_purpose'] = 'geography'
        elif any(keyword in header_lower for keyword in ['industry', 'sector']):
            analysis['inferred_purpose'] = 'industry'
        elif any(keyword in header_lower for keyword in ['coupon', 'rate', 'spread']):
            analysis['inferred_purpose'] = 'rate'
        
        return analysis
    
    def map_to_database_schema(self):
        """Map Excel columns to database schema"""
        print("\nMapping columns to database schema...")
        
        column_analysis = self.analysis['column_analysis']
        
        # Database field mapping based on CLO Asset model
        database_fields = {
            'blkrock_id': ['blkrock_id', 'blk_rock_id', 'asset_id', 'id', 'blackrock_id'],
            'issue_name': ['issue_name', 'security_name', 'name', 'description'],
            'issuer_name': ['issuer_name', 'issuer', 'company', 'obligor'],
            'issuer_id': ['issuer_id', 'issuer_code'],
            'tranche': ['tranche', 'class'],
            'par_amount': ['par_amount', 'par', 'amount', 'principal', 'face_amount'],
            'market_value': ['market_value', 'mv', 'price', 'market_price', 'current_price'],
            'currency': ['currency', 'ccy'],
            'maturity': ['maturity', 'maturity_date', 'final_maturity'],
            'dated_date': ['dated_date', 'issue_date'],
            'first_payment_date': ['first_payment_date', 'first_coupon_date'],
            'coupon': ['coupon', 'rate', 'interest_rate', 'cpn'],
            'coupon_type': ['coupon_type', 'rate_type'],
            'index_name': ['index_name', 'index', 'base_rate'],
            'cpn_spread': ['cpn_spread', 'spread', 'margin'],
            'payment_freq': ['payment_freq', 'frequency', 'payment_frequency'],
            'sp_rating': ['sp_rating', 's&p_rating', 'sp', 's&p'],
            'mdy_rating': ['mdy_rating', 'moody_rating', 'moodys', 'moody'],
            'country': ['country', 'geography', 'region'],
            'mdy_industry': ['mdy_industry', 'moody_industry', 'industry'],
            'sp_industry': ['sp_industry', 'sp_sector', 'sector'],
            'seniority': ['seniority', 'lien', 'priority'],
            'amount_issued': ['amount_issued', 'issue_size', 'total_issued'],
            'facility_size': ['facility_size', 'commitment'],
            'bond_loan': ['bond_loan', 'instrument_type', 'security_type']
        }
        
        mapping = {}
        
        for excel_header in column_analysis.keys():
            excel_lower = excel_header.lower().replace(' ', '_').replace('&', '').replace('-', '_')
            
            matched_db_field = None
            for db_field, patterns in database_fields.items():
                if any(pattern in excel_lower for pattern in patterns):
                    matched_db_field = db_field
                    break
            
            if matched_db_field:
                mapping[excel_header] = {
                    'database_field': matched_db_field,
                    'confidence': 'high' if any(pattern == excel_lower for pattern in database_fields[matched_db_field]) else 'medium',
                    'transformation_required': self._needs_transformation(column_analysis[excel_header], matched_db_field)
                }
            else:
                mapping[excel_header] = {
                    'database_field': None,
                    'confidence': 'none',
                    'transformation_required': False,
                    'note': 'No matching database field found'
                }
        
        self.analysis['database_mapping'] = mapping
        
        # Print mapping results
        print("Database field mapping:")
        for excel_col, mapping_info in mapping.items():
            db_field = mapping_info.get('database_field', 'UNMAPPED')
            confidence = mapping_info.get('confidence', 'none')
            print(f"  {excel_col:30} → {db_field:20} ({confidence})")
        
        return mapping
    
    def _needs_transformation(self, column_info: Dict, db_field: str) -> bool:
        """Determine if column needs transformation for database storage"""
        data_type = column_info.get('data_type', 'unknown')
        
        # Fields that typically need transformation
        monetary_fields = ['par_amount', 'market_value', 'amount_issued', 'facility_size']
        date_fields = ['maturity', 'dated_date', 'first_payment_date']
        numeric_fields = ['coupon', 'cpn_spread', 'payment_freq']
        
        if db_field in monetary_fields and data_type != 'numeric':
            return True
        elif db_field in date_fields and data_type != 'date':
            return True
        elif db_field in numeric_fields and data_type != 'numeric':
            return True
        
        return False
    
    def assess_data_quality(self):
        """Assess data quality issues"""
        print("\nAssessing data quality...")
        
        column_analysis = self.analysis['column_analysis']
        
        quality_issues = []
        quality_score = 100
        
        for header, analysis in column_analysis.items():
            total_count = analysis['total_count']
            null_count = analysis.get('null_count', 0)
            
            if total_count == 0:
                quality_issues.append(f"Column '{header}' is completely empty")
                quality_score -= 10
            elif null_count > total_count * 0.5:
                quality_issues.append(f"Column '{header}' has high null percentage: {null_count/total_count*100:.1f}%")
                quality_score -= 5
            
            # Check for required fields
            if header.lower() in ['blkrock_id', 'asset_id', 'id'] and null_count > 0:
                quality_issues.append(f"Critical ID field '{header}' has missing values")
                quality_score -= 20
        
        self.analysis['data_quality'] = {
            'quality_score': max(0, quality_score),
            'issues': quality_issues,
            'total_columns_analyzed': len(column_analysis),
            'columns_with_data': sum(1 for a in column_analysis.values() if a['total_count'] > 0)
        }
        
        print(f"Data quality score: {quality_score}/100")
        if quality_issues:
            print("Quality issues found:")
            for issue in quality_issues:
                print(f"  - {issue}")
        else:
            print("No major quality issues detected")
    
    def create_migration_plan(self):
        """Create specific migration plan for All Assets tab"""
        print("\nCreating migration plan...")
        
        database_mapping = self.analysis['database_mapping']
        
        # Count mapped vs unmapped fields
        mapped_fields = {k: v for k, v in database_mapping.items() if v['database_field']}
        unmapped_fields = {k: v for k, v in database_mapping.items() if not v['database_field']}
        
        migration_plan = {
            'mapped_fields': len(mapped_fields),
            'unmapped_fields': len(unmapped_fields),
            'high_confidence_mappings': sum(1 for v in mapped_fields.values() if v['confidence'] == 'high'),
            'fields_requiring_transformation': sum(1 for v in mapped_fields.values() if v['transformation_required']),
            'critical_fields_mapped': self._check_critical_fields_mapped(mapped_fields),
            'extraction_strategy': self._design_extraction_strategy(),
            'transformation_rules': self._design_transformation_rules(mapped_fields),
            'validation_requirements': self._design_validation_requirements()
        }
        
        self.analysis['migration_plan'] = migration_plan
        
        print(f"Migration plan summary:")
        print(f"  Mapped fields: {migration_plan['mapped_fields']}")
        print(f"  High confidence mappings: {migration_plan['high_confidence_mappings']}")
        print(f"  Fields requiring transformation: {migration_plan['fields_requiring_transformation']}")
        
        return migration_plan
    
    def _check_critical_fields_mapped(self, mapped_fields: Dict) -> List[str]:
        """Check if critical fields are mapped"""
        critical_fields = ['blkrock_id', 'issue_name', 'issuer_name', 'par_amount']
        mapped_critical = []
        
        for excel_col, mapping_info in mapped_fields.items():
            if mapping_info['database_field'] in critical_fields:
                mapped_critical.append(mapping_info['database_field'])
        
        return mapped_critical
    
    def _design_extraction_strategy(self) -> Dict[str, Any]:
        """Design data extraction strategy"""
        return {
            'method': 'pandas_openpyxl',
            'batch_size': 1000,
            'skip_rows': 1,  # Skip header row
            'max_rows': None,  # Extract all rows
            'date_columns': [col for col, info in self.analysis['column_analysis'].items() 
                           if info.get('data_type') == 'date'],
            'numeric_columns': [col for col, info in self.analysis['column_analysis'].items() 
                              if info.get('data_type') == 'numeric']
        }
    
    def _design_transformation_rules(self, mapped_fields: Dict) -> List[Dict[str, Any]]:
        """Design transformation rules for each field"""
        rules = []
        
        for excel_col, mapping_info in mapped_fields.items():
            if mapping_info['transformation_required']:
                db_field = mapping_info['database_field']
                
                if db_field in ['par_amount', 'market_value', 'amount_issued', 'facility_size']:
                    rules.append({
                        'field': excel_col,
                        'target': db_field,
                        'transformation': 'clean_monetary',
                        'steps': [
                            'Remove currency symbols ($)',
                            'Remove thousands separators (,)',
                            'Convert to Decimal',
                            'Validate positive value'
                        ]
                    })
                elif db_field in ['maturity', 'dated_date', 'first_payment_date']:
                    rules.append({
                        'field': excel_col,
                        'target': db_field,
                        'transformation': 'parse_date',
                        'steps': [
                            'Parse various date formats',
                            'Convert to YYYY-MM-DD format',
                            'Validate reasonable date range'
                        ]
                    })
                elif db_field in ['coupon', 'cpn_spread']:
                    rules.append({
                        'field': excel_col,
                        'target': db_field,
                        'transformation': 'parse_percentage',
                        'steps': [
                            'Remove % symbol if present',
                            'Convert to decimal format',
                            'Validate reasonable range'
                        ]
                    })
        
        return rules
    
    def _design_validation_requirements(self) -> List[Dict[str, Any]]:
        """Design validation requirements"""
        return [
            {
                'type': 'required_fields',
                'description': 'Validate all critical fields are present',
                'fields': ['blkrock_id', 'issue_name', 'par_amount']
            },
            {
                'type': 'data_types',
                'description': 'Validate data types after transformation',
                'rules': {
                    'par_amount': 'positive_decimal',
                    'market_value': 'positive_decimal',
                    'maturity': 'future_date',
                    'coupon': 'percentage_range_0_50'
                }
            },
            {
                'type': 'business_rules',
                'description': 'Validate business logic constraints',
                'rules': [
                    'Par amount > 0',
                    'Market value between 0 and 200',
                    'Maturity date > issue date',
                    'Coupon rate between 0% and 50%'
                ]
            }
        ]
    
    def extract_sample_data(self, sample_size: int = 10):
        """Extract sample data for testing"""
        print(f"\nExtracting sample data ({sample_size} rows)...")
        
        if not self.worksheet:
            return
        
        headers = self.analysis['sheet_info']['headers']
        sample_data = []
        
        for row in range(2, min(sample_size + 2, self.analysis['sheet_info']['max_row'] + 1)):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = self.worksheet.cell(row=row, column=col_idx).value
                row_data[header] = cell_value
            sample_data.append(row_data)
        
        self.analysis['sample_data'] = sample_data
        
        print("Sample data extracted:")
        for i, row in enumerate(sample_data[:3], 1):  # Show first 3 rows
            print(f"  Row {i}:")
            for header, value in row.items():
                if value is not None:
                    print(f"    {header}: {value}")
        
        return sample_data
    
    def generate_migration_report(self) -> str:
        """Generate comprehensive migration report"""
        report_lines = []
        report_lines.append("=" * 80)
        report_lines.append("ALL ASSETS TAB MIGRATION ANALYSIS REPORT")
        report_lines.append("=" * 80)
        
        # Sheet info
        sheet_info = self.analysis['sheet_info']
        report_lines.append(f"\nSheet Information:")
        report_lines.append(f"  Name: {sheet_info.get('name', 'Unknown')}")
        report_lines.append(f"  Dimensions: {sheet_info.get('column_count', 0)} columns × {sheet_info.get('max_row', 0)} rows")
        report_lines.append(f"  Estimated Assets: {sheet_info.get('max_row', 0) - 1}")
        
        # Database mapping
        database_mapping = self.analysis['database_mapping']
        migration_plan = self.analysis['migration_plan']
        
        report_lines.append(f"\nDatabase Mapping Summary:")
        report_lines.append(f"  Total Excel columns: {len(database_mapping)}")
        report_lines.append(f"  Successfully mapped: {migration_plan.get('mapped_fields', 0)}")
        report_lines.append(f"  High confidence mappings: {migration_plan.get('high_confidence_mappings', 0)}")
        report_lines.append(f"  Require transformation: {migration_plan.get('fields_requiring_transformation', 0)}")
        
        # Critical fields
        critical_mapped = migration_plan.get('critical_fields_mapped', [])
        report_lines.append(f"  Critical fields mapped: {', '.join(critical_mapped) if critical_mapped else 'None'}")
        
        # Data quality
        data_quality = self.analysis['data_quality']
        report_lines.append(f"\nData Quality Assessment:")
        report_lines.append(f"  Quality Score: {data_quality.get('quality_score', 0)}/100")
        report_lines.append(f"  Columns with data: {data_quality.get('columns_with_data', 0)}")
        
        quality_issues = data_quality.get('issues', [])
        if quality_issues:
            report_lines.append(f"  Issues found:")
            for issue in quality_issues:
                report_lines.append(f"    - {issue}")
        else:
            report_lines.append(f"  No major issues detected")
        
        # Migration readiness
        report_lines.append(f"\nMigration Readiness Assessment:")
        if migration_plan.get('mapped_fields', 0) >= 10 and len(critical_mapped) >= 3:
            report_lines.append(f"  Status: ✅ READY - Sufficient field mapping for migration")
        elif migration_plan.get('mapped_fields', 0) >= 5:
            report_lines.append(f"  Status: ⚠️  PARTIAL - Some fields mapped, review required")
        else:
            report_lines.append(f"  Status: ❌ NOT READY - Insufficient field mapping")
        
        report_lines.append("=" * 80)
        
        return "\n".join(report_lines)
    
    def save_analysis(self, output_file: str = None):
        """Save analysis results to JSON file"""
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"all_assets_migration_analysis_{timestamp}.json"
        
        # Convert datetime objects to strings for JSON serialization
        analysis_copy = self.analysis.copy()
        
        with open(output_file, 'w') as f:
            json.dump(analysis_copy, f, indent=2, default=str)
        
        print(f"\nAnalysis saved to: {output_file}")
        return output_file

def main():
    """Main execution function"""
    excel_file = "TradeHypoPrelimv32.xlsm"
    
    if not Path(excel_file).exists():
        print(f"Error: Excel file not found: {excel_file}")
        return
    
    print("Starting All Assets Tab Migration Analysis")
    print("=" * 50)
    
    analyzer = AllAssetsAnalyzer(excel_file)
    
    # Execute analysis steps
    if not analyzer.load_workbook():
        print("Failed to load workbook")
        return
    
    analyzer.analyze_sheet_structure()
    analyzer.analyze_columns()
    analyzer.map_to_database_schema()
    analyzer.assess_data_quality()
    analyzer.create_migration_plan()
    analyzer.extract_sample_data()
    
    # Generate and display report
    report = analyzer.generate_migration_report()
    print("\n" + report)
    
    # Save analysis
    output_file = analyzer.save_analysis()
    
    print(f"\n✅ Analysis complete! Results saved to {output_file}")

if __name__ == "__main__":
    main()