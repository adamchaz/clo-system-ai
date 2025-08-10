#!/usr/bin/env python3
"""
CLO Data Migration Validation Script
Comprehensive validation of data migration from Excel to PostgreSQL

This script provides detailed validation including:
- Data completeness and accuracy verification
- Business rule validation
- Cross-system reconciliation  
- Performance benchmarking
- Audit trail generation
"""

import json
import logging
import sys
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Any, Optional, Tuple
import pandas as pd

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))

# Excel processing
import openpyxl

# Database and ORM
from sqlalchemy import create_engine, text, func
from sqlalchemy.orm import sessionmaker, Session

# CLO System imports
from backend.app.models import (
    Asset, AssetCashFlow, CLODeal, CLOTranche,
    Fee, FeeCalculation, OCTrigger, ICTrigger
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('migration_validation.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class ValidationConfig:
    """Configuration for migration validation"""
    def __init__(self):
        self.excel_path = Path("TradeHypoPrelimv32.xlsm")
        self.database_url = "postgresql://clo_user:password@localhost:5432/clo_db"
        self.tolerance = Decimal('0.001')  # 0.1% tolerance for numerical comparisons
        self.generate_detailed_report = True
        self.export_discrepancies = True

class ValidationReport:
    """Container for validation results"""
    def __init__(self):
        self.validation_id = f"VALIDATION_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.start_time = datetime.now()
        self.end_time = None
        self.overall_success = False
        self.test_results = {}
        self.discrepancies = []
        self.performance_metrics = {}
        self.recommendations = []
    
    def add_test_result(self, test_name: str, passed: bool, details: Dict[str, Any]):
        """Add test result to report"""
        self.test_results[test_name] = {
            'passed': passed,
            'details': details,
            'timestamp': datetime.now()
        }
    
    def add_discrepancy(self, category: str, description: str, severity: str, details: Dict[str, Any] = None):
        """Add discrepancy to report"""
        self.discrepancies.append({
            'category': category,
            'description': description,
            'severity': severity,
            'details': details or {},
            'timestamp': datetime.now()
        })
    
    def finalize_report(self):
        """Finalize validation report"""
        self.end_time = datetime.now()
        self.overall_success = all(result['passed'] for result in self.test_results.values())
        
        # Generate recommendations based on results
        if not self.overall_success:
            high_severity_issues = [d for d in self.discrepancies if d['severity'] == 'HIGH']
            if high_severity_issues:
                self.recommendations.append("Address HIGH severity discrepancies before production deployment")
            
            failed_tests = [name for name, result in self.test_results.items() if not result['passed']]
            if failed_tests:
                self.recommendations.append(f"Re-run migration for failed test areas: {', '.join(failed_tests)}")
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert report to dictionary"""
        return {
            'validation_id': self.validation_id,
            'start_time': self.start_time.isoformat(),
            'end_time': self.end_time.isoformat() if self.end_time else None,
            'duration_minutes': (self.end_time - self.start_time).total_seconds() / 60 if self.end_time else None,
            'overall_success': self.overall_success,
            'test_results': self.test_results,
            'discrepancy_count': len(self.discrepancies),
            'high_severity_discrepancies': len([d for d in self.discrepancies if d['severity'] == 'HIGH']),
            'discrepancies': self.discrepancies,
            'performance_metrics': self.performance_metrics,
            'recommendations': self.recommendations
        }

class ExcelReferenceData:
    """Extract reference data from Excel for comparison"""
    
    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self.workbook = None
        self.reference_data = {}
        
    def load_workbook(self):
        """Load Excel workbook for reference data extraction"""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True, read_only=True)
            logger.info(f"Loaded Excel workbook with {len(self.workbook.worksheets)} worksheets")
        except Exception as e:
            logger.error(f"Failed to load Excel workbook: {e}")
            raise
    
    def extract_asset_summary_stats(self) -> Dict[str, Any]:
        """Extract summary statistics about assets from Excel"""
        stats = {
            'total_assets': 0,
            'total_par_amount': Decimal('0'),
            'asset_count_by_rating': {},
            'asset_count_by_country': {},
            'asset_count_by_industry': {}
        }
        
        try:
            # Look for asset-related worksheets
            for sheet in self.workbook.worksheets:
                if any(keyword in sheet.title.lower() for keyword in ['asset', 'portfolio', 'collateral']):
                    sheet_stats = self._extract_stats_from_asset_sheet(sheet)
                    self._merge_stats(stats, sheet_stats)
            
            return stats
            
        except Exception as e:
            logger.error(f"Error extracting asset summary stats: {e}")
            return stats
    
    def _extract_stats_from_asset_sheet(self, sheet) -> Dict[str, Any]:
        """Extract statistics from individual asset sheet"""
        stats = {
            'asset_count': 0,
            'total_par': Decimal('0'),
            'ratings': {},
            'countries': {},
            'industries': {}
        }
        
        # Convert to DataFrame for easier processing
        data = []
        for row in sheet.iter_rows(values_only=True, max_row=1000):  # Limit for performance
            if any(cell for cell in row):  # Skip empty rows
                data.append(row)
        
        if len(data) < 2:  # No header + data
            return stats
        
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # Find relevant columns
        id_col = self._find_column(df.columns, ['blkrock_id', 'asset_id', 'id'])
        par_col = self._find_column(df.columns, ['par_amount', 'par', 'amount'])
        rating_col = self._find_column(df.columns, ['sp_rating', 's&p', 'rating'])
        country_col = self._find_column(df.columns, ['country', 'geography'])
        industry_col = self._find_column(df.columns, ['industry', 'sector'])
        
        # Count assets and sum par amounts
        if id_col:
            valid_assets = df[df[id_col].notna() & (df[id_col] != "")]
            stats['asset_count'] = len(valid_assets)
            
            # Sum par amounts if column exists
            if par_col:
                par_values = pd.to_numeric(valid_assets[par_col], errors='coerce')
                stats['total_par'] = Decimal(str(par_values.sum()))
            
            # Count by rating
            if rating_col:
                rating_counts = valid_assets[rating_col].value_counts().to_dict()
                stats['ratings'] = {str(k): v for k, v in rating_counts.items() if pd.notna(k)}
            
            # Count by country
            if country_col:
                country_counts = valid_assets[country_col].value_counts().to_dict()
                stats['countries'] = {str(k): v for k, v in country_counts.items() if pd.notna(k)}
            
            # Count by industry
            if industry_col:
                industry_counts = valid_assets[industry_col].value_counts().to_dict()
                stats['industries'] = {str(k): v for k, v in industry_counts.items() if pd.notna(k)}
        
        return stats
    
    def _find_column(self, columns, patterns: List[str]) -> Optional[str]:
        """Find column matching patterns"""
        for col in columns:
            if col and any(pattern.lower() in str(col).lower().replace(' ', '_').replace('&', '')
                          for pattern in patterns):
                return col
        return None
    
    def _merge_stats(self, main_stats: Dict, sheet_stats: Dict):
        """Merge sheet statistics into main statistics"""
        main_stats['total_assets'] += sheet_stats.get('asset_count', 0)
        main_stats['total_par_amount'] += sheet_stats.get('total_par', Decimal('0'))
        
        # Merge rating counts
        for rating, count in sheet_stats.get('ratings', {}).items():
            main_stats['asset_count_by_rating'][rating] = main_stats['asset_count_by_rating'].get(rating, 0) + count
        
        # Merge country counts
        for country, count in sheet_stats.get('countries', {}).items():
            main_stats['asset_count_by_country'][country] = main_stats['asset_count_by_country'].get(country, 0) + count
        
        # Merge industry counts
        for industry, count in sheet_stats.get('industries', {}).items():
            main_stats['asset_count_by_industry'][industry] = main_stats['asset_count_by_industry'].get(industry, 0) + count

class DatabaseValidator:
    """Validate migrated data in database"""
    
    def __init__(self, db_session: Session, tolerance: Decimal):
        self.db = db_session
        self.tolerance = tolerance
        
    def validate_asset_completeness(self) -> Dict[str, Any]:
        """Validate asset data completeness"""
        logger.info("Validating asset data completeness...")
        
        validation_result = {
            'passed': True,
            'total_assets': 0,
            'missing_critical_fields': {},
            'data_quality_issues': []
        }
        
        try:
            # Get total asset count
            total_assets = self.db.query(Asset).count()
            validation_result['total_assets'] = total_assets
            
            if total_assets == 0:
                validation_result['passed'] = False
                validation_result['data_quality_issues'].append("No assets found in database")
                return validation_result
            
            # Check for missing critical fields
            critical_fields = {
                'blkrock_id': 'Asset identifier',
                'issue_name': 'Issue name',
                'issuer_name': 'Issuer name',
                'par_amount': 'Par amount',
                'maturity': 'Maturity date'
            }
            
            for field, description in critical_fields.items():
                null_count = self.db.query(Asset).filter(
                    getattr(Asset, field).is_(None)
                ).count()
                
                if null_count > 0:
                    percentage = (null_count / total_assets) * 100
                    validation_result['missing_critical_fields'][field] = {
                        'count': null_count,
                        'percentage': round(percentage, 2),
                        'description': description
                    }
                    
                    if percentage > 10:  # More than 10% missing
                        validation_result['passed'] = False
                        validation_result['data_quality_issues'].append(
                            f"High percentage of missing {description}: {percentage:.1f}%"
                        )
            
            # Check for duplicate asset IDs
            duplicate_count = self.db.query(Asset.blkrock_id).group_by(Asset.blkrock_id).having(
                func.count(Asset.blkrock_id) > 1
            ).count()
            
            if duplicate_count > 0:
                validation_result['passed'] = False
                validation_result['data_quality_issues'].append(
                    f"Found {duplicate_count} duplicate asset IDs"
                )
            
            logger.info(f"Asset completeness validation: {'PASSED' if validation_result['passed'] else 'FAILED'}")
            return validation_result
            
        except Exception as e:
            logger.error(f"Error validating asset completeness: {e}")
            validation_result['passed'] = False
            validation_result['data_quality_issues'].append(f"Validation error: {e}")
            return validation_result
    
    def validate_data_accuracy(self, excel_stats: Dict[str, Any]) -> Dict[str, Any]:
        """Validate data accuracy against Excel source"""
        logger.info("Validating data accuracy against Excel...")
        
        validation_result = {
            'passed': True,
            'comparison_results': {},
            'discrepancies': []
        }
        
        try:
            # Compare asset count
            db_asset_count = self.db.query(Asset).count()
            excel_asset_count = excel_stats.get('total_assets', 0)
            
            count_diff = abs(db_asset_count - excel_asset_count)
            count_diff_percentage = (count_diff / max(excel_asset_count, 1)) * 100
            
            validation_result['comparison_results']['asset_count'] = {
                'database_count': db_asset_count,
                'excel_count': excel_asset_count,
                'difference': count_diff,
                'difference_percentage': round(count_diff_percentage, 2)
            }
            
            if count_diff_percentage > 5:  # More than 5% difference
                validation_result['passed'] = False
                validation_result['discrepancies'].append(
                    f"Asset count difference: DB={db_asset_count}, Excel={excel_asset_count} ({count_diff_percentage:.1f}% difference)"
                )
            
            # Compare total par amount
            db_total_par = self.db.query(func.sum(Asset.par_amount)).scalar() or Decimal('0')
            excel_total_par = excel_stats.get('total_par_amount', Decimal('0'))
            
            if excel_total_par > 0:
                par_diff = abs(db_total_par - excel_total_par)
                par_diff_percentage = (par_diff / excel_total_par) * 100
                
                validation_result['comparison_results']['total_par_amount'] = {
                    'database_total': float(db_total_par),
                    'excel_total': float(excel_total_par),
                    'difference': float(par_diff),
                    'difference_percentage': round(float(par_diff_percentage), 4)
                }
                
                if par_diff_percentage > 1:  # More than 1% difference
                    validation_result['passed'] = False
                    validation_result['discrepancies'].append(
                        f"Total par amount difference: DB={db_total_par}, Excel={excel_total_par} ({par_diff_percentage:.2f}% difference)"
                    )
            
            # Compare rating distribution
            self._compare_rating_distribution(validation_result, excel_stats)
            
            logger.info(f"Data accuracy validation: {'PASSED' if validation_result['passed'] else 'FAILED'}")
            return validation_result
            
        except Exception as e:
            logger.error(f"Error validating data accuracy: {e}")
            validation_result['passed'] = False
            validation_result['discrepancies'].append(f"Validation error: {e}")
            return validation_result
    
    def _compare_rating_distribution(self, validation_result: Dict, excel_stats: Dict):
        """Compare rating distribution between database and Excel"""
        try:
            # Get database rating distribution
            db_rating_counts = {}
            rating_query = self.db.query(Asset.sp_rating, func.count(Asset.sp_rating)).group_by(Asset.sp_rating).all()
            
            for rating, count in rating_query:
                if rating:
                    db_rating_counts[rating] = count
            
            excel_rating_counts = excel_stats.get('asset_count_by_rating', {})
            
            # Compare distributions
            all_ratings = set(db_rating_counts.keys()) | set(excel_rating_counts.keys())
            rating_discrepancies = []
            
            for rating in all_ratings:
                db_count = db_rating_counts.get(rating, 0)
                excel_count = excel_rating_counts.get(rating, 0)
                
                if db_count != excel_count:
                    diff = abs(db_count - excel_count)
                    rating_discrepancies.append({
                        'rating': rating,
                        'db_count': db_count,
                        'excel_count': excel_count,
                        'difference': diff
                    })
            
            validation_result['comparison_results']['rating_distribution'] = {
                'total_ratings_compared': len(all_ratings),
                'discrepancies_found': len(rating_discrepancies),
                'discrepancy_details': rating_discrepancies
            }
            
            if len(rating_discrepancies) > len(all_ratings) * 0.2:  # More than 20% of ratings have discrepancies
                validation_result['passed'] = False
                validation_result['discrepancies'].append(
                    f"Significant rating distribution discrepancies: {len(rating_discrepancies)}/{len(all_ratings)} ratings differ"
                )
                
        except Exception as e:
            logger.error(f"Error comparing rating distribution: {e}")
            validation_result['discrepancies'].append(f"Rating comparison error: {e}")
    
    def validate_business_rules(self) -> Dict[str, Any]:
        """Validate business rule compliance"""
        logger.info("Validating business rule compliance...")
        
        validation_result = {
            'passed': True,
            'rule_violations': []
        }
        
        try:
            # Rule 1: Par amounts should be positive
            negative_par_count = self.db.query(Asset).filter(Asset.par_amount < 0).count()
            if negative_par_count > 0:
                validation_result['passed'] = False
                validation_result['rule_violations'].append(
                    f"Found {negative_par_count} assets with negative par amounts"
                )
            
            # Rule 2: Maturity dates should be in the future or recent past
            old_maturity_count = self.db.query(Asset).filter(
                Asset.maturity < date(2000, 1, 1)
            ).count()
            if old_maturity_count > 0:
                validation_result['passed'] = False
                validation_result['rule_violations'].append(
                    f"Found {old_maturity_count} assets with maturity dates before 2000"
                )
            
            # Rule 3: Market values should be between 0 and 200 (as percentage)
            invalid_mv_count = self.db.query(Asset).filter(
                (Asset.market_value < 0) | (Asset.market_value > 200)
            ).count()
            if invalid_mv_count > 0:
                validation_result['passed'] = False
                validation_result['rule_violations'].append(
                    f"Found {invalid_mv_count} assets with invalid market values (outside 0-200 range)"
                )
            
            # Rule 4: Ratings should be valid rating symbols
            invalid_ratings = self._validate_rating_symbols()
            if invalid_ratings:
                validation_result['passed'] = False
                validation_result['rule_violations'].extend(invalid_ratings)
            
            logger.info(f"Business rule validation: {'PASSED' if validation_result['passed'] else 'FAILED'}")
            return validation_result
            
        except Exception as e:
            logger.error(f"Error validating business rules: {e}")
            validation_result['passed'] = False
            validation_result['rule_violations'].append(f"Validation error: {e}")
            return validation_result
    
    def _validate_rating_symbols(self) -> List[str]:
        """Validate rating symbols are in expected format"""
        violations = []
        
        try:
            # Valid S&P ratings pattern
            valid_sp_ratings = [
                'AAA', 'AA+', 'AA', 'AA-', 'A+', 'A', 'A-',
                'BBB+', 'BBB', 'BBB-', 'BB+', 'BB', 'BB-',
                'B+', 'B', 'B-', 'CCC+', 'CCC', 'CCC-', 'CC', 'C', 'D'
            ]
            
            # Check S&P ratings
            invalid_sp_count = self.db.query(Asset).filter(
                Asset.sp_rating.is_not(None) &
                ~Asset.sp_rating.in_(valid_sp_ratings)
            ).count()
            
            if invalid_sp_count > 0:
                violations.append(f"Found {invalid_sp_count} assets with invalid S&P ratings")
            
            # Similar check could be done for Moody's ratings
            
        except Exception as e:
            violations.append(f"Error validating rating symbols: {e}")
        
        return violations

class MigrationValidator:
    """Main migration validation controller"""
    
    def __init__(self, config: ValidationConfig):
        self.config = config
        
        # Initialize database session
        engine = create_engine(config.database_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        self.db_session = SessionLocal()
        
        # Initialize components
        self.excel_reference = ExcelReferenceData(config.excel_path)
        self.db_validator = DatabaseValidator(self.db_session, config.tolerance)
        
        # Initialize report
        self.report = ValidationReport()
    
    def execute_full_validation(self) -> ValidationReport:
        """Execute comprehensive migration validation"""
        logger.info(f"Starting migration validation: {self.report.validation_id}")
        
        try:
            # Load Excel reference data
            logger.info("Loading Excel reference data...")
            self.excel_reference.load_workbook()
            excel_stats = self.excel_reference.extract_asset_summary_stats()
            logger.info(f"Excel reference data loaded: {excel_stats.get('total_assets', 0)} assets found")
            
            # Test 1: Data completeness validation
            logger.info("Running data completeness validation...")
            completeness_result = self.db_validator.validate_asset_completeness()
            self.report.add_test_result("data_completeness", completeness_result['passed'], completeness_result)
            
            for issue in completeness_result.get('data_quality_issues', []):
                self.report.add_discrepancy("DATA_COMPLETENESS", issue, "HIGH")
            
            # Test 2: Data accuracy validation
            logger.info("Running data accuracy validation...")
            accuracy_result = self.db_validator.validate_data_accuracy(excel_stats)
            self.report.add_test_result("data_accuracy", accuracy_result['passed'], accuracy_result)
            
            for discrepancy in accuracy_result.get('discrepancies', []):
                self.report.add_discrepancy("DATA_ACCURACY", discrepancy, "HIGH")
            
            # Test 3: Business rule validation
            logger.info("Running business rule validation...")
            business_rule_result = self.db_validator.validate_business_rules()
            self.report.add_test_result("business_rules", business_rule_result['passed'], business_rule_result)
            
            for violation in business_rule_result.get('rule_violations', []):
                self.report.add_discrepancy("BUSINESS_RULES", violation, "MEDIUM")
            
            # Test 4: Performance validation
            logger.info("Running performance validation...")
            performance_result = self._validate_performance()
            self.report.add_test_result("performance", performance_result['passed'], performance_result)
            
            # Finalize report
            self.report.finalize_report()
            
            if self.report.overall_success:
                logger.info("✅ All validation tests passed!")
            else:
                logger.warning("❌ Some validation tests failed. Check report for details.")
            
            return self.report
            
        except Exception as e:
            logger.error(f"Validation failed with error: {e}")
            self.report.add_test_result("validation_execution", False, {"error": str(e)})
            self.report.finalize_report()
            return self.report
        
        finally:
            self.db_session.close()
    
    def _validate_performance(self) -> Dict[str, Any]:
        """Validate database performance metrics"""
        logger.info("Testing database performance...")
        
        result = {
            'passed': True,
            'query_times': {},
            'performance_issues': []
        }
        
        try:
            import time
            
            # Test 1: Asset count query
            start_time = time.time()
            asset_count = self.db_session.query(Asset).count()
            count_time = time.time() - start_time
            result['query_times']['asset_count'] = round(count_time, 3)
            
            if count_time > 2:  # Should take less than 2 seconds
                result['passed'] = False
                result['performance_issues'].append(f"Asset count query too slow: {count_time:.2f}s")
            
            # Test 2: Complex asset query with filters
            start_time = time.time()
            filtered_assets = self.db_session.query(Asset).filter(
                Asset.par_amount > 1000000,
                Asset.sp_rating.in_(['BBB+', 'BBB', 'BBB-'])
            ).limit(100).all()
            filter_time = time.time() - start_time
            result['query_times']['filtered_asset_query'] = round(filter_time, 3)
            
            if filter_time > 1:  # Should take less than 1 second
                result['passed'] = False
                result['performance_issues'].append(f"Filtered asset query too slow: {filter_time:.2f}s")
            
            # Test 3: Aggregation query
            start_time = time.time()
            total_par = self.db_session.query(func.sum(Asset.par_amount)).scalar()
            agg_time = time.time() - start_time
            result['query_times']['aggregation_query'] = round(agg_time, 3)
            
            if agg_time > 1:  # Should take less than 1 second
                result['passed'] = False
                result['performance_issues'].append(f"Aggregation query too slow: {agg_time:.2f}s")
            
            self.report.performance_metrics = result['query_times']
            
            return result
            
        except Exception as e:
            logger.error(f"Error validating performance: {e}")
            result['passed'] = False
            result['performance_issues'].append(f"Performance validation error: {e}")
            return result
    
    def save_report(self, output_path: Optional[Path] = None) -> Path:
        """Save validation report to file"""
        if not output_path:
            output_path = Path(f"validation_reports/validation_report_{self.report.validation_id}.json")
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w') as f:
            json.dump(self.report.to_dict(), f, indent=2, default=str)
        
        logger.info(f"Validation report saved to: {output_path}")
        return output_path

def main():
    """Main validation execution function"""
    try:
        logger.info("Starting CLO Data Migration Validation")
        
        # Initialize configuration
        config = ValidationConfig()
        
        # Create validator
        validator = MigrationValidator(config)
        
        # Execute validation
        report = validator.execute_full_validation()
        
        # Save report
        report_path = validator.save_report()
        
        # Print summary
        print("\n" + "="*70)
        print("CLO DATA MIGRATION VALIDATION SUMMARY")
        print("="*70)
        print(f"Validation ID: {report.validation_id}")
        print(f"Overall Result: {'✅ PASSED' if report.overall_success else '❌ FAILED'}")
        print(f"Total Tests: {len(report.test_results)}")
        print(f"Tests Passed: {sum(1 for r in report.test_results.values() if r['passed'])}")
        print(f"Tests Failed: {sum(1 for r in report.test_results.values() if not r['passed'])}")
        print(f"Total Discrepancies: {len(report.discrepancies)}")
        print(f"High Severity Issues: {len([d for d in report.discrepancies if d['severity'] == 'HIGH'])}")
        print(f"Report Location: {report_path}")
        
        if not report.overall_success:
            print("\n⚠️  VALIDATION FAILED - Review report for details")
            print("Recommendations:")
            for rec in report.recommendations:
                print(f"  • {rec}")
            return 1
        else:
            print("\n✅ VALIDATION PASSED - Migration data is ready for production")
            return 0
        
    except Exception as e:
        logger.error(f"Fatal validation error: {e}")
        print(f"\n❌ VALIDATION ERROR: {e}")
        return 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)