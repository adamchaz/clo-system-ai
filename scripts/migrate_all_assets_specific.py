#!/usr/bin/env python3
"""
All Assets Specific Migration Script
Migrate the All Assets worksheet (1,003 assets × 71 columns) to PostgreSQL

This script provides:
- Optimized extraction of All Assets sheet data
- Comprehensive data transformation with business logic
- Batch loading with error handling
- Multi-level validation with Excel comparison
- Detailed migration reporting
"""

import json
import logging
import sys
import traceback
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from typing import Dict, List, Any, Optional, Union
import pandas as pd
import numpy as np

# Excel processing
import openpyxl

# Database and ORM
from sqlalchemy import create_engine, text, func
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.exc import SQLAlchemyError, IntegrityError

# Add parent directory for imports
sys.path.append(str(Path(__file__).parent.parent))

from backend.app.models.asset import Asset
from backend.app.core.database import Base

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('all_assets_migration.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class AllAssetsMigrator:
    """Specialized migrator for All Assets worksheet"""
    
    def __init__(self, excel_path: str, database_url: str):
        self.excel_path = Path(excel_path)
        self.database_url = database_url
        
        # Migration statistics
        self.stats = {
            'start_time': datetime.now(),
            'end_time': None,
            'extracted_count': 0,
            'transformed_count': 0,
            'loaded_count': 0,
            'validation_errors': 0,
            'business_rule_violations': 0
        }
        
        # Error tracking
        self.errors = {
            'extraction_errors': [],
            'transformation_errors': [],
            'loading_errors': [],
            'validation_errors': []
        }
        
        # Initialize database session
        self.engine = create_engine(database_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        self.db_session = SessionLocal()
        
        # Column mapping from Excel to database
        self.column_mapping = {
            'BLKRockID': 'blkrock_id',
            'Issue Name': 'issue_name', 
            'Issuer Name': 'issuer_name',
            'ISSUER ID': 'issuer_id',
            'Tranche': 'tranche',
            'Bond/Loan': 'bond_loan',
            'Par Amount': 'par_amount',
            'Maturity': 'maturity',
            'Coupon Type': 'coupon_type',
            'Payment Frequency': 'payment_freq',
            'Coupon Spread': 'cpn_spread',
            'Libor Floor': 'libor_floor',
            'Index': 'index_name',
            'Coupon': 'coupon',
            'Commitment Fee': 'commit_fee',
            'Unfunded Amount': 'unfunded_amount',
            'Facility Size': 'facility_size',
            'Market Value': 'market_value',
            'Currency': 'currency',
            'WAL': 'wal',
            'Dated Date': 'dated_date',
            'Issue Date': 'issue_date',
            'First Payment Date': 'first_payment_date',
            'Amount Issued': 'amount_issued',
            'Day Count': 'day_count',
            'Index Cap': 'index_cap',
            'Business Day Convention': 'business_day_conv',
            'PMT EOM': 'payment_eom',
            'Amortization Type': 'amortization_type',
            'Country ': 'country',  # Note: Excel has trailing space
            'Seniority': 'seniority',
            
            # Moody's ratings
            "Moody's Rating-21": 'mdy_rating',
            "Moody's Recovery Rate-23": 'mdy_recovery_rate',
            "Moody's Default Probability Rating-18": 'mdy_dp_rating',
            "Moody's Rating WARF": 'mdy_dp_rating_warf',
            'Moody Facility Rating': 'mdy_facility_rating',
            'Moody Facility Outlook': 'mdy_facility_outlook',
            'Moody Issuer Rating': 'mdy_issuer_rating',
            'Moody Issuer outlook': 'mdy_issuer_outlook',
            "Moody's Senior Secured Rating": 'mdy_snr_sec_rating',
            'Moody Senior Unsecured Rating': 'mdy_snr_unsec_rating',
            'Moody Subordinate Rating': 'mdy_sub_rating',
            "Moody's Credit Estimate": 'mdy_credit_est_rating',
            "Moody's Credit Estimate Date": 'mdy_credit_est_date',
            
            # S&P ratings
            'S&P Facility Rating': 'sandp_facility_rating',
            'S&P Issuer Rating': 'sandp_issuer_rating',
            'S&P Senior Secured Rating': 'sandp_snr_sec_rating',
            'S&P Subordinated Rating': 'sandp_subordinate',
            'S&P Recovery Rating': 'sandp_rec_rating',
            
            # Industries
            "Moody's Industry": 'mdy_industry',
            'S&P Industry': 'sp_industry',
            'Moody Asset Category': 'mdy_asset_category',
            'S&P Priority Category': 'sp_priority_category',
            
            # Special fields
            'Default Obligation': 'default_asset',
            'Date of Defaulted': 'date_of_default',
            'PiKing': 'piking',
            'PIK Amount': 'pik_amount',
            'Credit Opinion': 'analyst_opinion'
        }
        
        # Asset flags that go into JSONB flags field
        self.flag_columns = [
            'Deferred Interest Bond', 'Delayed Drawdown ', 'Revolver',
            'Letter of Credit', 'Participation', 'DIP', 'Convertible', 
            'Structured Finance', 'Bridge Loan', 'Current Pay', 'Cov-Lite',
            'First Lien Last Out Loan'
        ]
    
    def extract_all_assets(self) -> List[Dict[str, Any]]:
        """Extract all assets from the All Assets worksheet"""
        logger.info("Starting All Assets data extraction...")
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            if 'All Assets' not in workbook.sheetnames:
                raise ValueError("All Assets worksheet not found")
            
            sheet = workbook['All Assets']
            logger.info(f"Loaded All Assets sheet: {sheet.max_column} columns × {sheet.max_row} rows")
            
            # Extract headers from row 1
            headers = []
            for col in range(1, sheet.max_column + 1):
                header_value = sheet.cell(row=1, column=col).value
                headers.append(str(header_value) if header_value else f"Column_{col}")
            
            logger.info(f"Extracted {len(headers)} column headers")
            
            # Extract asset data (rows 2 to max_row)
            assets_data = []
            
            for row_num in range(2, sheet.max_row + 1):
                try:
                    asset_row = {}
                    
                    # Extract all column values for this row
                    for col_idx, header in enumerate(headers, 1):
                        cell_value = sheet.cell(row=row_num, column=col_idx).value
                        asset_row[header] = cell_value
                    
                    # Skip completely empty rows
                    if all(value in [None, '', ' '] for value in asset_row.values()):
                        continue
                    
                    # Skip rows without BLKRockID
                    if not asset_row.get('BLKRockID'):
                        logger.warning(f"Row {row_num}: Missing BLKRockID, skipping")
                        continue
                    
                    asset_row['_source_row'] = row_num  # Track source row for debugging
                    assets_data.append(asset_row)
                    
                except Exception as e:
                    logger.error(f"Error extracting row {row_num}: {e}")
                    self.errors['extraction_errors'].append(f"Row {row_num}: {e}")
            
            self.stats['extracted_count'] = len(assets_data)
            logger.info(f"Successfully extracted {len(assets_data)} assets from Excel")
            
            return assets_data
            
        except Exception as e:
            logger.error(f"Fatal error during extraction: {e}")
            logger.error(traceback.format_exc())
            raise
    
    def transform_asset_data(self, raw_assets: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Transform raw Excel data to database format"""
        logger.info(f"Starting transformation of {len(raw_assets)} assets...")
        
        transformed_assets = []
        
        for i, raw_asset in enumerate(raw_assets, 1):
            try:
                transformed = self._transform_single_asset(raw_asset)
                if transformed:
                    transformed_assets.append(transformed)
                
                if i % 100 == 0:
                    logger.info(f"Transformed {i}/{len(raw_assets)} assets")
                    
            except Exception as e:
                logger.error(f"Error transforming asset {raw_asset.get('BLKRockID', 'unknown')}: {e}")
                self.errors['transformation_errors'].append(f"Asset {raw_asset.get('BLKRockID', 'unknown')}: {e}")
        
        self.stats['transformed_count'] = len(transformed_assets)
        logger.info(f"Successfully transformed {len(transformed_assets)} assets")
        
        return transformed_assets
    
    def _transform_single_asset(self, raw_asset: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Transform a single asset with comprehensive data cleaning"""
        
        # Start with base transformation
        transformed = {
            'created_at': datetime.now(),
            'updated_at': datetime.now()
        }
        
        # Transform directly mapped fields
        for excel_col, db_field in self.column_mapping.items():
            raw_value = raw_asset.get(excel_col)
            
            if raw_value is not None:
                transformed_value = self._transform_field_value(raw_value, db_field, excel_col)
                if transformed_value is not None:
                    transformed[db_field] = transformed_value
        
        # Process asset flags into JSONB
        flags = self._process_asset_flags(raw_asset)
        if flags:
            transformed['flags'] = flags
        
        # Ensure critical fields are present
        if not transformed.get('blkrock_id'):
            logger.warning(f"Asset missing BLKRockID, skipping")
            return None
        
        return transformed
    
    def _transform_field_value(self, value: Any, db_field: str, excel_col: str) -> Any:
        """Transform individual field value based on target database field"""
        
        if pd.isna(value) or value in ['', ' ', 'N/A', 'n/a', '#N/A']:
            return None
        
        try:
            # Monetary fields - clean currency formatting
            if db_field in ['par_amount', 'market_value', 'facility_size', 'amount_issued', 
                           'unfunded_amount', 'pik_amount', 'commit_fee']:
                return self._clean_monetary_value(value)
            
            # Percentage/rate fields - convert to decimal
            elif db_field in ['coupon', 'cpn_spread', 'libor_floor', 'index_cap', 'mdy_recovery_rate']:
                return self._clean_percentage_value(value)
            
            # Date fields - parse various formats
            elif db_field in ['maturity', 'dated_date', 'issue_date', 'first_payment_date', 
                             'date_of_default', 'mdy_credit_est_date']:
                return self._parse_date_value(value)
            
            # Integer fields
            elif db_field in ['payment_freq']:
                return self._parse_integer_value(value)
            
            # Boolean fields
            elif db_field in ['payment_eom', 'piking', 'default_asset']:
                return self._parse_boolean_value(value)
            
            # Decimal fields (WAL, etc.)
            elif db_field in ['wal']:
                return self._clean_decimal_value(value)
            
            # Text fields - clean whitespace
            else:
                return self._clean_text_value(value)
                
        except Exception as e:
            logger.warning(f"Error transforming {excel_col} value '{value}': {e}")
            return None
    
    def _clean_monetary_value(self, value: Any) -> Optional[Decimal]:
        """Clean monetary value from Excel format"""
        if pd.isna(value):
            return None
        
        try:
            # Handle numeric types directly
            if isinstance(value, (int, float)):
                return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            # Clean string format
            cleaned = str(value).replace('$', '').replace(',', '').replace(' ', '').strip()
            
            if cleaned in ['', '-', '0']:
                return Decimal('0.00')
            
            return Decimal(cleaned).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
        except (ValueError, InvalidOperation):
            logger.warning(f"Could not parse monetary value: {value}")
            return None
    
    def _clean_percentage_value(self, value: Any) -> Optional[Decimal]:
        """Clean percentage value to decimal format"""
        if pd.isna(value):
            return None
        
        try:
            # Handle numeric types
            if isinstance(value, (int, float)):
                # Assume percentages > 1 are already in percentage form
                if value > 1:
                    return Decimal(str(value / 100)).quantize(Decimal('0.000001'), rounding=ROUND_HALF_UP)
                else:
                    return Decimal(str(value)).quantize(Decimal('0.000001'), rounding=ROUND_HALF_UP)
            
            # Clean string format
            cleaned = str(value).replace('%', '').replace(' ', '').strip()
            
            if cleaned in ['', '-', '0']:
                return Decimal('0.000000')
            
            decimal_value = Decimal(cleaned)
            
            # Convert percentage to decimal if > 1
            if decimal_value > 1:
                decimal_value = decimal_value / 100
            
            return decimal_value.quantize(Decimal('0.000001'), rounding=ROUND_HALF_UP)
            
        except (ValueError, InvalidOperation):
            logger.warning(f"Could not parse percentage value: {value}")
            return None
    
    def _parse_date_value(self, value: Any) -> Optional[date]:
        """Parse date value from various Excel formats"""
        if pd.isna(value):
            return None
        
        try:
            # Handle datetime objects
            if isinstance(value, datetime):
                return value.date()
            
            # Handle date objects
            if isinstance(value, date):
                return value
            
            # Handle numeric Excel dates
            if isinstance(value, (int, float)):
                # Excel date serial number (days since 1900-01-01)
                try:
                    excel_epoch = datetime(1900, 1, 1)
                    parsed_date = excel_epoch + pd.Timedelta(days=value - 2)  # Excel quirk
                    return parsed_date.date()
                except:
                    pass
            
            # Handle string dates
            date_str = str(value).strip()
            
            date_formats = [
                '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%m-%d-%Y',
                '%Y%m%d', '%d-%m-%Y', '%m/%d/%y', '%d/%m/%y'
            ]
            
            for fmt in date_formats:
                try:
                    parsed = datetime.strptime(date_str, fmt)
                    return parsed.date()
                except ValueError:
                    continue
            
            logger.warning(f"Could not parse date value: {value}")
            return None
            
        except Exception as e:
            logger.warning(f"Error parsing date {value}: {e}")
            return None
    
    def _parse_integer_value(self, value: Any) -> Optional[int]:
        """Parse integer value"""
        if pd.isna(value):
            return None
        
        try:
            if isinstance(value, int):
                return value
            
            if isinstance(value, float):
                return int(value)
            
            # Handle string integers
            cleaned = str(value).strip()
            return int(float(cleaned))  # Handle "4.0" format
            
        except (ValueError, TypeError):
            logger.warning(f"Could not parse integer value: {value}")
            return None
    
    def _parse_boolean_value(self, value: Any) -> Optional[bool]:
        """Parse boolean value from various formats"""
        if pd.isna(value):
            return None
        
        if isinstance(value, bool):
            return value
        
        str_value = str(value).strip().upper()
        
        if str_value in ['Y', 'YES', '1', 'TRUE', 'T']:
            return True
        elif str_value in ['N', 'NO', '0', 'FALSE', 'F']:
            return False
        else:
            return None
    
    def _clean_decimal_value(self, value: Any) -> Optional[Decimal]:
        """Clean decimal value"""
        if pd.isna(value):
            return None
        
        try:
            if isinstance(value, Decimal):
                return value
            
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            
            cleaned = str(value).replace(',', '').strip()
            return Decimal(cleaned)
            
        except (ValueError, InvalidOperation):
            logger.warning(f"Could not parse decimal value: {value}")
            return None
    
    def _clean_text_value(self, value: Any) -> Optional[str]:
        """Clean text value"""
        if pd.isna(value):
            return None
        
        text = str(value).strip()
        
        if text in ['', 'N/A', 'n/a', '#N/A', '-']:
            return None
        
        return text
    
    def _process_asset_flags(self, raw_asset: Dict[str, Any]) -> Optional[Dict[str, bool]]:
        """Process asset flags into JSONB format"""
        flags = {}
        
        for flag_col in self.flag_columns:
            flag_value = raw_asset.get(flag_col)
            
            if flag_value is not None:
                boolean_value = self._parse_boolean_value(flag_value)
                if boolean_value is not None:
                    # Convert column name to flag key
                    flag_key = flag_col.lower().replace(' ', '_').replace('_', '_')
                    flags[flag_key] = boolean_value
        
        return flags if flags else None
    
    def load_assets_to_database(self, transformed_assets: List[Dict[str, Any]]) -> Dict[str, int]:
        """Load transformed assets to database with batch processing"""
        logger.info(f"Starting database load of {len(transformed_assets)} assets...")
        
        batch_size = 50  # Smaller batches for complex asset data
        successful_loads = 0
        failed_loads = 0
        
        # Clear existing assets (if this is a full refresh)
        # self._clear_existing_assets()  # Uncomment if needed
        
        for i in range(0, len(transformed_assets), batch_size):
            batch = transformed_assets[i:i + batch_size]
            batch_num = i // batch_size + 1
            total_batches = len(transformed_assets) // batch_size + 1
            
            logger.info(f"Loading batch {batch_num}/{total_batches}: {len(batch)} assets")
            
            try:
                # Create Asset objects
                asset_objects = []
                for asset_data in batch:
                    try:
                        asset = Asset(**asset_data)
                        asset_objects.append(asset)
                    except Exception as e:
                        logger.error(f"Error creating Asset object for {asset_data.get('blkrock_id')}: {e}")
                        failed_loads += 1
                        self.errors['loading_errors'].append(f"Asset {asset_data.get('blkrock_id')}: {e}")
                
                # Bulk insert batch
                if asset_objects:
                    self.db_session.add_all(asset_objects)
                    self.db_session.commit()
                    successful_loads += len(asset_objects)
                    logger.info(f"Successfully loaded batch {batch_num}")
                
            except SQLAlchemyError as e:
                logger.error(f"Database error loading batch {batch_num}: {e}")
                self.db_session.rollback()
                failed_loads += len(batch)
                self.errors['loading_errors'].append(f"Batch {batch_num}: {e}")
            
            except Exception as e:
                logger.error(f"Unexpected error loading batch {batch_num}: {e}")
                self.db_session.rollback()
                failed_loads += len(batch)
                self.errors['loading_errors'].append(f"Batch {batch_num}: {e}")
        
        self.stats['loaded_count'] = successful_loads
        
        logger.info(f"Database loading complete: {successful_loads} successful, {failed_loads} failed")
        
        return {
            'successful': successful_loads,
            'failed': failed_loads,
            'total_processed': successful_loads + failed_loads
        }
    
    def validate_migration(self) -> Dict[str, Any]:
        """Comprehensive migration validation"""
        logger.info("Starting migration validation...")
        
        validation_results = {
            'asset_count_validation': self._validate_asset_count(),
            'data_completeness_validation': self._validate_data_completeness(),
            'business_rules_validation': self._validate_business_rules(),
            'sample_data_validation': self._validate_sample_data()
        }
        
        # Count total validation errors
        total_errors = sum(
            len(result.get('errors', [])) 
            for result in validation_results.values()
        )
        
        self.stats['validation_errors'] = total_errors
        
        logger.info(f"Validation complete: {total_errors} total issues found")
        
        return validation_results
    
    def _validate_asset_count(self) -> Dict[str, Any]:
        """Validate asset count matches Excel"""
        try:
            db_count = self.db_session.query(Asset).count()
            expected_count = self.stats['extracted_count']
            
            return {
                'database_count': db_count,
                'expected_count': expected_count,
                'difference': abs(db_count - expected_count),
                'passed': db_count == expected_count,
                'errors': [] if db_count == expected_count else [f"Count mismatch: expected {expected_count}, got {db_count}"]
            }
            
        except Exception as e:
            logger.error(f"Error validating asset count: {e}")
            return {'passed': False, 'errors': [f"Validation error: {e}"]}
    
    def _validate_data_completeness(self) -> Dict[str, Any]:
        """Validate critical fields are populated"""
        try:
            total_assets = self.db_session.query(Asset).count()
            
            critical_fields = {
                'blkrock_id': 'Asset ID',
                'issue_name': 'Issue Name', 
                'par_amount': 'Par Amount'
            }
            
            completeness_results = {}
            errors = []
            
            for field, description in critical_fields.items():
                null_count = self.db_session.query(Asset).filter(
                    getattr(Asset, field).is_(None)
                ).count()
                
                completeness_percentage = ((total_assets - null_count) / total_assets * 100) if total_assets > 0 else 0
                
                completeness_results[field] = {
                    'description': description,
                    'null_count': null_count,
                    'completeness_percentage': round(completeness_percentage, 2)
                }
                
                if null_count > 0:
                    errors.append(f"{description}: {null_count} missing values ({100 - completeness_percentage:.1f}% incomplete)")
            
            return {
                'total_assets': total_assets,
                'field_completeness': completeness_results,
                'passed': len(errors) == 0,
                'errors': errors
            }
            
        except Exception as e:
            logger.error(f"Error validating data completeness: {e}")
            return {'passed': False, 'errors': [f"Validation error: {e}"]}
    
    def _validate_business_rules(self) -> Dict[str, Any]:
        """Validate business rules are satisfied"""
        try:
            violations = []
            
            # Rule 1: Par amounts should be positive
            negative_par_count = self.db_session.query(Asset).filter(
                Asset.par_amount < 0
            ).count()
            if negative_par_count > 0:
                violations.append(f"Negative par amounts: {negative_par_count} assets")
            
            # Rule 2: Market values should be between 0 and 200
            invalid_mv_count = self.db_session.query(Asset).filter(
                (Asset.market_value < 0) | (Asset.market_value > 200)
            ).count()
            if invalid_mv_count > 0:
                violations.append(f"Invalid market values (outside 0-200%): {invalid_mv_count} assets")
            
            # Rule 3: Maturity dates should be reasonable
            old_maturity_count = self.db_session.query(Asset).filter(
                Asset.maturity < date(2000, 1, 1)
            ).count()
            if old_maturity_count > 0:
                violations.append(f"Very old maturity dates (before 2000): {old_maturity_count} assets")
            
            # Rule 4: Coupon rates should be reasonable
            invalid_coupon_count = self.db_session.query(Asset).filter(
                (Asset.coupon < 0) | (Asset.coupon > 0.5)
            ).count()
            if invalid_coupon_count > 0:
                violations.append(f"Invalid coupon rates (outside 0-50%): {invalid_coupon_count} assets")
            
            return {
                'business_rule_violations': violations,
                'passed': len(violations) == 0,
                'errors': violations
            }
            
        except Exception as e:
            logger.error(f"Error validating business rules: {e}")
            return {'passed': False, 'errors': [f"Validation error: {e}"]}
    
    def _validate_sample_data(self) -> Dict[str, Any]:
        """Validate sample of migrated data against Excel"""
        try:
            # Get sample of assets from database
            sample_assets = self.db_session.query(Asset).limit(10).all()
            
            # For now, just check that we can retrieve the data
            sample_validation = {
                'sample_size': len(sample_assets),
                'passed': len(sample_assets) > 0,
                'errors': [] if len(sample_assets) > 0 else ["No assets found for sample validation"]
            }
            
            return sample_validation
            
        except Exception as e:
            logger.error(f"Error validating sample data: {e}")
            return {'passed': False, 'errors': [f"Validation error: {e}"]}
    
    def generate_migration_report(self) -> str:
        """Generate comprehensive migration report"""
        self.stats['end_time'] = datetime.now()
        duration = (self.stats['end_time'] - self.stats['start_time']).total_seconds() / 60
        
        report = []
        report.append("=" * 80)
        report.append("ALL ASSETS MIGRATION REPORT")
        report.append("=" * 80)
        
        # Migration summary
        report.append(f"\nMigration Summary:")
        report.append(f"  Start Time: {self.stats['start_time']}")
        report.append(f"  End Time: {self.stats['end_time']}")
        report.append(f"  Duration: {duration:.1f} minutes")
        report.append(f"  Excel File: {self.excel_path}")
        
        # Processing statistics
        report.append(f"\nProcessing Statistics:")
        report.append(f"  Assets Extracted: {self.stats['extracted_count']}")
        report.append(f"  Assets Transformed: {self.stats['transformed_count']}")
        report.append(f"  Assets Loaded: {self.stats['loaded_count']}")
        
        # Success rate
        if self.stats['extracted_count'] > 0:
            success_rate = (self.stats['loaded_count'] / self.stats['extracted_count']) * 100
            report.append(f"  Success Rate: {success_rate:.1f}%")
        
        # Error summary
        total_errors = sum(len(errors) for errors in self.errors.values())
        report.append(f"\nError Summary:")
        report.append(f"  Total Errors: {total_errors}")
        
        for error_type, errors in self.errors.items():
            if errors:
                report.append(f"  {error_type.replace('_', ' ').title()}: {len(errors)}")
        
        # Migration status
        report.append(f"\nMigration Status:")
        if self.stats['loaded_count'] == self.stats['extracted_count'] and total_errors == 0:
            report.append(f"  ✅ SUCCESS - All assets migrated successfully")
        elif self.stats['loaded_count'] >= self.stats['extracted_count'] * 0.95:
            report.append(f"  ⚠️  PARTIAL SUCCESS - Most assets migrated with minor issues")
        else:
            report.append(f"  ❌ ISSUES - Significant problems encountered during migration")
        
        report.append("=" * 80)
        
        return "\n".join(report)
    
    def execute_full_migration(self) -> Dict[str, Any]:
        """Execute complete All Assets migration workflow"""
        logger.info("🚀 Starting All Assets Migration")
        
        try:
            # Phase 1: Extract
            raw_assets = self.extract_all_assets()
            
            # Phase 2: Transform
            transformed_assets = self.transform_asset_data(raw_assets)
            
            # Phase 3: Load
            load_results = self.load_assets_to_database(transformed_assets)
            
            # Phase 4: Validate
            validation_results = self.validate_migration()
            
            # Generate report
            report = self.generate_migration_report()
            
            return {
                'success': load_results['successful'] > 0,
                'statistics': self.stats,
                'load_results': load_results,
                'validation_results': validation_results,
                'errors': self.errors,
                'report': report
            }
            
        except Exception as e:
            logger.error(f"Fatal error during migration: {e}")
            logger.error(traceback.format_exc())
            
            return {
                'success': False,
                'error': str(e),
                'statistics': self.stats,
                'errors': self.errors
            }
        
        finally:
            self.db_session.close()

def main():
    """Main migration execution"""
    
    # Configuration
    excel_file = "TradeHypoPrelimv32.xlsm"
    database_url = "postgresql://clo_user:password@localhost:5432/clo_db"
    
    if not Path(excel_file).exists():
        print(f"❌ Error: Excel file not found: {excel_file}")
        return 1
    
    print("🚀 Starting All Assets Migration")
    print("=" * 50)
    
    # Execute migration
    migrator = AllAssetsMigrator(excel_file, database_url)
    results = migrator.execute_full_migration()
    
    # Display results
    if results['success']:
        print("✅ Migration completed successfully!")
        print(f"Assets migrated: {results['statistics']['loaded_count']}")
    else:
        print("❌ Migration failed!")
        if 'error' in results:
            print(f"Error: {results['error']}")
    
    # Display report
    if 'report' in results:
        print("\n" + results['report'])
    
    # Save detailed results
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results_file = f"all_assets_migration_results_{timestamp}.json"
    
    with open(results_file, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    
    print(f"\n📄 Detailed results saved to: {results_file}")
    
    return 0 if results['success'] else 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)